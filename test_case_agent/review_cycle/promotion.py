from __future__ import annotations

import hashlib
import json
import os
import re
import shutil
import time
import uuid
import zipfile
from contextlib import contextmanager
from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from typing import Any, Callable, Iterator, Mapping, Sequence

from test_case_agent.review_cycle.production_tc_gate import (
    ProductionTcGateResult,
    validate_production_tc_draft,
)
from test_case_agent.review_cycle.prepared_package import (
    PACKAGE_VERSION,
    PreparedObligation,
    load_obligations,
    load_prepared_package,
)
from test_case_agent.review_cycle.dimension_bindings import (
    parse_reviewer_dimension_source_bindings,
)
from test_case_agent.review_cycle.runtime import StageRuntimeError
from test_case_agent.review_cycle.source_assertions import (
    SourceAssertionContractError,
    SourceAssertionManifest,
    normalize_exact_source_text,
    parse_embedded_source_assertion_contract,
)


PROMOTION_BASIS_SCHEMA_VERSION = 3
PROMOTION_BASIS_SEED_SCHEMA_VERSION = 2
REVIEW_RESULT_SCHEMA_VERSION = 2
REQUIRED_GATE_VALIDATOR = "prepared-quality-gate-bundle-v1"
TARGET_SLO_MS = 300_000
HARD_SLO_MS = 420_000
SHA256_RE = re.compile(r"[0-9a-f]{64}")
KNOWN_STATE_SCHEMAS = {
    "codex-exec-cycle-state-v1",
    "workflow-state-final-aliases-v1",
}
PROMOTION_SEED_AVAILABLE_INPUTS = {
    "final_findings",
    "final_traceability_matrix",
    "final_traceability_matrix_xlsx",
    "handoff_prompt",
    "promotion_contract",
    "promotion_readiness",
}
PROMOTION_SEED_MISSING_INPUTS = {
    "cycle-state-signed-off-replacement",
    "workflow-state-signed-off-replacement",
    "final-traceability-matrix",
    "final-traceability-matrix-xlsx",
    "final-writer-response-decision",
    "reviewer-to-ui-prep-handoff-prompt",
}
PROMOTION_SEED_INPUT_TO_MISSING_INPUT = {
    "final_traceability_matrix": "final-traceability-matrix",
    "final_traceability_matrix_xlsx": "final-traceability-matrix-xlsx",
    "handoff_prompt": "reviewer-to-ui-prep-handoff-prompt",
}
PROMOTION_READY_BLOCKING_REASON = (
    "promotion is disabled; review the promotion-ready candidate before a "
    "controlled production write"
)
RECOVERY_JOURNAL_SCHEMA_VERSION = 1
LOCK_SCHEMA_VERSION = 2


class PromotionBlocked(RuntimeError):
    """A fail-closed, user-actionable promotion refusal."""

    def __init__(self, code: str, message: str):
        self.code = code
        self.message = message
        super().__init__(f"blocked-input[{code}]: {message}")


@dataclass(frozen=True)
class ArtifactBinding:
    path: str
    sha256: str


@dataclass(frozen=True)
class GateBinding:
    path: str
    sha256: str
    validator: str


@dataclass(frozen=True)
class StateUpdate:
    role: str
    path: str
    before_sha256: str
    after_path: str
    after_sha256: str
    schema: str


@dataclass(frozen=True)
class Publication:
    canonical_path: str
    expected_prior_sha256: str | None


@dataclass(frozen=True)
class PromotionBasis:
    schema_version: int
    ft_root: str
    cycle_dir: str
    scope_slug: str
    candidate: ArtifactBinding
    writer: ArtifactBinding
    source_basis: ArtifactBinding
    obligation_set: ArtifactBinding
    prepared_package: ArtifactBinding
    reviewer: ArtifactBinding
    publication: Publication
    gate_reports: tuple[GateBinding, ...]
    state_updates: tuple[StateUpdate, ...]
    final_aliases: Mapping[str, ArtifactBinding | None]


@dataclass(frozen=True)
class PromotionBasisSeed:
    ft_root: str
    cycle_dir: str
    scope_slug: str
    candidate: ArtifactBinding
    writer: ArtifactBinding
    source_basis: ArtifactBinding
    obligation_set: ArtifactBinding
    prepared_package: ArtifactBinding
    reviewer: ArtifactBinding
    gate_reports: tuple[GateBinding, ...]
    canonical_path: str
    canonical_prior_sha256: str | None
    production_test_case_hashes: Mapping[str, str]
    available_builder_inputs: Mapping[str, ArtifactBinding | None]


@dataclass(frozen=True)
class PromotionResult:
    status: str
    transaction_id: str
    basis_path: Path
    canonical_path: Path
    receipt_path: Path | None
    metrics_path: Path | None
    byte_identical: bool
    production_gate: Mapping[str, Any]
    phase_timings_ms: Mapping[str, int]

    def as_dict(self) -> dict[str, Any]:
        return {
            "status": self.status,
            "transaction_id": self.transaction_id,
            "basis_path": str(self.basis_path),
            "canonical_path": str(self.canonical_path),
            "receipt_path": str(self.receipt_path) if self.receipt_path else None,
            "metrics_path": str(self.metrics_path) if self.metrics_path else None,
            "byte_identical": self.byte_identical,
            "production_gate": dict(self.production_gate),
            "phase_timings_ms": dict(self.phase_timings_ms),
        }


@dataclass(frozen=True)
class PromotionBasisBuildResult:
    status: str
    seed_path: Path
    basis_path: Path
    basis_sha256: str
    reused: bool
    build_duration_ms: int

    def as_dict(self) -> dict[str, Any]:
        return {
            "status": self.status,
            "seed_path": str(self.seed_path),
            "basis_path": str(self.basis_path),
            "basis_sha256": self.basis_sha256,
            "reused": self.reused,
            "build_duration_ms": self.build_duration_ms,
        }


def build_normalized_review_result(
    *,
    reviewed_draft_path: str,
    reviewed_draft_sha256: str,
    reviewed_source_basis_sha256: str,
    reviewed_obligation_set_sha256: str,
    obligation_reviews: Sequence[Mapping[str, Any]],
    dimension_reviews: Sequence[Mapping[str, Any]],
    findings: Sequence[Mapping[str, Any]],
    summary: str,
) -> dict[str, Any]:
    """Build the immutable runner-owned projection of an accepted v4 review."""

    return {
        "schema_version": REVIEW_RESULT_SCHEMA_VERSION,
        "contract_version": 4,
        "decision": "accepted",
        "reviewed_draft_path": reviewed_draft_path,
        "reviewed_draft_sha256": reviewed_draft_sha256,
        "reviewed_source_basis_sha256": reviewed_source_basis_sha256,
        "reviewed_obligation_set_sha256": reviewed_obligation_set_sha256,
        "obligation_reviews": [dict(item) for item in obligation_reviews],
        "dimension_reviews": [dict(item) for item in dimension_reviews],
        "findings": [dict(item) for item in findings],
        "summary": summary,
    }


def build_promotion_basis_seed(
    *,
    ft_root: str,
    cycle_dir: str,
    scope_slug: str,
    candidate: Mapping[str, str],
    writer: Mapping[str, str],
    source_basis: Mapping[str, str],
    obligation_set: Mapping[str, str],
    reviewer: Mapping[str, str],
    prepared_package: Mapping[str, str],
    gate_reports: Sequence[Mapping[str, str]],
    canonical_path: str,
    canonical_prior_sha256: str | None,
    production_test_case_hashes: Mapping[str, str],
    available_builder_inputs: Mapping[str, Mapping[str, str] | None],
) -> dict[str, Any]:
    """Build a non-promotable seed without fabricating terminal handoff data."""

    normalized_available = {
        key: (dict(value) if value is not None else None)
        for key, value in sorted(available_builder_inputs.items())
    }
    missing_builder_inputs = set(PROMOTION_SEED_MISSING_INPUTS)
    for name, missing_name in PROMOTION_SEED_INPUT_TO_MISSING_INPUT.items():
        if normalized_available.get(name) is not None:
            missing_builder_inputs.discard(missing_name)

    return {
        "schema_version": PROMOTION_BASIS_SEED_SCHEMA_VERSION,
        "artifact_kind": "promotion-basis-seed",
        "status": "builder-input-required",
        "ft_root": ft_root,
        "cycle_dir": cycle_dir,
        "scope_slug": scope_slug,
        "candidate": dict(candidate),
        "writer": dict(writer),
        "source_basis": dict(source_basis),
        "obligation_set": dict(obligation_set),
        "reviewer": dict(reviewer),
        "prepared_package": dict(prepared_package),
        "gate_reports": [dict(item) for item in gate_reports],
        "production_baseline": {
            "captured_at": "cycle-start",
            "canonical_path": canonical_path,
            "canonical_prior_sha256": canonical_prior_sha256,
            "test_case_sha256": dict(sorted(production_test_case_hashes.items())),
        },
        "available_builder_inputs": normalized_available,
        "missing_builder_inputs": sorted(missing_builder_inputs),
    }


def _blocked(code: str, message: str) -> PromotionBlocked:
    return PromotionBlocked(code, message)


def _parse_json_bytes(content: bytes, *, path: Path, label: str) -> Mapping[str, Any]:
    try:
        payload = json.loads(content.decode("utf-8"))
    except (UnicodeError, json.JSONDecodeError) as exc:
        raise _blocked("PROMO-INVALID-JSON", f"cannot load {label} {path}: {exc}") from exc
    if not isinstance(payload, Mapping):
        raise _blocked("PROMO-INVALID-SCHEMA", f"{label} must be a JSON object")
    return payload


def _load_json(path: Path, *, missing_code: str, label: str) -> Mapping[str, Any]:
    if not path.is_file():
        raise _blocked(missing_code, f"missing normalized {label}: {path}")
    try:
        content = path.read_bytes()
    except OSError as exc:
        raise _blocked("PROMO-INVALID-JSON", f"cannot load {label} {path}: {exc}") from exc
    return _parse_json_bytes(content, path=path, label=label)


def _exact_fields(payload: Mapping[str, Any], expected: set[str], label: str) -> None:
    if set(payload) != expected:
        missing = sorted(expected - set(payload))
        unknown = sorted(set(payload) - expected)
        raise _blocked(
            "PROMO-INVALID-SCHEMA",
            f"{label} fields mismatch; missing={missing}, unknown={unknown}",
        )


def _text(value: Any, label: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise _blocked("PROMO-INVALID-SCHEMA", f"{label} must be non-empty text")
    return value.strip()


def _sha(value: Any, label: str) -> str:
    result = _text(value, label)
    if not SHA256_RE.fullmatch(result):
        raise _blocked("PROMO-INVALID-SCHEMA", f"{label} must be a lowercase SHA-256")
    return result


def _relative_path(value: Any, label: str) -> str:
    result = _text(value, label)
    parsed = PurePosixPath(result)
    if (
        parsed.is_absolute()
        or parsed.as_posix() != result
        or any(part in {"", ".", ".."} for part in parsed.parts)
    ):
        raise _blocked("PROMO-UNSAFE-PATH", f"{label} must be a normalized repository-relative path")
    return result


def _artifact(payload: Any, label: str) -> ArtifactBinding:
    if not isinstance(payload, Mapping):
        raise _blocked("PROMO-INVALID-SCHEMA", f"{label} must be an object")
    _exact_fields(payload, {"path", "sha256"}, label)
    return ArtifactBinding(
        path=_relative_path(payload["path"], f"{label}.path"),
        sha256=_sha(payload["sha256"], f"{label}.sha256"),
    )


def _gate(payload: Any, index: int) -> GateBinding:
    label = f"promotion-basis.gate_reports[{index}]"
    if not isinstance(payload, Mapping):
        raise _blocked("PROMO-INVALID-SCHEMA", f"{label} must be an object")
    _exact_fields(payload, {"path", "sha256", "validator"}, label)
    return GateBinding(
        path=_relative_path(payload["path"], f"{label}.path"),
        sha256=_sha(payload["sha256"], f"{label}.sha256"),
        validator=_text(payload["validator"], f"{label}.validator"),
    )


def _load_seed(path: Path) -> PromotionBasisSeed:
    payload = _load_json(
        path,
        missing_code="PROMO-MISSING-SEED",
        label="promotion-basis.seed.json",
    )
    _exact_fields(
        payload,
        {
            "schema_version",
            "artifact_kind",
            "status",
            "ft_root",
            "cycle_dir",
            "scope_slug",
            "candidate",
            "writer",
            "source_basis",
            "obligation_set",
            "prepared_package",
            "reviewer",
            "gate_reports",
            "production_baseline",
            "available_builder_inputs",
            "missing_builder_inputs",
        },
        "promotion-basis seed",
    )
    if payload["schema_version"] != PROMOTION_BASIS_SEED_SCHEMA_VERSION:
        raise _blocked(
            "PROMO-INVALID-SEED",
            "promotion seed schema_version must equal "
            f"{PROMOTION_BASIS_SEED_SCHEMA_VERSION}",
        )
    if (
        payload["artifact_kind"] != "promotion-basis-seed"
        or payload["status"] != "builder-input-required"
    ):
        raise _blocked(
            "PROMO-INVALID-SEED",
            "promotion seed must be the runner-owned builder-input-required artifact",
        )
    raw_missing = payload["missing_builder_inputs"]
    if not isinstance(raw_missing, list) or len(raw_missing) != len(set(raw_missing)):
        raise _blocked(
            "PROMO-INVALID-SEED",
            "promotion seed missing_builder_inputs must be a duplicate-free array",
        )
    raw_gates = payload["gate_reports"]
    if not isinstance(raw_gates, list) or not raw_gates:
        raise _blocked("PROMO-INVALID-SEED", "promotion seed requires gate_reports")
    gates = tuple(_gate(item, index) for index, item in enumerate(raw_gates))
    if len({item.path for item in gates}) != len(gates):
        raise _blocked("PROMO-INVALID-SEED", "promotion seed gate paths must be unique")
    if REQUIRED_GATE_VALIDATOR not in {item.validator for item in gates}:
        raise _blocked(
            "PROMO-INVALID-SEED",
            f"promotion seed must include {REQUIRED_GATE_VALIDATOR}",
        )

    baseline = payload["production_baseline"]
    if not isinstance(baseline, Mapping):
        raise _blocked("PROMO-INVALID-SEED", "production_baseline must be an object")
    _exact_fields(
        baseline,
        {
            "captured_at",
            "canonical_path",
            "canonical_prior_sha256",
            "test_case_sha256",
        },
        "promotion seed production_baseline",
    )
    if baseline["captured_at"] != "cycle-start":
        raise _blocked(
            "PROMO-INVALID-SEED",
            "production_baseline.captured_at must equal cycle-start",
        )
    canonical_prior = baseline["canonical_prior_sha256"]
    if canonical_prior is not None:
        canonical_prior = _sha(
            canonical_prior,
            "production_baseline.canonical_prior_sha256",
        )
    raw_hashes = baseline["test_case_sha256"]
    if not isinstance(raw_hashes, Mapping):
        raise _blocked(
            "PROMO-INVALID-SEED",
            "production_baseline.test_case_sha256 must be an object",
        )
    production_hashes: dict[str, str] = {}
    for raw_path, raw_sha in raw_hashes.items():
        relative = _relative_path(
            raw_path,
            "production_baseline.test_case_sha256 path",
        )
        production_hashes[relative] = _sha(
            raw_sha,
            f"production_baseline.test_case_sha256[{relative}]",
        )

    available = payload["available_builder_inputs"]
    if not isinstance(available, Mapping):
        raise _blocked("PROMO-INVALID-SEED", "available_builder_inputs must be an object")
    unknown_available = set(available) - PROMOTION_SEED_AVAILABLE_INPUTS
    if unknown_available:
        raise _blocked(
            "PROMO-INVALID-SEED",
            f"promotion seed contains unknown available builder inputs: {sorted(unknown_available)}",
        )
    normalized_available: dict[str, ArtifactBinding | None] = {}
    for name, value in available.items():
        normalized_available[str(name)] = (
            None if value is None else _artifact(value, f"available_builder_inputs.{name}")
        )
    expected_missing = set(PROMOTION_SEED_MISSING_INPUTS)
    for name, missing_name in PROMOTION_SEED_INPUT_TO_MISSING_INPUT.items():
        if normalized_available.get(name) is not None:
            expected_missing.discard(missing_name)
    if set(raw_missing) != expected_missing:
        raise _blocked(
            "PROMO-INVALID-SEED",
            "promotion seed missing_builder_inputs does not match its hash-bound "
            "available_builder_inputs",
        )

    candidate = _artifact(payload["candidate"], "promotion seed candidate")
    writer = _artifact(payload["writer"], "promotion seed writer")
    if candidate != writer:
        raise _blocked(
            "PROMO-INVALID-SEED",
            "promotion seed candidate and writer bindings must be byte-identical and path-identical",
        )
    return PromotionBasisSeed(
        ft_root=_relative_path(payload["ft_root"], "promotion seed ft_root"),
        cycle_dir=_relative_path(payload["cycle_dir"], "promotion seed cycle_dir"),
        scope_slug=_text(payload["scope_slug"], "promotion seed scope_slug"),
        candidate=candidate,
        writer=writer,
        source_basis=_artifact(payload["source_basis"], "promotion seed source_basis"),
        obligation_set=_artifact(
            payload["obligation_set"], "promotion seed obligation_set"
        ),
        prepared_package=_artifact(
            payload["prepared_package"], "promotion seed prepared_package"
        ),
        reviewer=_artifact(payload["reviewer"], "promotion seed reviewer"),
        gate_reports=gates,
        canonical_path=_relative_path(
            baseline["canonical_path"], "production_baseline.canonical_path"
        ),
        canonical_prior_sha256=canonical_prior,
        production_test_case_hashes=production_hashes,
        available_builder_inputs=normalized_available,
    )


def _state_update(payload: Any, index: int) -> StateUpdate:
    label = f"promotion-basis.state_updates[{index}]"
    if not isinstance(payload, Mapping):
        raise _blocked("PROMO-INVALID-SCHEMA", f"{label} must be an object")
    _exact_fields(
        payload,
        {"role", "path", "before_sha256", "after_path", "after_sha256", "schema"},
        label,
    )
    schema = _text(payload["schema"], f"{label}.schema")
    if schema not in KNOWN_STATE_SCHEMAS:
        raise _blocked(
            "PROMO-STATE-SCHEMA-UNSUPPORTED",
            f"{label}.schema is not supported: {schema}",
        )
    return StateUpdate(
        role=_text(payload["role"], f"{label}.role"),
        path=_relative_path(payload["path"], f"{label}.path"),
        before_sha256=_sha(payload["before_sha256"], f"{label}.before_sha256"),
        after_path=_relative_path(payload["after_path"], f"{label}.after_path"),
        after_sha256=_sha(payload["after_sha256"], f"{label}.after_sha256"),
        schema=schema,
    )


def _load_basis(path: Path, *, content: bytes | None = None) -> PromotionBasis:
    payload = (
        _load_json(
            path,
            missing_code="PROMO-MISSING-BASIS",
            label="promotion-basis.json",
        )
        if content is None
        else _parse_json_bytes(
            content,
            path=path,
            label="promotion-basis.json",
        )
    )
    _exact_fields(
        payload,
        {
            "schema_version",
            "ft_root",
            "cycle_dir",
            "scope_slug",
            "candidate",
            "writer",
            "source_basis",
            "obligation_set",
            "prepared_package",
            "reviewer",
            "publication",
            "gate_reports",
            "state_updates",
            "final_aliases",
        },
        "promotion-basis",
    )
    if payload["schema_version"] != PROMOTION_BASIS_SCHEMA_VERSION:
        raise _blocked(
            "PROMO-INVALID-SCHEMA",
            f"promotion-basis.schema_version must equal {PROMOTION_BASIS_SCHEMA_VERSION}",
        )
    publication = payload["publication"]
    if not isinstance(publication, Mapping):
        raise _blocked("PROMO-INVALID-SCHEMA", "promotion-basis.publication must be an object")
    _exact_fields(
        publication,
        {"canonical_path", "expected_prior_sha256"},
        "promotion-basis.publication",
    )
    expected_prior = publication["expected_prior_sha256"]
    if expected_prior is not None:
        expected_prior = _sha(expected_prior, "publication.expected_prior_sha256")

    raw_gates = payload["gate_reports"]
    if not isinstance(raw_gates, list) or not raw_gates:
        raise _blocked(
            "PROMO-INVALID-SCHEMA", "promotion-basis.gate_reports must be a non-empty array"
        )
    gates = tuple(_gate(item, index) for index, item in enumerate(raw_gates))
    if len({item.path for item in gates}) != len(gates):
        raise _blocked("PROMO-INVALID-SCHEMA", "promotion gate report paths must be unique")
    if REQUIRED_GATE_VALIDATOR not in {item.validator for item in gates}:
        raise _blocked(
            "PROMO-MISSING-GATE",
            f"promotion basis must include {REQUIRED_GATE_VALIDATOR}",
        )

    raw_updates = payload["state_updates"]
    if not isinstance(raw_updates, list):
        raise _blocked("PROMO-INVALID-SCHEMA", "promotion-basis.state_updates must be an array")
    updates = tuple(_state_update(item, index) for index, item in enumerate(raw_updates))
    expected_roles = {"cycle-state", "workflow-state"}
    actual_roles = {item.role for item in updates}
    if len(updates) != 2 or actual_roles != expected_roles:
        raise _blocked(
            "PROMO-MISSING-HANDOFF-SCHEMA",
            "promotion requires exactly one cycle-state and one workflow-state replacement",
        )

    aliases = payload["final_aliases"]
    if not isinstance(aliases, Mapping):
        raise _blocked("PROMO-INVALID-SCHEMA", "promotion-basis.final_aliases must be an object")
    alias_fields = {
        "final_findings",
        "final_traceability_matrix",
        "final_traceability_matrix_xlsx",
        "final_writer_response",
        "handoff_prompt",
    }
    _exact_fields(aliases, alias_fields, "promotion-basis.final_aliases")
    normalized_aliases: dict[str, ArtifactBinding | None] = {}
    for name in sorted(alias_fields):
        value = aliases[name]
        if name == "final_writer_response" and value is None:
            normalized_aliases[name] = None
        else:
            normalized_aliases[name] = _artifact(value, f"final_aliases.{name}")

    return PromotionBasis(
        schema_version=payload["schema_version"],
        ft_root=_relative_path(payload["ft_root"], "promotion-basis.ft_root"),
        cycle_dir=_relative_path(payload["cycle_dir"], "promotion-basis.cycle_dir"),
        scope_slug=_text(payload["scope_slug"], "promotion-basis.scope_slug"),
        candidate=_artifact(payload["candidate"], "promotion-basis.candidate"),
        writer=_artifact(payload["writer"], "promotion-basis.writer"),
        source_basis=_artifact(payload["source_basis"], "promotion-basis.source_basis"),
        obligation_set=_artifact(payload["obligation_set"], "promotion-basis.obligation_set"),
        prepared_package=_artifact(
            payload["prepared_package"], "promotion-basis.prepared_package"
        ),
        reviewer=_artifact(payload["reviewer"], "promotion-basis.reviewer"),
        publication=Publication(
            canonical_path=_relative_path(
                publication["canonical_path"], "publication.canonical_path"
            ),
            expected_prior_sha256=expected_prior,
        ),
        gate_reports=gates,
        state_updates=updates,
        final_aliases=normalized_aliases,
    )


def _read_basis_snapshot(path: Path) -> tuple[PromotionBasis, bytes, str]:
    """Parse and hash one immutable in-memory read of promotion-basis.json."""

    if not path.is_file():
        raise _blocked(
            "PROMO-MISSING-BASIS",
            f"missing normalized promotion-basis.json: {path}",
        )
    try:
        content = path.read_bytes()
    except OSError as exc:
        raise _blocked(
            "PROMO-INVALID-JSON",
            f"cannot load promotion-basis.json {path}: {exc}",
        ) from exc
    return _load_basis(path, content=content), content, _sha256_bytes(content)


def _resolve(repo_root: Path, relative: str, label: str) -> Path:
    path = (repo_root / Path(relative)).resolve()
    try:
        path.relative_to(repo_root.resolve())
    except ValueError as exc:
        raise _blocked("PROMO-UNSAFE-PATH", f"{label} escapes repository root: {relative}") from exc
    return path


def _relative(path: Path, root: Path) -> str:
    try:
        return path.resolve().relative_to(root.resolve()).as_posix()
    except ValueError as exc:
        raise _blocked("PROMO-UNSAFE-PATH", f"path escapes expected root: {path}") from exc


def _sha256_bytes(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def _sha256_file(path: Path) -> str:
    if not path.is_file():
        raise _blocked("PROMO-MISSING-ARTIFACT", f"required artifact is missing: {path}")
    return _sha256_bytes(path.read_bytes())


def _verify_binding(
    binding: ArtifactBinding,
    repo_root: Path,
    *,
    label: str,
    parent: Path | None = None,
    missing_code: str = "PROMO-MISSING-ARTIFACT",
) -> Path:
    path = _resolve(repo_root, binding.path, label)
    if parent is not None:
        try:
            path.relative_to(parent.resolve())
        except ValueError as exc:
            raise _blocked("PROMO-UNSAFE-PATH", f"{label} must be inside {parent}: {path}") from exc
    if not path.is_file():
        raise _blocked(missing_code, f"required {label} is missing: {path}")
    actual = _sha256_file(path)
    if actual != binding.sha256:
        raise _blocked(
            "PROMO-HASH-MISMATCH",
            f"{label} SHA-256 mismatch: expected {binding.sha256}, got {actual}",
        )
    return path


def _parse_scalar(value: str) -> Any:
    value = value.strip()
    if not value:
        return []
    if value == "true":
        return True
    if value == "false":
        return False
    if value in {"null", "none", "~"}:
        return None
    if value == "[]":
        return []
    if value.startswith('"'):
        try:
            return json.loads(value)
        except json.JSONDecodeError as exc:
            raise _blocked(
                "PROMO-INVALID-YAML",
                f"invalid quoted YAML scalar: {value!r}",
            ) from exc
    if value.startswith("'") or value.endswith("'"):
        raise _blocked(
            "PROMO-INVALID-YAML",
            "single-quoted scalars are outside the supported workflow-state subset",
        )
    return value


def _top_level_yaml(text: str) -> dict[str, Any]:
    result: dict[str, Any] = {}
    current_block: str | None = None
    block_kinds: dict[str, str] = {}
    nested_keys: dict[str, set[str]] = {}
    for line_no, raw in enumerate(text.splitlines(), start=1):
        if not raw.strip() or raw.lstrip().startswith("#"):
            continue
        if "\t" in raw[: len(raw) - len(raw.lstrip())]:
            raise _blocked(
                "PROMO-INVALID-YAML",
                f"tab indentation is not allowed at line {line_no}",
            )
        match = re.fullmatch(r"([A-Za-z0-9_-]+):(?:\s*(.*))?", raw)
        if match:
            key = match.group(1)
            if key in result:
                raise _blocked(
                    "PROMO-INVALID-YAML",
                    f"duplicate top-level YAML key {key!r} at line {line_no}",
                )
            value = (match.group(2) or "").strip()
            result[key] = _parse_scalar(value)
            current_block = key if not value else None
            if current_block is not None:
                nested_keys[current_block] = set()
            continue
        if not raw.startswith(" "):
            raise _blocked(
                "PROMO-INVALID-YAML",
                f"unsupported top-level YAML syntax at line {line_no}",
            )
        if current_block is None:
            raise _blocked(
                "PROMO-INVALID-YAML",
                f"indented YAML content has no owning key at line {line_no}",
            )
        list_match = re.fullmatch(r"  -\s+(.+)", raw)
        if list_match:
            if block_kinds.get(current_block) == "mapping":
                raise _blocked(
                    "PROMO-INVALID-YAML",
                    f"mixed YAML mapping/list block for {current_block!r}",
                )
            block_kinds[current_block] = "list"
            result[current_block].append(_parse_scalar(list_match.group(1)))
            continue
        mapping_match = re.fullmatch(
            r"  ([A-Za-z0-9_-]+):(?:\s*(.*))?",
            raw,
        )
        if mapping_match:
            if block_kinds.get(current_block) == "list":
                raise _blocked(
                    "PROMO-INVALID-YAML",
                    f"mixed YAML mapping/list block for {current_block!r}",
                )
            if block_kinds.get(current_block) is None:
                result[current_block] = {}
            block_kinds[current_block] = "mapping"
            nested_key = mapping_match.group(1)
            if nested_key in nested_keys[current_block]:
                raise _blocked(
                    "PROMO-INVALID-YAML",
                    f"duplicate YAML mapping key {current_block}.{nested_key} at line {line_no}",
                )
            nested_keys[current_block].add(nested_key)
            result[current_block][nested_key] = _parse_scalar(  # type: ignore[index]
                (mapping_match.group(2) or "").strip()
            )
            continue
        raise _blocked(
            "PROMO-INVALID-YAML",
            f"unsupported indented YAML syntax at line {line_no}",
        )
    return result


def _yaml_mapping(text: str, heading: str) -> dict[str, Any]:
    lines = text.splitlines()
    starts = [index for index, line in enumerate(lines) if line == f"{heading}:"]
    if not starts:
        return {}
    if len(starts) != 1:
        raise _blocked(
            "PROMO-INVALID-YAML",
            f"YAML mapping {heading!r} must occur exactly once",
        )
    start = starts[0]
    result: dict[str, Any] = {}
    for line_no, line in enumerate(lines[start + 1 :], start=start + 2):
        if not line.strip() or line.lstrip().startswith("#"):
            continue
        if line and not line.startswith(" "):
            break
        match = re.fullmatch(r"  ([A-Za-z0-9_-]+):\s*(.*?)\s*", line)
        if not match:
            raise _blocked(
                "PROMO-INVALID-YAML",
                f"unsupported YAML mapping entry in {heading!r} at line {line_no}",
            )
        key = match.group(1)
        if key in result:
            raise _blocked(
                "PROMO-INVALID-YAML",
                f"duplicate YAML mapping key {heading}.{key} at line {line_no}",
            )
        result[key] = _parse_scalar(match.group(2))
    return result


def _yaml_list(text: str, heading: str) -> list[Any]:
    lines = text.splitlines()
    starts = [index for index, line in enumerate(lines) if line == f"{heading}:"]
    if not starts:
        return []
    if len(starts) != 1:
        raise _blocked(
            "PROMO-INVALID-YAML",
            f"YAML list {heading!r} must occur exactly once",
        )
    result: list[Any] = []
    for line_no, line in enumerate(lines[starts[0] + 1 :], start=starts[0] + 2):
        if not line.strip() or line.lstrip().startswith("#"):
            continue
        if not line.startswith(" "):
            break
        match = re.fullmatch(r"  -\s+(.+)", line)
        if not match:
            raise _blocked(
                "PROMO-INVALID-YAML",
                f"unsupported YAML list entry in {heading!r} at line {line_no}",
            )
        result.append(_parse_scalar(match.group(1)))
    return result


def _validate_cycle_state_before(text: str, candidate_sha: str) -> None:
    state = _top_level_yaml(text)
    if state.get("accepted_terminal_state") is not True:
        raise _blocked("PROMO-NOT-ACCEPTED", "cycle-state is not in an accepted terminal state")
    if state.get("reviewer_stage_status") != "accepted":
        raise _blocked("PROMO-NOT-ACCEPTED", "cycle-state reviewer_stage_status is not accepted")
    if state.get("final_promoted") is True:
        raise _blocked("PROMO-ALREADY-PROMOTED", "cycle-state already records final_promoted=true")
    if state.get("writer_draft_sha256") != candidate_sha:
        raise _blocked(
            "PROMO-HASH-MISMATCH",
            "cycle-state writer_draft_sha256 does not match the accepted candidate",
        )
    reviewer_rebind = (
        state.get("writer_stage_status") == "skipped-reviewer-rebind"
        and isinstance(state.get("reviewer_rebind_report"), str)
        and bool(state["reviewer_rebind_report"].strip())
    )
    ordinary_writer_review = state.get("writer_stage_status") == "completed"
    expected = (
        {
            "workflow_status": "accepted-not-promoted",
            "stage_status": "accepted-not-promoted",
            "current_stage": "reviewer-r1",
            "writer_stage_status": "skipped-reviewer-rebind",
            "reviewer_stage_status": "accepted",
            "accepted_terminal_state": True,
            "final_promoted": False,
            "draft_or_unsigned": True,
            "promotion_status": "pending",
        }
        if reviewer_rebind
        else {
            "workflow_status": "accepted-not-promoted",
            "stage_status": "accepted-not-promoted",
            "current_stage": "reviewer-r1",
            "writer_stage_status": "completed",
            "reviewer_stage_status": "accepted",
            "accepted_terminal_state": True,
            "final_promoted": False,
            "draft_or_unsigned": True,
            "promotion_status": "pending",
        }
        if ordinary_writer_review
        else {
            "workflow_status": "accepted-promotion-ready-not-promoted",
            "stage_status": "accepted-promotion-ready-not-promoted",
            "current_stage": "reviewer",
            "reviewer_stage_status": "accepted",
            "accepted_terminal_state": True,
            "final_promoted": False,
            "draft_or_unsigned": True,
            "promotion_status": "ready-not-promoted",
        }
    )
    mismatches = {
        key: {"expected": value, "actual": state.get(key)}
        for key, value in expected.items()
        if state.get(key) != value
    }
    if (
        mismatches
        or state.get("blocking_reasons") != [PROMOTION_READY_BLOCKING_REASON]
    ):
        raise _blocked(
            "PROMO-NOT-ACCEPTED",
            "cycle-state is not an exact accepted promotion-ready or reviewer-rebind terminal state: "
            f"mismatches={mismatches}",
        )


def _validate_cycle_state_after(
    text: str, *, candidate_sha: str, canonical_path: str
) -> None:
    state = _top_level_yaml(text)
    expected = {
        "workflow_status": "signed-off",
        "stage_status": "signed-off",
        "reviewer_stage_status": "accepted",
        "accepted_terminal_state": True,
        "final_promoted": True,
        "draft_or_unsigned": False,
        "promotion_status": "completed",
        "writer_draft_sha256": candidate_sha,
        "final_sha256": candidate_sha,
        "canonical_test_cases": canonical_path,
    }
    mismatches = {
        key: {"expected": value, "actual": state.get(key)}
        for key, value in expected.items()
        if state.get(key) != value
    }
    if mismatches or state.get("blocking_reasons") not in ([], None):
        raise _blocked(
            "PROMO-INVALID-CYCLE-STATE",
            f"signed-off cycle-state replacement is invalid: {mismatches}",
        )


def _repo_binding_relative(binding: ArtifactBinding, repo_root: Path, ft_root: Path) -> str:
    return _relative(_resolve(repo_root, binding.path, "final alias"), ft_root)


def _validate_workflow_state_after(
    text: str,
    *,
    basis: PromotionBasis,
    repo_root: Path,
    ft_root: Path,
    cycle_state_path: Path,
    signed_off_snapshot_rel: str,
) -> None:
    top = _top_level_yaml(text)
    if top.get("scope_slug") != basis.scope_slug:
        raise _blocked(
            "PROMO-INVALID-WORKFLOW-STATE",
            "workflow-state scope_slug does not match the bound promotion scope",
        )
    if top.get("current_stage") != "ft-test-case-iteration" or top.get("stage_status") != "signed-off":
        raise _blocked(
            "PROMO-INVALID-WORKFLOW-STATE",
            "workflow-state replacement must be a signed-off ft-test-case-iteration handoff",
        )
    if top.get("next_skill") != "ft-ui-automation-prep":
        raise _blocked(
            "PROMO-INVALID-WORKFLOW-STATE",
            "signed-off workflow-state must route to ft-ui-automation-prep",
        )
    if top.get("blocking_reasons") != []:
        raise _blocked(
            "PROMO-INVALID-WORKFLOW-STATE",
            "signed-off workflow-state must have an explicit empty blocking_reasons list",
        )
    latest = _yaml_mapping(text, "latest_artifacts")
    canonical_relative = _relative(
        _resolve(repo_root, basis.publication.canonical_path, "canonical path"), ft_root
    )
    expected: dict[str, Any] = {
        "canonical_test_cases": canonical_relative,
        "cycle_state": _relative(cycle_state_path, ft_root),
        "final_findings": _repo_binding_relative(
            basis.final_aliases["final_findings"], repo_root, ft_root  # type: ignore[arg-type]
        ),
        "final_traceability_matrix": _repo_binding_relative(
            basis.final_aliases["final_traceability_matrix"], repo_root, ft_root  # type: ignore[arg-type]
        ),
        "final_traceability_matrix_xlsx": _repo_binding_relative(
            basis.final_aliases["final_traceability_matrix_xlsx"], repo_root, ft_root  # type: ignore[arg-type]
        ),
        "final_writer_response": (
            None
            if basis.final_aliases["final_writer_response"] is None
            else _repo_binding_relative(
                basis.final_aliases["final_writer_response"], repo_root, ft_root  # type: ignore[arg-type]
            )
        ),
        "signed_off_snapshot": signed_off_snapshot_rel,
        "prompt_reviewer_to_ui_prep": _repo_binding_relative(
            basis.final_aliases["handoff_prompt"], repo_root, ft_root  # type: ignore[arg-type]
        ),
        "active_transition_prompt": _repo_binding_relative(
            basis.final_aliases["handoff_prompt"], repo_root, ft_root  # type: ignore[arg-type]
        ),
    }
    mismatches = {
        key: {"expected": value, "actual": latest.get(key)}
        for key, value in expected.items()
        if latest.get(key) != value
    }
    if mismatches:
        raise _blocked(
            "PROMO-INVALID-WORKFLOW-STATE",
            f"workflow-state final aliases are incomplete or stale: {mismatches}",
        )
    required_inputs = set(_yaml_list(text, "required_inputs"))
    required = {
        expected["cycle_state"],
        expected["signed_off_snapshot"],
        expected["prompt_reviewer_to_ui_prep"],
    }
    if not required.issubset(required_inputs):
        raise _blocked(
            "PROMO-INVALID-WORKFLOW-STATE",
            f"workflow-state required_inputs omit signed-off handoff artifacts: {sorted(required - required_inputs)}",
        )


def _validate_workflow_state_before(text: str, *, scope_slug: str) -> None:
    top = _top_level_yaml(text)
    if top.get("scope_slug") != scope_slug:
        raise _blocked(
            "PROMO-INVALID-WORKFLOW-STATE",
            "current workflow-state scope_slug does not match the promotion seed",
        )
    expected = {
        "current_stage": "ft-test-case-iteration",
        "stage_status": "ready-for-review",
        "next_skill": "ft-test-case-reviewer",
    }
    mismatches = {
        key: {"expected": value, "actual": top.get(key)}
        for key, value in expected.items()
        if top.get(key) != value
    }
    if top.get("stage_status") == "signed-off":
        raise _blocked(
            "PROMO-ALREADY-PROMOTED",
            "current workflow-state is already signed-off without a reusable committed transaction",
        )
    if mismatches:
        raise _blocked(
            "PROMO-INVALID-WORKFLOW-STATE",
            "current workflow-state is not the exact active reviewer handoff: "
            f"{mismatches}",
        )
    if top.get("blocking_reasons") != []:
        raise _blocked(
            "PROMO-INVALID-WORKFLOW-STATE",
            "current workflow-state must have an explicit empty blocking_reasons list",
        )


def _binding_payload(binding: ArtifactBinding) -> dict[str, str]:
    return {"path": binding.path, "sha256": binding.sha256}


def _gate_payload(binding: GateBinding) -> dict[str, str]:
    return {
        "path": binding.path,
        "sha256": binding.sha256,
        "validator": binding.validator,
    }


def _state_update_payload(update: StateUpdate) -> dict[str, str]:
    return {
        "role": update.role,
        "path": update.path,
        "before_sha256": update.before_sha256,
        "after_path": update.after_path,
        "after_sha256": update.after_sha256,
        "schema": update.schema,
    }


def _ft_relative_artifact_binding(
    value: Any,
    *,
    label: str,
    repo_root: Path,
    ft_root: Path,
    required_suffix: str,
    parent: Path,
) -> tuple[ArtifactBinding, Path]:
    relative = _relative_path(value, label)
    path = (ft_root / Path(relative)).resolve()
    try:
        path.relative_to(parent.resolve())
    except ValueError as exc:
        raise _blocked(
            "PROMO-UNSAFE-PATH",
            f"{label} must be inside {parent}: {relative}",
        ) from exc
    if path.suffix.lower() != required_suffix:
        raise _blocked(
            "PROMO-INVALID-BUILDER-INPUT",
            f"{label} must use {required_suffix}: {relative}",
        )
    if not path.is_file() or path.stat().st_size <= 0:
        raise _blocked(
            "PROMO-MISSING-BUILDER-INPUT",
            f"required non-empty {label} is missing: {path}",
        )
    binding = ArtifactBinding(
        path=_relative(path, repo_root),
        sha256=_sha256_file(path),
    )
    return binding, path


def _production_test_case_snapshot(ft_root: Path) -> dict[str, str]:
    production_root = ft_root / "test-cases"
    if not production_root.is_dir():
        return {}
    return {
        _relative(path, ft_root): _sha256_file(path)
        for path in sorted(production_root.rglob("*.md"))
        if path.is_file()
    }


def _canonical_basis_bytes(payload: Mapping[str, Any]) -> bytes:
    return (
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    ).encode("utf-8")


def _yaml_scalar_text(value: Any) -> str:
    if value is None:
        return "none"
    if value is True:
        return "true"
    if value is False:
        return "false"
    if isinstance(value, int) and not isinstance(value, bool):
        return str(value)
    if not isinstance(value, str):
        raise _blocked(
            "PROMO-INVALID-YAML",
            f"unsupported generated YAML scalar type: {type(value).__name__}",
        )
    if (
        value
        and value.strip() == value
        and re.fullmatch(r"[A-Za-z0-9_./@+-]+", value)
        and value.lower() not in {"true", "false", "null", "none"}
    ):
        return value
    return json.dumps(value, ensure_ascii=False)


def _render_yaml_block(key: str, value: Any) -> list[str]:
    if isinstance(value, Mapping):
        lines = [f"{key}:"]
        for child_key, child_value in value.items():
            if not re.fullmatch(r"[A-Za-z0-9_-]+", str(child_key)):
                raise _blocked(
                    "PROMO-INVALID-YAML",
                    f"invalid generated YAML mapping key: {child_key!r}",
                )
            lines.append(
                f"  {child_key}: {_yaml_scalar_text(child_value)}"
            )
        return lines
    if isinstance(value, (list, tuple)):
        if not value:
            return [f"{key}: []"]
        return [f"{key}:", *[f"  - {_yaml_scalar_text(item)}" for item in value]]
    return [f"{key}: {_yaml_scalar_text(value)}"]


def _replace_yaml_top_level(
    text: str,
    replacements: Mapping[str, Any],
) -> bytes:
    """Replace top-level YAML blocks while preserving unrelated state verbatim."""

    _top_level_yaml(text)
    lines = text.splitlines()
    starts: list[tuple[int, str]] = []
    for index, line in enumerate(lines):
        match = re.fullmatch(r"([A-Za-z0-9_-]+):(?:\s*(.*))?", line)
        if match:
            starts.append((index, match.group(1)))
    output: list[str] = []
    consumed: set[str] = set()
    cursor = 0
    for position, (start, key) in enumerate(starts):
        end = starts[position + 1][0] if position + 1 < len(starts) else len(lines)
        output.extend(lines[cursor:start])
        if key in replacements:
            output.extend(_render_yaml_block(key, replacements[key]))
            consumed.add(key)
        else:
            output.extend(lines[start:end])
        cursor = end
    output.extend(lines[cursor:])
    for key, value in replacements.items():
        if key not in consumed:
            output.extend(_render_yaml_block(key, value))
    return ("\n".join(output).rstrip() + "\n").encode("utf-8")


def _read_utf8(path: Path, *, label: str) -> tuple[bytes, str]:
    if not path.is_file():
        raise _blocked("PROMO-MISSING-STATE", f"required {label} is missing: {path}")
    content = path.read_bytes()
    try:
        return content, content.decode("utf-8")
    except UnicodeError as exc:
        raise _blocked("PROMO-INVALID-STATE", f"{label} must be UTF-8") from exc


def _discover_active_handoff(
    *,
    ft_root: Path,
    scope_slug: str,
) -> tuple[Path, Path, str]:
    handoff_root = ft_root / "work" / "stage-handoffs"
    candidates: list[tuple[Path, Path, str]] = []
    if handoff_root.is_dir():
        for prompt in sorted(handoff_root.rglob("prompt.reviewer-to-ui-prep.md")):
            workflow = prompt.parent / "workflow-state.yaml"
            if not workflow.is_file():
                continue
            _, workflow_text = _read_utf8(workflow, label="workflow-state")
            scope_lines = re.findall(
                r"(?m)^scope_slug:\s*(.*?)\s*$",
                workflow_text,
            )
            if len(scope_lines) != 1 or _parse_scalar(scope_lines[0]) != scope_slug:
                continue
            workflow_state = _top_level_yaml(workflow_text)
            if (
                workflow_state.get("stage_status") != "signed-off"
            ):
                candidates.append((prompt.resolve(), workflow.resolve(), workflow_text))
    if len(candidates) != 1:
        raise _blocked(
            "PROMO-MISSING-BUILDER-INPUT",
            "expected exactly one active reviewer-to-ui-prep handoff for scope "
            f"{scope_slug!r}, found {len(candidates)}",
        )
    return candidates[0]


def _normalize_traceability_cell(value: Any) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"\s+", " ", text.replace("\u00a0", " ")).strip()
    if text.startswith("`") and text.endswith("`"):
        leading = len(text) - len(text.lstrip("`"))
        trailing = len(text) - len(text.rstrip("`"))
        if leading == trailing and len(text) >= leading * 2:
            text = text[leading:-trailing].strip()
    return text


def _split_markdown_table_row(line: str) -> tuple[str, ...] | None:
    stripped = line.strip()
    if not (stripped.startswith("|") and stripped.endswith("|")):
        return None
    cells: list[str] = []
    current: list[str] = []
    escaped = False
    code_ticks = 0
    for char in stripped[1:-1]:
        if escaped:
            if char not in {"|", "\\"}:
                current.append("\\")
            current.append(char)
            escaped = False
            continue
        if char == "\\":
            escaped = True
            continue
        if char == "`":
            code_ticks = 0 if code_ticks else 1
            current.append(char)
            continue
        if char == "|" and not code_ticks:
            cells.append(
                _normalize_traceability_cell("".join(current)).replace("`", "")
            )
            current = []
            continue
        current.append(char)
    if escaped:
        current.append("\\")
    cells.append(_normalize_traceability_cell("".join(current)).replace("`", ""))
    return tuple(cells)


def _markdown_traceability_tables(path: Path) -> tuple[tuple[tuple[str, ...], ...], ...]:
    lines = path.read_text(encoding="utf-8").splitlines()
    tables: list[tuple[tuple[str, ...], ...]] = []
    index = 0
    while index + 1 < len(lines):
        header = _split_markdown_table_row(lines[index])
        separator = _split_markdown_table_row(lines[index + 1])
        if (
            header is None
            or separator is None
            or len(header) != len(separator)
            or not all(re.fullmatch(r":?-{3,}:?", cell) for cell in separator)
        ):
            index += 1
            continue
        rows: list[tuple[str, ...]] = [header]
        index += 2
        while index < len(lines):
            row = _split_markdown_table_row(lines[index])
            if row is None:
                break
            if len(row) != len(header):
                raise _blocked(
                    "PROMO-TRACEABILITY-PARITY-MISMATCH",
                    "traceability Markdown table contains a row with a different column count",
                )
            rows.append(row)
            index += 1
        tables.append(tuple(rows))
    return tuple(tables)


def _xlsx_traceability_table(path: Path) -> tuple[tuple[str, ...], ...]:
    try:
        with zipfile.ZipFile(path) as archive:
            names = set(archive.namelist())
            required_parts = {"[Content_Types].xml", "_rels/.rels", "xl/workbook.xml"}
            if not required_parts <= names or archive.testzip() is not None:
                raise ValueError("missing or corrupt OOXML parts")
    except (OSError, zipfile.BadZipFile, ValueError) as exc:
        raise _blocked(
            "PROMO-TRACEABILITY-XLSX-INVALID",
            "final traceability matrix XLSX is not a readable OOXML workbook",
        ) from exc
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(
            path,
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except Exception as exc:
        raise _blocked(
            "PROMO-TRACEABILITY-XLSX-INVALID",
            "final traceability matrix XLSX cannot be opened as an OOXML workbook",
        ) from exc
    try:
        named = [sheet for sheet in workbook.worksheets if sheet.title.casefold() == "traceability"]
        non_meta = [sheet for sheet in workbook.worksheets if sheet.title.casefold() != "meta"]
        candidates = named or non_meta
        if len(candidates) != 1:
            raise _blocked(
                "PROMO-TRACEABILITY-XLSX-INVALID",
                "XLSX must contain exactly one traceability data sheet and optional meta sheet",
            )
        sheet = candidates[0]
        if sheet.sheet_state != "visible":
            raise _blocked(
                "PROMO-TRACEABILITY-XLSX-INVALID",
                "traceability XLSX data sheet must be visible",
            )
        unexpected = [
            item.title
            for item in workbook.worksheets
            if item is not sheet and item.title.casefold() != "meta"
        ]
        if unexpected:
            raise _blocked(
                "PROMO-TRACEABILITY-XLSX-INVALID",
                "traceability XLSX contains unexpected data sheets: "
                + ", ".join(unexpected),
            )
        rows: list[list[str]] = []
        for workbook_row in sheet.iter_rows():
            values: list[str] = []
            for cell in workbook_row:
                if cell.data_type == "f":
                    raise _blocked(
                        "PROMO-TRACEABILITY-XLSX-INVALID",
                        "traceability XLSX must contain values, not formulas",
                    )
                values.append(_normalize_traceability_cell(cell.value))
            rows.append(values)
        while rows and not any(rows[-1]):
            rows.pop()
        if not rows:
            raise _blocked(
                "PROMO-TRACEABILITY-XLSX-INVALID",
                "traceability XLSX data sheet is empty",
            )
        last_column = max(
            (index for row in rows for index, value in enumerate(row) if value),
            default=-1,
        )
        if last_column < 0:
            raise _blocked(
                "PROMO-TRACEABILITY-XLSX-INVALID",
                "traceability XLSX data sheet is empty",
            )
        normalized_rows = tuple(tuple(row[: last_column + 1]) for row in rows)
        header = normalized_rows[0]
        if not header or any(not value for value in header) or len(header) != len(set(header)):
            raise _blocked(
                "PROMO-TRACEABILITY-XLSX-INVALID",
                "traceability XLSX requires non-empty unique column names",
            )
        return normalized_rows
    finally:
        workbook.close()


def _validate_traceability_matrix_pair(markdown: Path, xlsx: Path) -> None:
    workbook_table = _xlsx_traceability_table(xlsx)
    markdown_tables = _markdown_traceability_tables(markdown)
    matching_headers = [table for table in markdown_tables if table[0] == workbook_table[0]]
    if len(matching_headers) != 1:
        raise _blocked(
            "PROMO-TRACEABILITY-PARITY-MISMATCH",
            "traceability Markdown must contain exactly one table with the XLSX columns",
        )
    if matching_headers[0] != workbook_table:
        raise _blocked(
            "PROMO-TRACEABILITY-PARITY-MISMATCH",
            "traceability Markdown and XLSX must contain identical columns, rows, order, and values",
        )


def validate_traceability_matrix_pair(markdown: Path, xlsx: Path) -> None:
    """Validate the canonical Markdown/XLSX pair without discovering workflow aliases."""

    _validate_traceability_matrix_pair(markdown.resolve(), xlsx.resolve())


def _validate_seed_terminal_binding(
    *,
    seed: PromotionBasisSeed,
    name: str,
    path: Path,
    repo_root: Path,
) -> None:
    expected = seed.available_builder_inputs.get(name)
    if expected is None:
        raise _blocked(
            "PROMO-MISSING-BUILDER-INPUT",
            f"runner seed does not bind required terminal artifact {name}",
        )
    actual = ArtifactBinding(
        path=_relative(path.resolve(), repo_root.resolve()),
        sha256=_sha256_file(path),
    )
    if actual != expected:
        raise _blocked(
            "PROMO-BUILDER-BINDING-MISMATCH",
            f"discovered {name} is not the exact artifact bound by the runner seed",
        )


def _discover_traceability_matrix(
    *,
    repo_root: Path,
    ft_root: Path,
    cycle_dir: Path,
    workflow_text: str,
) -> tuple[Path, Path]:
    latest = _yaml_mapping(workflow_text, "latest_artifacts")
    declared_md = latest.get("final_traceability_matrix")
    declared_xlsx = latest.get("final_traceability_matrix_xlsx")
    if declared_md is not None or declared_xlsx is not None:
        if declared_md is None or declared_xlsx is None:
            raise _blocked(
                "PROMO-MISSING-BUILDER-INPUT",
                "active workflow-state must declare both traceability matrix aliases",
            )
        md = (ft_root / Path(_relative_path(declared_md, "final traceability matrix"))).resolve()
        xlsx = (
            ft_root
            / Path(_relative_path(declared_xlsx, "final traceability matrix XLSX"))
        ).resolve()
        try:
            md.relative_to(cycle_dir)
            xlsx.relative_to(cycle_dir)
        except ValueError as exc:
            raise _blocked(
                "PROMO-UNSAFE-PATH",
                "declared traceability matrix aliases must be inside cycle_dir",
            ) from exc
        candidates = [(md, xlsx)]
    else:
        outputs = cycle_dir / "outputs"
        candidates = [
            (md.resolve(), md.with_suffix(".xlsx").resolve())
            for md in sorted(outputs.glob("*traceability-matrix.md"))
            if md.is_file() and md.with_suffix(".xlsx").is_file()
        ]
    valid = [
        (md, xlsx)
        for md, xlsx in candidates
        if md.is_file()
        and md.stat().st_size > 0
        and xlsx.is_file()
        and xlsx.stat().st_size > 0
        and xlsx.read_bytes().startswith(b"PK")
    ]
    if len(valid) != 1:
        raise _blocked(
            "PROMO-MISSING-BUILDER-INPUT",
            "expected exactly one non-empty same-basename Markdown/XLSX traceability "
            f"matrix pair, found {len(valid)}",
        )
    md, xlsx = valid[0]
    if md.parent != xlsx.parent or md.stem != xlsx.stem:
        raise _blocked(
            "PROMO-INVALID-BUILDER-INPUT",
            "traceability matrix Markdown/XLSX must be same-basename siblings",
        )
    _validate_traceability_matrix_pair(md, xlsx)
    del repo_root
    return md, xlsx


def _write_generated_replacement_pair(
    *,
    cycle_path: Path,
    cycle_bytes: bytes,
    workflow_path: Path,
    workflow_bytes: bytes,
) -> None:
    desired = ((cycle_path, cycle_bytes), (workflow_path, workflow_bytes))
    for path, content in desired:
        if path.is_file() and path.read_bytes() != content:
            raise _blocked(
                "PROMO-BASIS-CONFLICT",
                f"existing generated state replacement differs from accepted state: {path}",
            )
    if (
        cycle_path.parent == workflow_path.parent
        and not cycle_path.parent.exists()
    ):
        parent = cycle_path.parent
        staged = parent.with_name(f".{parent.name}.{os.getpid()}.promotion-tmp")
        if staged.exists():
            raise _blocked(
                "PROMO-TEMP-CONFLICT",
                f"promotion replacement staging directory exists: {staged}",
            )
        try:
            staged.mkdir(parents=True)
            _atomic_write(staged / cycle_path.name, cycle_bytes)
            _atomic_write(staged / workflow_path.name, workflow_bytes)
            os.replace(staged, parent)
        finally:
            shutil.rmtree(staged, ignore_errors=True)
        return
    created: list[Path] = []
    try:
        for path, content in desired:
            if not path.exists():
                _atomic_write(path, content)
                created.append(path)
    except Exception:
        for path in created:
            path.unlink(missing_ok=True)
        raise


def _generate_signed_off_state_replacements(
    *,
    repo_root: Path,
    ft_root: Path,
    cycle_dir: Path,
    seed: PromotionBasisSeed,
    cycle_state_replacement_path: Path,
    workflow_state_replacement_path: Path,
) -> None:
    """Project accepted terminal state into deterministic signed-off replacements.

    Semantic findings/matrix/handoff content is never created here; the generator
    only discovers, validates and aliases artifacts already accepted by the cycle.
    """

    cycle_state_path = cycle_dir / "cycle-state.yaml"
    _, cycle_text = _read_utf8(cycle_state_path, label="cycle-state")
    _validate_cycle_state_before(cycle_text, seed.candidate.sha256)
    prompt, workflow_state_path, workflow_text = _discover_active_handoff(
        ft_root=ft_root,
        scope_slug=seed.scope_slug,
    )
    _validate_workflow_state_before(workflow_text, scope_slug=seed.scope_slug)
    matrix, matrix_xlsx = _discover_traceability_matrix(
        repo_root=repo_root,
        ft_root=ft_root,
        cycle_dir=cycle_dir,
        workflow_text=workflow_text,
    )
    _validate_seed_terminal_binding(
        seed=seed,
        name="final_traceability_matrix",
        path=matrix,
        repo_root=repo_root,
    )
    _validate_seed_terminal_binding(
        seed=seed,
        name="final_traceability_matrix_xlsx",
        path=matrix_xlsx,
        repo_root=repo_root,
    )
    _validate_seed_terminal_binding(
        seed=seed,
        name="handoff_prompt",
        path=prompt,
        repo_root=repo_root,
    )
    findings = seed.available_builder_inputs.get("final_findings")
    if findings is None:
        raise _blocked(
            "PROMO-MISSING-BUILDER-INPUT",
            "runner seed does not bind the accepted final findings artifact",
        )
    findings_path = _verify_binding(
        findings,
        repo_root,
        label="seed final findings",
        parent=cycle_dir,
    )
    latest_before = _yaml_mapping(workflow_text, "latest_artifacts")
    final_writer_value = latest_before.get("final_writer_response")
    final_writer_relative: str | None = None
    if final_writer_value is not None:
        final_writer_path = (
            ft_root
            / Path(_relative_path(final_writer_value, "final writer response"))
        ).resolve()
        try:
            final_writer_path.relative_to(cycle_dir)
        except ValueError as exc:
            raise _blocked(
                "PROMO-UNSAFE-PATH",
                "final writer response must be inside cycle_dir",
            ) from exc
        if not final_writer_path.is_file() or final_writer_path.stat().st_size <= 0:
            raise _blocked(
                "PROMO-MISSING-BUILDER-INPUT",
                f"declared final writer response is missing: {final_writer_path}",
            )
        final_writer_relative = _relative(final_writer_path, ft_root)

    canonical_ft_relative = _relative(
        _resolve(repo_root, seed.canonical_path, "canonical path"),
        ft_root,
    )
    cycle_relative = _relative(cycle_state_path, ft_root)
    snapshot_relative = _relative(
        cycle_dir / "versions" / "signed-off",
        ft_root,
    )
    prompt_relative = _relative(prompt, ft_root)
    latest_after = dict(latest_before)
    latest_after.update(
        {
            "canonical_test_cases": canonical_ft_relative,
            "cycle_state": cycle_relative,
            "final_findings": _relative(findings_path, ft_root),
            "final_traceability_matrix": _relative(matrix, ft_root),
            "final_traceability_matrix_xlsx": _relative(matrix_xlsx, ft_root),
            "final_writer_response": final_writer_relative,
            "signed_off_snapshot": snapshot_relative,
            "prompt_reviewer_to_ui_prep": prompt_relative,
            "active_transition_prompt": prompt_relative,
        }
    )
    required_inputs = _yaml_list(workflow_text, "required_inputs")
    if any(not isinstance(item, str) or not item for item in required_inputs):
        raise _blocked(
            "PROMO-INVALID-WORKFLOW-STATE",
            "workflow-state required_inputs must contain only non-empty paths",
        )
    for required in (cycle_relative, snapshot_relative, prompt_relative):
        if required not in required_inputs:
            required_inputs.append(required)

    cycle_after = _replace_yaml_top_level(
        cycle_text,
        {
            "workflow_status": "signed-off",
            "stage_status": "signed-off",
            "reviewer_stage_status": "accepted",
            "writer_draft_sha256": seed.candidate.sha256,
            "accepted_terminal_state": True,
            "final_promoted": True,
            "draft_or_unsigned": False,
            "promotion_status": "completed",
            "final_sha256": seed.candidate.sha256,
            "canonical_test_cases": canonical_ft_relative,
            "blocking_reasons": [],
        },
    )
    workflow_after = _replace_yaml_top_level(
        workflow_text,
        {
            "current_stage": "ft-test-case-iteration",
            "stage_status": "signed-off",
            "next_skill": "ft-ui-automation-prep",
            "canonical_test_cases": canonical_ft_relative,
            "required_inputs": required_inputs,
            "latest_artifacts": latest_after,
            "blocking_reasons": [],
        },
    )
    _write_generated_replacement_pair(
        cycle_path=cycle_state_replacement_path,
        cycle_bytes=cycle_after,
        workflow_path=workflow_state_replacement_path,
        workflow_bytes=workflow_after,
    )


def build_full_promotion_basis(
    *,
    repo_root: Path,
    cycle_dir: Path,
    seed_path: Path | None = None,
    basis_path: Path | None = None,
    cycle_state_replacement_path: Path | None = None,
    workflow_state_replacement_path: Path | None = None,
    monotonic_ns: Callable[[], int] = time.monotonic_ns,
) -> PromotionBasisBuildResult:
    """Build schema-v3 promotion basis only from real hash-bound terminal inputs.

    The builder binds real accepted findings, matrix and handoff artifacts. When
    explicit replacements are absent, it deterministically projects the accepted
    cycle/workflow states to signed-off form. It never creates findings, matrices,
    handoff prompts or reviewer-owned semantic content.
    """

    started = monotonic_ns()
    repo_root = repo_root.resolve()
    cycle_dir = cycle_dir.resolve()
    try:
        cycle_relative = cycle_dir.relative_to(repo_root).as_posix()
    except ValueError as exc:
        raise _blocked("PROMO-UNSAFE-PATH", "cycle_dir must be inside repo_root") from exc
    if not cycle_dir.is_dir():
        raise _blocked("PROMO-MISSING-CYCLE", f"cycle directory is missing: {cycle_dir}")
    explicit_cycle_replacement = cycle_state_replacement_path is not None
    explicit_workflow_replacement = workflow_state_replacement_path is not None
    seed_path = (seed_path or cycle_dir / "promotion-basis.seed.json").resolve()
    basis_path = (basis_path or cycle_dir / "promotion-basis.json").resolve()
    cycle_state_replacement_path = (
        cycle_state_replacement_path
        or cycle_dir / "promotion-inputs" / "cycle-state.signed-off.yaml"
    ).resolve()
    workflow_state_replacement_path = (
        workflow_state_replacement_path
        or cycle_dir / "promotion-inputs" / "workflow-state.signed-off.yaml"
    ).resolve()
    for label, path in (
        ("promotion seed", seed_path),
        ("promotion basis", basis_path),
        ("cycle-state replacement", cycle_state_replacement_path),
        ("workflow-state replacement", workflow_state_replacement_path),
    ):
        try:
            path.relative_to(cycle_dir)
        except ValueError as exc:
            raise _blocked(
                "PROMO-UNSAFE-PATH",
                f"{label} must be inside cycle_dir: {path}",
            ) from exc

    seed = _load_seed(seed_path)
    if seed.cycle_dir != cycle_relative:
        raise _blocked(
            "PROMO-CYCLE-MISMATCH",
            f"promotion seed cycle_dir={seed.cycle_dir} does not match {cycle_relative}",
        )
    ft_root = _resolve(repo_root, seed.ft_root, "promotion seed ft_root")
    if not ft_root.is_dir():
        raise _blocked("PROMO-MISSING-FT-ROOT", f"ft_root is missing: {ft_root}")
    try:
        cycle_dir.relative_to(ft_root)
    except ValueError as exc:
        raise _blocked(
            "PROMO-CYCLE-MISMATCH",
            "promotion seed cycle directory must be inside ft_root",
        ) from exc

    candidate_path = _verify_binding(
        seed.candidate, repo_root, label="seed candidate", parent=cycle_dir
    )
    reviewer_path = _verify_binding(
        seed.reviewer, repo_root, label="seed reviewer", parent=cycle_dir
    )
    source_path = _verify_binding(
        seed.source_basis, repo_root, label="seed source basis", parent=ft_root
    )
    obligation_path = _verify_binding(
        seed.obligation_set, repo_root, label="seed obligation set", parent=ft_root
    )
    prepared_package_path = _verify_binding(
        seed.prepared_package,
        repo_root,
        label="seed prepared package",
        parent=ft_root,
    )
    if seed.writer.sha256 != seed.candidate.sha256:
        raise _blocked(
            "PROMO-HASH-MISMATCH",
            "seed candidate is not byte-identical to the reviewed writer draft",
        )
    del candidate_path
    try:
        prepared_package = load_prepared_package(prepared_package_path, repo_root)
    except StageRuntimeError as exc:
        raise _blocked(
            "PROMO-PREPARED-PACKAGE-INVALID",
            f"prepared package is stale or invalid: {exc}",
        ) from exc
    release_status = prepared_package.release_status
    if prepared_package.package_version != PACKAGE_VERSION or release_status is None:
        raise _blocked(
            "PROMO-PREPARED-PACKAGE-INVALID",
            "promotion seed must bind the current structured prepared package",
        )
    if prepared_package.scope_slug != seed.scope_slug:
        raise _blocked(
            "PROMO-SCOPE-MISMATCH",
            "promotion seed package scope_slug does not match the accepted cycle",
        )
    if release_status.execution_dependency_registry:
        raise _blocked(
            "PROMO-BLOCKED-EXECUTION-DEPENDENCIES",
            "execution dependencies remain unresolved: "
            + ", ".join(release_status.blocking_gap_ids),
        )
    if release_status.blocking_gap_ids:
        raise _blocked(
            "PROMO-BLOCKING-SOURCE-GAPS",
            "blocking source gaps remain unresolved: "
            + ", ".join(release_status.blocking_gap_ids),
        )
    if (
        release_status.output_mode != "release"
        or not release_status.release_eligible
        or release_status.unsigned_status != "none"
    ):
        raise _blocked(
            "PROMO-PREPARED-PACKAGE-NOT-RELEASE-ELIGIBLE",
            "promotion seed package is not release-eligible for the bound scope",
        )
    package_artifacts = {
        item.kind: ArtifactBinding(item.path, item.sha256)
        for item in prepared_package.package_artifacts
    }
    if (
        package_artifacts.get("source-evidence") != seed.source_basis
        or package_artifacts.get("atomic-obligations") != seed.obligation_set
    ):
        raise _blocked(
            "PROMO-PACKAGE-BINDING-MISMATCH",
            "promotion seed source/obligation bindings do not match the prepared package",
        )
    _validate_review_result(
        reviewer_path,
        repo_root=repo_root,
        scope_slug=seed.scope_slug,
        writer=seed.writer,
        source_basis=seed.source_basis,
        source_basis_path=source_path,
        obligation_set=seed.obligation_set,
        obligation_path=obligation_path,
    )
    _validate_gate_reports(
        seed.gate_reports,
        repo_root,
        cycle_dir,
        seed.candidate.sha256,
    )

    canonical_path = _resolve(repo_root, seed.canonical_path, "seed canonical path")
    production_root = (ft_root / "test-cases").resolve()
    try:
        canonical_path.relative_to(production_root)
    except ValueError as exc:
        raise _blocked(
            "PROMO-PRODUCTION-BOUNDARY",
            f"canonical path must be inside {production_root}",
        ) from exc
    canonical_ft_relative = _relative(canonical_path, ft_root)
    baseline_canonical_sha = seed.production_test_case_hashes.get(canonical_ft_relative)
    if baseline_canonical_sha != seed.canonical_prior_sha256:
        raise _blocked(
            "PROMO-INVALID-SEED",
            "promotion seed canonical prior hash conflicts with its production snapshot",
        )

    existing_basis = _load_basis(basis_path) if basis_path.is_file() else None
    transaction_id = f"{seed.candidate.sha256[:16]}-{seed.reviewer.sha256[:8]}"
    completion_markers_exist = (
        (cycle_dir / "promotion" / transaction_id).exists()
        or (cycle_dir / "versions" / "signed-off").exists()
    )
    if existing_basis is None or not completion_markers_exist:
        current_production = _production_test_case_snapshot(ft_root)
        if current_production != dict(seed.production_test_case_hashes):
            missing = sorted(set(seed.production_test_case_hashes) - set(current_production))
            changed = sorted(
                path
                for path in set(seed.production_test_case_hashes) & set(current_production)
                if seed.production_test_case_hashes[path] != current_production[path]
            )
            unexpected = sorted(set(current_production) - set(seed.production_test_case_hashes))
            raise _blocked(
                "PROMO-PRODUCTION-DRIFT",
                "production test-case snapshot changed after cycle start: "
                f"missing={missing}, changed={changed}, unexpected={unexpected}",
            )

    replacements_missing = (
        not cycle_state_replacement_path.is_file()
        or not workflow_state_replacement_path.is_file()
    )
    if replacements_missing and (
        explicit_cycle_replacement or explicit_workflow_replacement
    ):
        raise _blocked(
            "PROMO-MISSING-BUILDER-INPUT",
            "explicit signed-off state replacement paths must both already exist",
        )
    if replacements_missing:
        _generate_signed_off_state_replacements(
            repo_root=repo_root,
            ft_root=ft_root,
            cycle_dir=cycle_dir,
            seed=seed,
            cycle_state_replacement_path=cycle_state_replacement_path,
            workflow_state_replacement_path=workflow_state_replacement_path,
        )
    if not cycle_state_replacement_path.is_file():
        raise _blocked(
            "PROMO-MISSING-BUILDER-INPUT",
            f"cycle-state signed-off replacement is missing: {cycle_state_replacement_path}",
        )
    if not workflow_state_replacement_path.is_file():
        raise _blocked(
            "PROMO-MISSING-BUILDER-INPUT",
            f"workflow-state signed-off replacement is missing: {workflow_state_replacement_path}",
        )
    try:
        cycle_after_bytes = cycle_state_replacement_path.read_bytes()
        workflow_after_bytes = workflow_state_replacement_path.read_bytes()
        cycle_after_text = cycle_after_bytes.decode("utf-8")
        workflow_after_text = workflow_after_bytes.decode("utf-8")
    except UnicodeError as exc:
        raise _blocked(
            "PROMO-INVALID-BUILDER-INPUT",
            "signed-off state replacements must be UTF-8",
        ) from exc
    latest = _yaml_mapping(workflow_after_text, "latest_artifacts")
    required_latest = {
        "final_findings",
        "final_traceability_matrix",
        "final_traceability_matrix_xlsx",
        "final_writer_response",
        "prompt_reviewer_to_ui_prep",
    }
    missing_latest = sorted(required_latest - set(latest))
    if missing_latest:
        raise _blocked(
            "PROMO-MISSING-BUILDER-INPUT",
            f"workflow-state replacement omits final aliases: {missing_latest}",
        )

    findings_binding, _ = _ft_relative_artifact_binding(
        latest["final_findings"],
        label="final findings",
        repo_root=repo_root,
        ft_root=ft_root,
        required_suffix=".md",
        parent=cycle_dir,
    )
    seed_findings = seed.available_builder_inputs.get("final_findings")
    if seed_findings is None or findings_binding != seed_findings:
        raise _blocked(
            "PROMO-BUILDER-BINDING-MISMATCH",
            "workflow-state final_findings is not the exact reviewer artifact bound by the runner seed",
        )
    matrix_binding, matrix_path = _ft_relative_artifact_binding(
        latest["final_traceability_matrix"],
        label="final traceability matrix",
        repo_root=repo_root,
        ft_root=ft_root,
        required_suffix=".md",
        parent=cycle_dir,
    )
    matrix_xlsx_binding, matrix_xlsx_path = _ft_relative_artifact_binding(
        latest["final_traceability_matrix_xlsx"],
        label="final traceability matrix XLSX",
        repo_root=repo_root,
        ft_root=ft_root,
        required_suffix=".xlsx",
        parent=cycle_dir,
    )
    if (
        matrix_path.parent != matrix_xlsx_path.parent
        or matrix_path.stem != matrix_xlsx_path.stem
    ):
        raise _blocked(
            "PROMO-INVALID-BUILDER-INPUT",
            "final traceability matrix Markdown/XLSX must be same-basename siblings",
        )
    _validate_traceability_matrix_pair(matrix_path, matrix_xlsx_path)
    _validate_seed_terminal_binding(
        seed=seed,
        name="final_traceability_matrix",
        path=matrix_path,
        repo_root=repo_root,
    )
    _validate_seed_terminal_binding(
        seed=seed,
        name="final_traceability_matrix_xlsx",
        path=matrix_xlsx_path,
        repo_root=repo_root,
    )

    handoff_root = (ft_root / "work" / "stage-handoffs").resolve()
    handoff_binding, handoff_path = _ft_relative_artifact_binding(
        latest["prompt_reviewer_to_ui_prep"],
        label="reviewer-to-ui-prep handoff prompt",
        repo_root=repo_root,
        ft_root=ft_root,
        required_suffix=".md",
        parent=handoff_root,
    )
    if handoff_path.name != "prompt.reviewer-to-ui-prep.md":
        raise _blocked(
            "PROMO-INVALID-BUILDER-INPUT",
            "handoff prompt must use the canonical prompt.reviewer-to-ui-prep.md name",
        )
    _validate_seed_terminal_binding(
        seed=seed,
        name="handoff_prompt",
        path=handoff_path,
        repo_root=repo_root,
    )
    workflow_state_path = handoff_path.parent / "workflow-state.yaml"
    if not workflow_state_path.is_file():
        raise _blocked(
            "PROMO-MISSING-STATE",
            f"active workflow-state is missing beside the bound handoff prompt: {workflow_state_path}",
        )
    cycle_state_path = cycle_dir / "cycle-state.yaml"
    if not cycle_state_path.is_file():
        raise _blocked(
            "PROMO-MISSING-STATE",
            f"accepted cycle-state is missing: {cycle_state_path}",
        )

    final_writer_response: ArtifactBinding | None
    if latest["final_writer_response"] is None:
        final_writer_response = None
    else:
        final_writer_response, _ = _ft_relative_artifact_binding(
            latest["final_writer_response"],
            label="final writer response",
            repo_root=repo_root,
            ft_root=ft_root,
            required_suffix=".md",
            parent=cycle_dir,
        )

    existing_updates = (
        {item.role: item for item in existing_basis.state_updates}
        if existing_basis is not None
        else {}
    )

    def build_state_update(
        *,
        role: str,
        target: Path,
        replacement: Path,
        replacement_bytes: bytes,
        schema: str,
    ) -> StateUpdate:
        target_bytes = target.read_bytes()
        replacement_sha = _sha256_bytes(replacement_bytes)
        target_relative = _relative(target, repo_root)
        replacement_relative = _relative(replacement, repo_root)
        existing = existing_updates.get(role)
        if existing is not None:
            if (
                existing.path != target_relative
                or existing.after_path != replacement_relative
                or existing.after_sha256 != replacement_sha
                or existing.schema != schema
            ):
                raise _blocked(
                    "PROMO-BASIS-CONFLICT",
                    f"existing promotion basis {role} binding conflicts with current terminal inputs",
                )
            actual_target_sha = _sha256_bytes(target_bytes)
            if actual_target_sha not in {existing.before_sha256, existing.after_sha256}:
                raise _blocked(
                    "PROMO-HASH-MISMATCH",
                    f"{role} is neither the bound pre-promotion nor signed-off state",
                )
            return existing
        return StateUpdate(
            role=role,
            path=target_relative,
            before_sha256=_sha256_bytes(target_bytes),
            after_path=replacement_relative,
            after_sha256=replacement_sha,
            schema=schema,
        )

    cycle_update = build_state_update(
        role="cycle-state",
        target=cycle_state_path,
        replacement=cycle_state_replacement_path,
        replacement_bytes=cycle_after_bytes,
        schema="codex-exec-cycle-state-v1",
    )
    workflow_update = build_state_update(
        role="workflow-state",
        target=workflow_state_path,
        replacement=workflow_state_replacement_path,
        replacement_bytes=workflow_after_bytes,
        schema="workflow-state-final-aliases-v1",
    )

    aliases: dict[str, ArtifactBinding | None] = {
        "final_findings": findings_binding,
        "final_traceability_matrix": matrix_binding,
        "final_traceability_matrix_xlsx": matrix_xlsx_binding,
        "final_writer_response": final_writer_response,
        "handoff_prompt": handoff_binding,
    }
    basis = PromotionBasis(
        schema_version=PROMOTION_BASIS_SCHEMA_VERSION,
        ft_root=seed.ft_root,
        cycle_dir=seed.cycle_dir,
        scope_slug=seed.scope_slug,
        candidate=seed.candidate,
        writer=seed.writer,
        source_basis=seed.source_basis,
        obligation_set=seed.obligation_set,
        prepared_package=seed.prepared_package,
        reviewer=seed.reviewer,
        publication=Publication(
            canonical_path=seed.canonical_path,
            expected_prior_sha256=seed.canonical_prior_sha256,
        ),
        gate_reports=seed.gate_reports,
        state_updates=(cycle_update, workflow_update),
        final_aliases=aliases,
    )
    _validate_cycle_state_after(
        cycle_after_text,
        candidate_sha=seed.candidate.sha256,
        canonical_path=canonical_ft_relative,
    )
    signed_off_snapshot_rel = _relative(
        cycle_dir / "versions" / "signed-off",
        ft_root,
    )
    _validate_workflow_state_after(
        workflow_after_text,
        basis=basis,
        repo_root=repo_root,
        ft_root=ft_root,
        cycle_state_path=cycle_state_path,
        signed_off_snapshot_rel=signed_off_snapshot_rel,
    )
    current_cycle_bytes = cycle_state_path.read_bytes()
    current_workflow_bytes = workflow_state_path.read_bytes()
    if _sha256_bytes(current_cycle_bytes) == cycle_update.before_sha256:
        try:
            _validate_cycle_state_before(
                current_cycle_bytes.decode("utf-8"),
                seed.candidate.sha256,
            )
        except UnicodeError as exc:
            raise _blocked("PROMO-INVALID-STATE", "cycle-state must be UTF-8") from exc
    elif _sha256_bytes(current_cycle_bytes) != cycle_update.after_sha256:
        raise _blocked("PROMO-HASH-MISMATCH", "cycle-state binding is stale")
    if _sha256_bytes(current_workflow_bytes) == workflow_update.before_sha256:
        try:
            _validate_workflow_state_before(
                current_workflow_bytes.decode("utf-8"),
                scope_slug=seed.scope_slug,
            )
        except UnicodeError as exc:
            raise _blocked("PROMO-INVALID-STATE", "workflow-state must be UTF-8") from exc
    elif _sha256_bytes(current_workflow_bytes) != workflow_update.after_sha256:
        raise _blocked("PROMO-HASH-MISMATCH", "workflow-state binding is stale")

    basis_payload: dict[str, Any] = {
        "schema_version": PROMOTION_BASIS_SCHEMA_VERSION,
        "ft_root": seed.ft_root,
        "cycle_dir": seed.cycle_dir,
        "scope_slug": seed.scope_slug,
        "candidate": _binding_payload(seed.candidate),
        "writer": _binding_payload(seed.writer),
        "source_basis": _binding_payload(seed.source_basis),
        "obligation_set": _binding_payload(seed.obligation_set),
        "prepared_package": _binding_payload(seed.prepared_package),
        "reviewer": _binding_payload(seed.reviewer),
        "publication": {
            "canonical_path": seed.canonical_path,
            "expected_prior_sha256": seed.canonical_prior_sha256,
        },
        "gate_reports": [_gate_payload(item) for item in seed.gate_reports],
        "state_updates": [
            _state_update_payload(item)
            for item in (cycle_update, workflow_update)
        ],
        "final_aliases": {
            name: (None if binding is None else _binding_payload(binding))
            for name, binding in aliases.items()
        },
    }
    desired_bytes = _canonical_basis_bytes(basis_payload)
    reused = False
    if basis_path.is_file():
        if basis_path.read_bytes() != desired_bytes:
            raise _blocked(
                "PROMO-BASIS-CONFLICT",
                "existing promotion-basis.json differs from the deterministic seed projection",
            )
        reused = True
    else:
        _atomic_write(basis_path, desired_bytes)
    loaded = _load_basis(basis_path)
    if loaded != basis:
        raise _blocked(
            "PROMO-BASIS-CONFLICT",
            "written promotion basis does not round-trip to the deterministic builder model",
        )
    duration_ms = _phase_ms(started, monotonic_ns())
    return PromotionBasisBuildResult(
        status="reused" if reused else "built",
        seed_path=seed_path,
        basis_path=basis_path,
        basis_sha256=_sha256_file(basis_path),
        reused=reused,
        build_duration_ms=duration_ms,
    )


def _validate_source_assertion_obligation_chain(
    manifest: SourceAssertionManifest,
    obligations: Sequence[PreparedObligation],
) -> None:
    """Recheck the source-to-obligation invariants at the promotion boundary."""

    assertion_by_atom = {item.atom_id: item for item in manifest.assertions}
    expected_atoms = {item.traceability_atom_id for item in obligations}
    actual_atoms = set(assertion_by_atom)
    if expected_atoms != actual_atoms:
        raise _blocked(
            "PROMO-INVALID-SOURCE-BASIS",
            "authenticated source assertion ATOM set does not match the bound "
            f"obligation set: missing={sorted(expected_atoms - actual_atoms)}, "
            f"unexpected={sorted(actual_atoms - expected_atoms)}",
        )

    expected_disposition = {
        "testable": "testable",
        "gap": "ambiguous",
        "unclear": "ambiguous",
        "not-applicable": "not-applicable",
    }
    for obligation in obligations:
        assertion = assertion_by_atom[obligation.traceability_atom_id]
        disposition = expected_disposition[obligation.coverage_status]
        if assertion.semantic_disposition != disposition:
            raise _blocked(
                "PROMO-INVALID-SOURCE-BASIS",
                f"{obligation.obligation_id} coverage status "
                f"{obligation.coverage_status} conflicts with authenticated "
                f"{assertion.assertion_id} semantic disposition "
                f"{assertion.semantic_disposition}",
            )
        if (
            obligation.coverage_status == "testable"
            and obligation.obligation_id not in assertion.obligation_ids
        ):
            raise _blocked(
                "PROMO-INVALID-SOURCE-BASIS",
                f"{obligation.obligation_id} is bound to the wrong authenticated "
                f"source assertion for {obligation.traceability_atom_id}",
            )
        if normalize_exact_source_text(obligation.atomic_statement) != (
            normalize_exact_source_text(assertion.canonical_statement)
        ):
            raise _blocked(
                "PROMO-INVALID-SOURCE-BASIS",
                f"{obligation.obligation_id} atomic statement does not match "
                f"authenticated {assertion.assertion_id}",
            )
        if obligation.coverage_status == "testable" and (
            normalize_exact_source_text(obligation.observable_oracle)
            != normalize_exact_source_text("; ".join(assertion.oracle_clauses))
        ):
            raise _blocked(
                "PROMO-INVALID-SOURCE-BASIS",
                f"{obligation.obligation_id} oracle does not match authenticated "
                f"{assertion.assertion_id}",
            )

        # Requirement codes are the canonical external references when present.
        # Source-row ids are the fallback for uncoded source assertions. Additional
        # compiler-derived aliases and dictionary ids remain allowed.
        canonical_refs = (
            assertion.requirement_codes
            if assertion.requirement_codes
            else (assertion.source_row_id,)
        )
        missing_refs = sorted(set(canonical_refs) - set(obligation.source_refs))
        if missing_refs:
            raise _blocked(
                "PROMO-INVALID-SOURCE-BASIS",
                f"{obligation.obligation_id} source_refs omit authenticated canonical "
                f"refs: {missing_refs}",
            )


def _validate_review_result(
    path: Path,
    *,
    repo_root: Path,
    scope_slug: str,
    writer: ArtifactBinding,
    source_basis: ArtifactBinding,
    source_basis_path: Path,
    obligation_set: ArtifactBinding,
    obligation_path: Path,
) -> Mapping[str, Any]:
    payload = _load_json(
        path, missing_code="PROMO-MISSING-REVIEW-RESULT", label="review-result.json"
    )
    expected_fields = {
        "schema_version",
        "contract_version",
        "decision",
        "reviewed_draft_path",
        "reviewed_draft_sha256",
        "reviewed_source_basis_sha256",
        "reviewed_obligation_set_sha256",
        "obligation_reviews",
        "dimension_reviews",
        "findings",
        "summary",
    }
    _exact_fields(payload, expected_fields, "review-result")
    if payload["schema_version"] != REVIEW_RESULT_SCHEMA_VERSION or payload["contract_version"] != 4:
        raise _blocked(
            "PROMO-INVALID-REVIEW-RESULT",
            "review-result must use normalized schema v2 and source-first reviewer contract v4",
        )
    if payload["decision"] != "accepted":
        raise _blocked("PROMO-NOT-ACCEPTED", "review-result decision is not accepted")
    comparisons = {
        "reviewed_draft_path": writer.path,
        "reviewed_draft_sha256": writer.sha256,
        "reviewed_source_basis_sha256": source_basis.sha256,
        "reviewed_obligation_set_sha256": obligation_set.sha256,
    }
    mismatches = {
        key: {"expected": value, "actual": payload.get(key)}
        for key, value in comparisons.items()
        if payload.get(key) != value
    }
    if mismatches:
        raise _blocked(
            "PROMO-HASH-MISMATCH", f"review-result hash chain mismatch: {mismatches}"
        )
    prepared_obligations = load_obligations(obligation_path)
    blocking_gap_ids = tuple(
        sorted(
            gap.gap_id
            for gap in prepared_obligations.coverage_gaps
            if gap.blocking
        )
    )
    if blocking_gap_ids:
        raise _blocked(
            "PROMO-BLOCKING-SOURCE-GAPS",
            "accepted draft contains blocking source gaps: "
            + ", ".join(blocking_gap_ids),
        )
    expected_obligations = prepared_obligations.obligations
    expected_by_id = {item.obligation_id: item for item in expected_obligations}
    try:
        source_basis_text = source_basis_path.read_text(encoding="utf-8")
        source_contract = parse_embedded_source_assertion_contract(
            source_basis_text,
            repo_root,
            expected_scope_slug=scope_slug,
            expected_obligation_ids=(
                item.obligation_id
                for item in expected_obligations
                if item.coverage_status == "testable"
            ),
        )
    except (OSError, UnicodeError, SourceAssertionContractError) as exc:
        raise _blocked(
            "PROMO-INVALID-SOURCE-BASIS",
            f"source basis does not contain an authentic accepted source assertion "
            f"contract: {exc}",
        ) from exc
    if source_contract is None:
        raise _blocked(
            "PROMO-INVALID-SOURCE-BASIS",
            "source-first reviewer contract v4 cannot be promoted from legacy plain "
            "Markdown evidence without an embedded source assertion contract",
        )
    execution_blocked = tuple(
        item
        for item in source_contract.manifest.assertions
        if item.semantic_disposition == "testable"
        and item.execution_readiness == "dependency-blocked"
    )
    if execution_blocked:
        raise _blocked(
            "PROMO-BLOCKED-EXECUTION-DEPENDENCIES",
            "authenticated source assertions retain execution dependencies: "
            + ", ".join(item.assertion_id for item in execution_blocked),
        )
    _validate_source_assertion_obligation_chain(
        source_contract.manifest,
        expected_obligations,
    )
    obligation_reviews = payload["obligation_reviews"]
    if not isinstance(obligation_reviews, list) or not obligation_reviews:
        raise _blocked(
            "PROMO-INVALID-REVIEW-RESULT",
            "accepted review-result requires a compact per-obligation receipt",
        )
    actual_verdicts: dict[str, str] = {}
    for item in obligation_reviews:
        if not isinstance(item, Mapping) or set(item) != {"obligation_id", "verdict"}:
            raise _blocked(
                "PROMO-INVALID-REVIEW-RESULT",
                "review-result obligation receipt item has invalid fields",
            )
        obligation_id = item.get("obligation_id")
        verdict = item.get("verdict")
        if not isinstance(obligation_id, str) or obligation_id in actual_verdicts:
            raise _blocked(
                "PROMO-INVALID-REVIEW-RESULT",
                "review-result obligation receipt contains an invalid or duplicate id",
            )
        actual_verdicts[obligation_id] = str(verdict or "")
    if set(actual_verdicts) != set(expected_by_id):
        raise _blocked(
            "PROMO-INVALID-REVIEW-RESULT",
            "review-result obligation receipt does not match the bound obligation set",
        )
    for obligation_id, obligation in expected_by_id.items():
        expected_verdict = (
            "covered" if obligation.coverage_status == "testable" else "gap-preserved"
        )
        if actual_verdicts[obligation_id] != expected_verdict:
            raise _blocked(
                "PROMO-INVALID-REVIEW-RESULT",
                f"accepted review-result has non-passing verdict for {obligation_id}",
            )
    try:
        dimension_source_refs = parse_reviewer_dimension_source_bindings(
            source_basis_text
        ).as_mapping()
    except (OSError, StageRuntimeError) as exc:
        raise _blocked(
            "PROMO-INVALID-REVIEW-RESULT",
            f"source basis has invalid reviewer dimension bindings: {exc}",
        ) from exc
    dimensions = payload["dimension_reviews"]
    if not isinstance(dimensions, list):
        raise _blocked(
            "PROMO-INVALID-REVIEW-RESULT", "review-result.dimension_reviews must be an array"
        )
    actual_dimensions: dict[str, Mapping[str, Any]] = {}
    for item in dimensions:
        if not isinstance(item, Mapping) or set(item) != {
            "dimension",
            "verdict",
            "source_refs",
            "note",
        }:
            raise _blocked(
                "PROMO-INVALID-REVIEW-RESULT",
                "review-result dimension receipt item has invalid fields",
            )
        dimension = item.get("dimension")
        if not isinstance(dimension, str) or dimension in actual_dimensions:
            raise _blocked(
                "PROMO-INVALID-REVIEW-RESULT",
                "review-result dimension receipt contains an invalid or duplicate dimension",
            )
        if item.get("verdict") != "verified":
            raise _blocked(
                "PROMO-INVALID-REVIEW-RESULT",
                "accepted review-result requires every dimension verdict to be verified",
            )
        source_refs = item.get("source_refs")
        if (
            not isinstance(source_refs, list)
            or not source_refs
            or any(not isinstance(value, str) or not value.strip() for value in source_refs)
            or len(source_refs) != len(set(source_refs))
        ):
            raise _blocked(
                "PROMO-INVALID-REVIEW-RESULT",
                f"review-result dimension {dimension} has invalid source refs",
            )
        if not isinstance(item.get("note"), str) or not str(item["note"]).strip():
            raise _blocked(
                "PROMO-INVALID-REVIEW-RESULT",
                f"review-result dimension {dimension} note must be non-empty",
            )
        actual_dimensions[dimension] = item
    if set(actual_dimensions) != set(dimension_source_refs):
        raise _blocked(
            "PROMO-INVALID-REVIEW-RESULT",
            "review-result dimension receipt does not match the bound source-basis set",
        )
    for dimension, item in actual_dimensions.items():
        expected_refs = dimension_source_refs[dimension]
        if tuple(item["source_refs"]) != expected_refs:
            raise _blocked(
                "PROMO-INVALID-REVIEW-RESULT",
                f"review-result dimension {dimension} must cite the exact canonical "
                f"bound source refs: expected={list(expected_refs)}, "
                f"actual={item['source_refs']}",
            )
    findings = payload["findings"]
    if not isinstance(findings, list):
        raise _blocked("PROMO-INVALID-REVIEW-RESULT", "review-result.findings must be an array")
    if any(isinstance(item, Mapping) and item.get("severity") == "error" for item in findings):
        raise _blocked(
            "PROMO-INVALID-REVIEW-RESULT", "accepted review-result contains an error finding"
        )
    if not isinstance(payload["summary"], str) or not payload["summary"].strip():
        raise _blocked("PROMO-INVALID-REVIEW-RESULT", "review-result.summary must be non-empty")
    return payload


def _validate_gate_reports(
    bindings: Sequence[GateBinding], repo_root: Path, cycle_dir: Path, candidate_sha: str
) -> tuple[dict[str, Any], ...]:
    results: list[dict[str, Any]] = []
    for binding in bindings:
        artifact = ArtifactBinding(binding.path, binding.sha256)
        path = _verify_binding(artifact, repo_root, label="gate report", parent=cycle_dir)
        payload = _load_json(path, missing_code="PROMO-MISSING-GATE", label="gate report")
        if payload.get("validator") != binding.validator:
            raise _blocked(
                "PROMO-GATE-MISMATCH",
                f"gate validator mismatch for {binding.path}: {payload.get('validator')}",
            )
        if payload.get("passed") is not True:
            raise _blocked("PROMO-GATE-FAILED", f"required gate did not pass: {binding.path}")
        if payload.get("draft_sha256") != candidate_sha:
            raise _blocked(
                "PROMO-HASH-MISMATCH",
                f"gate report is not bound to accepted draft: {binding.path}",
            )
        results.append(
            {
                "path": binding.path,
                "sha256": binding.sha256,
                "validator": binding.validator,
            }
        )
    return tuple(results)


def _atomic_write(path: Path, content: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.{os.getpid()}.promotion-tmp")
    if temporary.exists():
        raise _blocked("PROMO-TEMP-CONFLICT", f"promotion temporary path exists: {temporary}")
    try:
        with temporary.open("xb") as handle:
            handle.write(content)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def _captured_binding_bytes(
    path: Path,
    binding: ArtifactBinding,
    *,
    label: str,
) -> bytes:
    content = path.read_bytes()
    actual = _sha256_bytes(content)
    if actual != binding.sha256:
        raise _blocked(
            "PROMO-HASH-MISMATCH",
            f"{label} changed while it was being captured: expected "
            f"{binding.sha256}, got {actual}",
        )
    return content


def _verify_and_capture_binding(
    binding: ArtifactBinding,
    repo_root: Path,
    *,
    label: str,
    parent: Path | None = None,
    missing_code: str = "PROMO-MISSING-ARTIFACT",
) -> tuple[Path, bytes]:
    path = _resolve(repo_root, binding.path, label)
    if parent is not None:
        try:
            path.relative_to(parent.resolve())
        except ValueError as exc:
            raise _blocked(
                "PROMO-UNSAFE-PATH",
                f"{label} must be inside {parent}: {path}",
            ) from exc
    if not path.is_file():
        raise _blocked(missing_code, f"required {label} is missing: {path}")
    return path, _captured_binding_bytes(path, binding, label=label)


def _assert_file_sha(path: Path, expected_sha: str, *, label: str) -> None:
    if not path.is_file() or _sha256_file(path) != expected_sha:
        raise _blocked(
            "PROMO-HASH-MISMATCH",
            f"{label} changed after promotion preflight",
        )


def _recovery_journal_root(cycle_dir: Path, transaction_id: str) -> Path:
    return cycle_dir / "promotion" / f".{transaction_id}.journal"


def _recovery_prepare_root(
    cycle_dir: Path,
    transaction_id: str,
    owner_token: str,
) -> Path:
    return cycle_dir / "promotion" / f".{transaction_id}.{owner_token}.prepare"


def _recovery_targets(
    *,
    basis: PromotionBasis,
    repo_root: Path,
) -> list[dict[str, Any]]:
    targets: list[dict[str, Any]] = [
        {
            "role": "canonical",
            "path": basis.publication.canonical_path,
            "before_sha256": basis.publication.expected_prior_sha256,
            "after_sha256": basis.candidate.sha256,
            "before_snapshot_path": (
                None
                if basis.publication.expected_prior_sha256 is None
                else f"before/repo/{basis.publication.canonical_path}"
            ),
        }
    ]
    for update in sorted(basis.state_updates, key=lambda item: item.role):
        targets.append(
            {
                "role": update.role,
                "path": update.path,
                "before_sha256": update.before_sha256,
                "after_sha256": update.after_sha256,
                "before_snapshot_path": f"before/repo/{update.path}",
            }
        )
    resolved = [
        _resolve(repo_root, item["path"], f"recovery target {item['role']}")
        for item in targets
    ]
    if len(set(resolved)) != len(resolved):
        raise _blocked(
            "PROMO-RECOVERY-CONFLICT",
            "promotion recovery targets must be path-distinct",
        )
    return targets


def _recovery_journal_payload(
    *,
    basis: PromotionBasis,
    basis_path: Path,
    basis_sha: str,
    repo_root: Path,
    transaction_id: str,
    owner_token: str,
) -> dict[str, Any]:
    return {
        "schema_version": RECOVERY_JOURNAL_SCHEMA_VERSION,
        "status": "prepared",
        "transaction_id": transaction_id,
        "owner_token": owner_token,
        "promotion_basis": {
            "path": _relative(basis_path, repo_root),
            "sha256": basis_sha,
        },
        "targets": _recovery_targets(basis=basis, repo_root=repo_root),
    }


def _load_and_validate_recovery_journal(
    root: Path,
    *,
    basis: PromotionBasis,
    basis_path: Path,
    basis_sha: str,
    repo_root: Path,
    transaction_id: str,
) -> Mapping[str, Any]:
    payload = _load_json(
        root / "recovery-journal.json",
        missing_code="PROMO-RECOVERY-CONFLICT",
        label="promotion recovery journal",
    )
    owner_token = payload.get("owner_token")
    if (
        not isinstance(owner_token, str)
        or re.fullmatch(r"[0-9a-f]{32}", owner_token) is None
    ):
        raise _blocked(
            "PROMO-RECOVERY-CONFLICT",
            "promotion recovery journal has an invalid owner token",
        )
    expected = _recovery_journal_payload(
        basis=basis,
        basis_path=basis_path,
        basis_sha=basis_sha,
        repo_root=repo_root,
        transaction_id=transaction_id,
        owner_token=owner_token,
    )
    if payload != expected:
        raise _blocked(
            "PROMO-RECOVERY-CONFLICT",
            "promotion recovery journal does not match the immutable promotion basis",
        )
    return payload


def _recovery_target_state(
    path: Path,
    *,
    before_sha: str | None,
    after_sha: str,
) -> str:
    if not path.exists():
        return "before" if before_sha is None else "unknown"
    if not path.is_file():
        return "unknown"
    actual = _sha256_file(path)
    before_match = before_sha is not None and actual == before_sha
    after_match = actual == after_sha
    if before_match and after_match:
        return "both"
    if before_match:
        return "before"
    if after_match:
        return "after"
    return "unknown"


def _rollback_recovery_journal(
    journal_root: Path,
    *,
    basis: PromotionBasis,
    basis_path: Path,
    basis_sha: str,
    repo_root: Path,
    transaction_id: str,
) -> None:
    payload = _load_and_validate_recovery_journal(
        journal_root,
        basis=basis,
        basis_path=basis_path,
        basis_sha=basis_sha,
        repo_root=repo_root,
        transaction_id=transaction_id,
    )
    material: list[tuple[Mapping[str, Any], Path, bytes | None]] = []
    for item in payload["targets"]:
        target = _resolve(repo_root, item["path"], f"recovery target {item['role']}")
        state = _recovery_target_state(
            target,
            before_sha=item["before_sha256"],
            after_sha=item["after_sha256"],
        )
        if state == "unknown":
            raise _blocked(
                "PROMO-RECOVERY-AMBIGUOUS",
                f"recovery target {item['role']} is neither the before nor after image",
            )
        snapshot_relative = item["before_snapshot_path"]
        before_bytes: bytes | None = None
        if snapshot_relative is not None:
            snapshot = (journal_root / Path(snapshot_relative)).resolve()
            try:
                snapshot.relative_to(journal_root.resolve())
            except ValueError as exc:
                raise _blocked(
                    "PROMO-RECOVERY-CONFLICT",
                    f"recovery snapshot escapes journal root: {snapshot_relative}",
                ) from exc
            if not snapshot.is_file():
                raise _blocked(
                    "PROMO-RECOVERY-CONFLICT",
                    f"recovery before image is missing: {snapshot_relative}",
                )
            before_bytes = snapshot.read_bytes()
            if _sha256_bytes(before_bytes) != item["before_sha256"]:
                raise _blocked(
                    "PROMO-RECOVERY-CONFLICT",
                    f"recovery before image hash mismatch: {snapshot_relative}",
                )
        material.append((item, target, before_bytes))

    for item, target, before_bytes in material:
        if before_bytes is None:
            target.unlink(missing_ok=True)
        else:
            _atomic_write(target, before_bytes)
        if (
            _recovery_target_state(
                target,
                before_sha=item["before_sha256"],
                after_sha=item["after_sha256"],
            )
            not in {"before", "both"}
        ):
            raise _blocked(
                "PROMO-RECOVERY-AMBIGUOUS",
                f"recovery could not restore {item['role']} to its before image",
            )

    rolled_back = journal_root.with_suffix(".rolled-back")
    if rolled_back.exists():
        raise _blocked(
            "PROMO-RECOVERY-CONFLICT",
            f"promotion recovery tombstone already exists: {rolled_back}",
        )
    os.replace(journal_root, rolled_back)
    shutil.rmtree(rolled_back, ignore_errors=True)


def _targets_are_before(
    *,
    basis: PromotionBasis,
    repo_root: Path,
) -> bool:
    return all(
        _recovery_target_state(
            _resolve(repo_root, item["path"], f"recovery target {item['role']}"),
            before_sha=item["before_sha256"],
            after_sha=item["after_sha256"],
        )
        in {"before", "both"}
        for item in _recovery_targets(basis=basis, repo_root=repo_root)
    )


def _reconcile_interrupted_promotion(
    *,
    repo_root: Path,
    cycle_dir: Path,
    basis: PromotionBasis,
    basis_path: Path,
    basis_sha: str,
) -> None:
    transaction_id = f"{basis.candidate.sha256[:16]}-{basis.reviewer.sha256[:8]}"
    journal_root = _recovery_journal_root(cycle_dir, transaction_id)
    transaction_root = cycle_dir / "promotion" / transaction_id
    snapshot_root = cycle_dir / "versions" / "signed-off"
    if journal_root.exists() and transaction_root.exists():
        raise _blocked(
            "PROMO-RECOVERY-AMBIGUOUS",
            "both prepared and committed transaction roots exist",
        )
    if journal_root.exists():
        if snapshot_root.exists():
            raise _blocked(
                "PROMO-RECOVERY-AMBIGUOUS",
                "signed-off snapshot exists before the transaction commit marker",
            )
        _rollback_recovery_journal(
            journal_root,
            basis=basis,
            basis_path=basis_path,
            basis_sha=basis_sha,
            repo_root=repo_root,
            transaction_id=transaction_id,
        )
        return
    if not transaction_root.exists():
        return
    recovery_journal = transaction_root / "recovery-journal.json"
    if not recovery_journal.is_file():
        return
    _load_and_validate_recovery_journal(
        transaction_root,
        basis=basis,
        basis_path=basis_path,
        basis_sha=basis_sha,
        repo_root=repo_root,
        transaction_id=transaction_id,
    )
    if not all(
        _recovery_target_state(
            _resolve(repo_root, item["path"], f"recovery target {item['role']}"),
            before_sha=item["before_sha256"],
            after_sha=item["after_sha256"],
        )
        in {"after", "both"}
        for item in _recovery_targets(basis=basis, repo_root=repo_root)
    ):
        raise _blocked(
            "PROMO-RECOVERY-AMBIGUOUS",
            "committed transaction targets are not all in their after image",
        )
    internal_snapshot = transaction_root / "final-snapshot"
    if internal_snapshot.exists() and snapshot_root.exists():
        raise _blocked(
            "PROMO-RECOVERY-AMBIGUOUS",
            "both internal and published signed-off snapshots exist",
        )
    if internal_snapshot.is_dir():
        snapshot_root.parent.mkdir(parents=True, exist_ok=True)
        os.replace(internal_snapshot, snapshot_root)
    elif not snapshot_root.is_dir():
        raise _blocked(
            "PROMO-RECOVERY-AMBIGUOUS",
            "committed transaction has no recoverable signed-off snapshot",
        )


@contextmanager
def _materialized_candidate(cycle_dir: Path, content: bytes) -> Iterator[Path]:
    path = cycle_dir / f".promotion-candidate-{os.getpid()}.md"
    if path.exists():
        raise _blocked(
            "PROMO-TEMP-CONFLICT",
            f"promotion candidate materialization already exists: {path}",
        )
    try:
        with path.open("xb") as handle:
            handle.write(content)
            handle.flush()
            os.fsync(handle.fileno())
        yield path
    finally:
        path.unlink(missing_ok=True)


def _write_json(path: Path, payload: Mapping[str, Any]) -> None:
    _atomic_write(
        path,
        (json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n").encode("utf-8"),
    )


def _process_is_alive(pid: int) -> bool:
    if pid <= 0:
        return False
    if pid == os.getpid():
        return True
    if os.name == "nt":
        import ctypes

        process_query_limited_information = 0x1000
        still_active = 259
        kernel32 = ctypes.WinDLL("kernel32", use_last_error=True)
        kernel32.OpenProcess.argtypes = [
            ctypes.c_ulong,
            ctypes.c_int,
            ctypes.c_ulong,
        ]
        kernel32.OpenProcess.restype = ctypes.c_void_p
        kernel32.GetExitCodeProcess.argtypes = [
            ctypes.c_void_p,
            ctypes.POINTER(ctypes.c_ulong),
        ]
        kernel32.GetExitCodeProcess.restype = ctypes.c_int
        kernel32.CloseHandle.argtypes = [ctypes.c_void_p]
        kernel32.CloseHandle.restype = ctypes.c_int
        handle = kernel32.OpenProcess(
            process_query_limited_information,
            False,
            pid,
        )
        if not handle:
            return ctypes.get_last_error() == 5
        try:
            exit_code = ctypes.c_ulong()
            if not kernel32.GetExitCodeProcess(handle, ctypes.byref(exit_code)):
                return True
            return exit_code.value == still_active
        finally:
            kernel32.CloseHandle(handle)
    try:
        os.kill(pid, 0)
    except ProcessLookupError:
        return False
    except PermissionError:
        return True
    return True


def _promotion_target_lock_specs(
    repo_root: Path,
    canonical_path: Path,
    workflow_state_path: Path,
) -> list[tuple[str, str, Path, tuple[str, ...], Path]]:
    lock_root = repo_root / ".qa-agent-promotion-locks"
    target_roles: dict[str, tuple[Path, set[str]]] = {}
    for role, target in (
        ("canonical", canonical_path),
        ("workflow-state", workflow_state_path),
    ):
        resolved = target.resolve()
        normalized = os.path.normcase(os.path.normpath(str(resolved)))
        existing = target_roles.get(normalized)
        if existing is None:
            target_roles[normalized] = (resolved, {role})
        else:
            existing[1].add(role)
    return [
        (
            lock_key,
            normalized,
            resolved,
            tuple(sorted(roles)),
            lock_root / f"{lock_key}.lock",
        )
        for lock_key, normalized, resolved, roles in sorted(
            (
                _sha256_bytes((normalized + "\n").encode("utf-8")),
                normalized,
                resolved,
                roles,
            )
            for normalized, (resolved, roles) in target_roles.items()
        )
    ]


def _lock_metadata(
    *,
    repo_root: Path,
    cycle_dir: Path,
    basis_path: Path,
    basis_sha: str,
    transaction_id: str,
    owner_token: str,
    roles: Sequence[str],
    target: Path | None,
) -> dict[str, Any]:
    return {
        "schema_version": LOCK_SCHEMA_VERSION,
        "pid": os.getpid(),
        "owner_token": owner_token,
        "cycle_dir": _relative(cycle_dir, repo_root),
        "basis_path": _relative(basis_path, repo_root),
        "basis_sha256": basis_sha,
        "transaction_id": transaction_id,
        "roles": list(roles),
        "target": None if target is None else str(target.resolve()),
    }


def _write_lock_descriptor(descriptor: int, payload: Mapping[str, Any]) -> None:
    content = (
        json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        + "\n"
    ).encode("utf-8")
    os.write(descriptor, content)
    os.fsync(descriptor)


@contextmanager
def _promotion_lock(
    cycle_dir: Path,
    *,
    repo_root: Path,
    basis_path: Path,
    basis_sha: str,
    transaction_id: str,
    owner_token: str,
) -> Iterator[None]:
    lock_path = cycle_dir / "promotion.lock"
    try:
        descriptor = os.open(lock_path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
    except FileExistsError as exc:
        raise _blocked("PROMO-LOCKED", f"another promotion owns {lock_path}") from exc
    try:
        _write_lock_descriptor(
            descriptor,
            _lock_metadata(
                repo_root=repo_root,
                cycle_dir=cycle_dir,
                basis_path=basis_path,
                basis_sha=basis_sha,
                transaction_id=transaction_id,
                owner_token=owner_token,
                roles=("cycle",),
                target=None,
            ),
        )
        os.close(descriptor)
        descriptor = -1
        yield
    finally:
        if descriptor >= 0:
            os.close(descriptor)
        lock_path.unlink(missing_ok=True)


@contextmanager
def _promotion_target_lock(
    repo_root: Path,
    canonical_path: Path,
    workflow_state_path: Path,
    *,
    cycle_dir: Path,
    basis_path: Path,
    basis_sha: str,
    transaction_id: str,
    owner_token: str,
) -> Iterator[None]:
    """Serialize promotions that overlap on any production/handoff target."""

    lock_root = repo_root / ".qa-agent-promotion-locks"
    lock_root.mkdir(parents=True, exist_ok=True)
    lock_specs = _promotion_target_lock_specs(
        repo_root,
        canonical_path,
        workflow_state_path,
    )
    acquired: list[Path] = []
    try:
        for _, _, resolved, roles, lock_path in lock_specs:
            try:
                descriptor = os.open(
                    lock_path,
                    os.O_CREAT | os.O_EXCL | os.O_WRONLY,
                )
            except FileExistsError as exc:
                raise _blocked(
                    "PROMO-TARGET-LOCKED",
                    f"another promotion owns target {resolved}",
                ) from exc
            try:
                _write_lock_descriptor(
                    descriptor,
                    _lock_metadata(
                        repo_root=repo_root,
                        cycle_dir=cycle_dir,
                        basis_path=basis_path,
                        basis_sha=basis_sha,
                        transaction_id=transaction_id,
                        owner_token=owner_token,
                        roles=roles,
                        target=resolved,
                    ),
                )
                os.close(descriptor)
                descriptor = -1
            finally:
                if descriptor >= 0:
                    os.close(descriptor)
                    lock_path.unlink(missing_ok=True)
            acquired.append(lock_path)
        yield
    finally:
        for lock_path in reversed(acquired):
            lock_path.unlink(missing_ok=True)
        try:
            lock_root.rmdir()
        except OSError:
            pass


def _read_lock_metadata(path: Path) -> Mapping[str, Any] | None:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, UnicodeError, json.JSONDecodeError):
        return None
    expected_fields = {
        "schema_version",
        "pid",
        "owner_token",
        "cycle_dir",
        "basis_path",
        "basis_sha256",
        "transaction_id",
        "roles",
        "target",
    }
    if not isinstance(payload, Mapping) or set(payload) != expected_fields:
        return None
    if payload.get("schema_version") != LOCK_SCHEMA_VERSION:
        return None
    pid = payload.get("pid")
    owner_token = payload.get("owner_token")
    roles = payload.get("roles")
    target = payload.get("target")
    if (
        not isinstance(pid, int)
        or isinstance(pid, bool)
        or pid <= 0
        or not isinstance(owner_token, str)
        or re.fullmatch(r"[0-9a-f]{32}", owner_token) is None
        or not isinstance(roles, list)
        or not roles
        or any(not isinstance(item, str) or not item for item in roles)
        or len(roles) != len(set(roles))
        or (target is not None and (not isinstance(target, str) or not target))
    ):
        return None
    try:
        _relative_path(payload.get("cycle_dir"), "lock cycle_dir")
        _relative_path(payload.get("basis_path"), "lock basis_path")
        _sha(payload.get("basis_sha256"), "lock basis_sha256")
        _text(payload.get("transaction_id"), "lock transaction_id")
    except PromotionBlocked:
        return None
    return payload


@contextmanager
def _recovery_claim(
    repo_root: Path,
    owner_metadata: Mapping[str, Any],
) -> Iterator[None]:
    owner_token = owner_metadata["owner_token"]
    lock_root = repo_root / ".qa-agent-promotion-locks"
    lock_root.mkdir(parents=True, exist_ok=True)
    claim = lock_root / f"recovery-{owner_token}.lock"
    for _ in range(2):
        try:
            descriptor = os.open(claim, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        except FileExistsError as exc:
            metadata = _read_lock_metadata(claim)
            if metadata is None or _process_is_alive(metadata["pid"]):
                raise _blocked(
                    "PROMO-RECOVERY-IN-PROGRESS",
                    f"another process owns recovery claim {claim}",
                ) from exc
            claim.unlink(missing_ok=True)
            continue
        try:
            _write_lock_descriptor(
                descriptor,
                {
                    **owner_metadata,
                    "pid": os.getpid(),
                    "roles": ["recovery"],
                    "target": None,
                },
            )
            os.close(descriptor)
            descriptor = -1
            try:
                yield
            finally:
                claim.unlink(missing_ok=True)
            return
        finally:
            if descriptor >= 0:
                os.close(descriptor)
                claim.unlink(missing_ok=True)
    raise _blocked(
        "PROMO-RECOVERY-IN-PROGRESS",
        f"could not acquire recovery claim {claim}",
    )


def _recover_lock_owner(repo_root: Path, observed_lock: Path) -> bool:
    metadata = _read_lock_metadata(observed_lock)
    if metadata is None:
        return False
    if _process_is_alive(metadata["pid"]):
        raise _blocked(
            "PROMO-TARGET-LOCKED",
            f"live promotion process {metadata['pid']} owns {observed_lock}",
        )
    owner_token = metadata["owner_token"]
    with _recovery_claim(repo_root, metadata):
        current = _read_lock_metadata(observed_lock)
        if current is None:
            if observed_lock.exists():
                raise _blocked(
                    "PROMO-RECOVERY-AMBIGUOUS",
                    f"stale lock metadata became unreadable: {observed_lock}",
                )
            return True
        if current != metadata:
            raise _blocked(
                "PROMO-RECOVERY-AMBIGUOUS",
                f"stale lock ownership changed during recovery: {observed_lock}",
            )
        if _process_is_alive(current["pid"]):
            raise _blocked(
                "PROMO-TARGET-LOCKED",
                f"promotion process {current['pid']} revived or PID was reused",
            )
        cycle_dir = _resolve(repo_root, current["cycle_dir"], "stale lock cycle_dir")
        basis_path = _resolve(repo_root, current["basis_path"], "stale lock basis_path")
        basis, _, basis_sha = _read_basis_snapshot(basis_path)
        if basis_sha != current["basis_sha256"]:
            raise _blocked(
                "PROMO-RECOVERY-CONFLICT",
                "stale lock promotion basis no longer matches its owner binding",
            )
        if _relative(cycle_dir, repo_root) != basis.cycle_dir:
            raise _blocked(
                "PROMO-RECOVERY-CONFLICT",
                "stale lock cycle does not match its promotion basis",
            )
        transaction_id = (
            f"{basis.candidate.sha256[:16]}-{basis.reviewer.sha256[:8]}"
        )
        if transaction_id != current["transaction_id"]:
            raise _blocked(
                "PROMO-RECOVERY-CONFLICT",
                "stale lock transaction does not match its promotion basis",
            )
        _reconcile_interrupted_promotion(
            repo_root=repo_root,
            cycle_dir=cycle_dir,
            basis=basis,
            basis_path=basis_path,
            basis_sha=basis_sha,
        )
        journal = _recovery_journal_root(cycle_dir, transaction_id)
        transaction = cycle_dir / "promotion" / transaction_id
        if not journal.exists() and not transaction.exists() and not _targets_are_before(
            basis=basis,
            repo_root=repo_root,
        ):
            raise _blocked(
                "PROMO-RECOVERY-AMBIGUOUS",
                "stale lock has no journal and promotion targets are not in the before image",
            )

        for item in _recovery_targets(basis=basis, repo_root=repo_root):
            target = _resolve(
                repo_root,
                item["path"],
                f"stale lock temporary target {item['role']}",
            )
            temporary = target.with_name(
                f".{target.name}.{current['pid']}.promotion-tmp"
            )
            if temporary.is_file():
                temporary.unlink()
        materialized_candidate = (
            cycle_dir / f".promotion-candidate-{current['pid']}.md"
        )
        if materialized_candidate.is_file():
            materialized_candidate.unlink()

        canonical = _resolve(
            repo_root,
            basis.publication.canonical_path,
            "stale lock canonical target",
        )
        workflow_update = next(
            item for item in basis.state_updates if item.role == "workflow-state"
        )
        workflow = _resolve(
            repo_root,
            workflow_update.path,
            "stale lock workflow target",
        )
        owner_lock_paths = [
            item[4]
            for item in _promotion_target_lock_specs(repo_root, canonical, workflow)
        ]
        owner_lock_paths.append(cycle_dir / "promotion.lock")
        for path in owner_lock_paths:
            if not path.exists():
                continue
            lock_metadata = _read_lock_metadata(path)
            if (
                lock_metadata is None
                or lock_metadata.get("owner_token") != owner_token
                or _process_is_alive(lock_metadata["pid"])
            ):
                raise _blocked(
                    "PROMO-RECOVERY-AMBIGUOUS",
                    f"cannot prove stale ownership for lock {path}",
                )
        for path in owner_lock_paths:
            path.unlink(missing_ok=True)
        prepare = _recovery_prepare_root(cycle_dir, transaction_id, owner_token)
        shutil.rmtree(prepare, ignore_errors=True)
    try:
        (repo_root / ".qa-agent-promotion-locks").rmdir()
    except OSError:
        pass
    return True


def _recover_expected_locks(
    *,
    repo_root: Path,
    cycle_dir: Path,
    canonical_path: Path,
    workflow_state_path: Path,
) -> None:
    expected = [
        item[4]
        for item in _promotion_target_lock_specs(
            repo_root,
            canonical_path,
            workflow_state_path,
        )
    ]
    expected.append(cycle_dir / "promotion.lock")
    for lock_path in expected:
        if lock_path.exists() and not _recover_lock_owner(repo_root, lock_path):
            return


def _snapshot_write(root: Path, repo_relative: str, content: bytes) -> dict[str, Any]:
    path = root / Path(repo_relative)
    _atomic_write(path, content)
    return {
        "path": repo_relative,
        "sha256": _sha256_bytes(content),
        "size_bytes": len(content),
    }


def _write_snapshot_yaml(
    path: Path,
    *,
    cycle_id: str,
    entries: Sequence[Mapping[str, Any]],
) -> None:
    _atomic_write(path, _snapshot_yaml_bytes(cycle_id=cycle_id, entries=entries))


def _snapshot_yaml_bytes(
    *,
    cycle_id: str,
    entries: Sequence[Mapping[str, Any]],
) -> bytes:
    lines = [
        "version: 1",
        "snapshot_id: signed-off",
        f"cycle_id: {cycle_id}",
        "stage_status: signed-off",
        "files:",
    ]
    for item in sorted(entries, key=lambda value: str(value["path"])):
        lines.append(
            "  - "
            + " | ".join(
                (
                    str(item["path"]),
                    str(item["sha256"]),
                    str(item["size_bytes"]),
                )
            )
        )
    return ("\n".join(lines) + "\n").encode("utf-8")


def _phase_ms(start_ns: int, end_ns: int) -> int:
    return max(0, (end_ns - start_ns) // 1_000_000)


def _call_fault(fault_injector: Callable[[str], None] | None, phase: str) -> None:
    if fault_injector is not None:
        fault_injector(phase)


def _verify_signed_off_snapshot(
    snapshot_root: Path,
    *,
    cycle_id: str,
    expected_hashes: Mapping[str, str],
) -> None:
    manifest_path = snapshot_root / "snapshot-manifest.json"
    payload = _load_json(
        manifest_path,
        missing_code="PROMO-IDEMPOTENCY-CONFLICT",
        label="signed-off snapshot manifest",
    )
    if payload.get("schema_version") != 1 or payload.get("snapshot_id") != "signed-off":
        raise _blocked(
            "PROMO-IDEMPOTENCY-CONFLICT",
            "signed-off snapshot manifest has an unsupported identity",
        )
    files = payload.get("files")
    if not isinstance(files, list) or not files:
        raise _blocked(
            "PROMO-IDEMPOTENCY-CONFLICT",
            "signed-off snapshot manifest has no files",
        )
    seen: set[str] = set()
    hashes: dict[str, str] = {}
    for index, item in enumerate(files):
        if not isinstance(item, Mapping):
            raise _blocked(
                "PROMO-IDEMPOTENCY-CONFLICT",
                f"signed-off snapshot file[{index}] is invalid",
            )
        _exact_fields(
            item,
            {"path", "sha256", "size_bytes"},
            f"signed-off snapshot file[{index}]",
        )
        relative = _relative_path(item["path"], f"snapshot file[{index}].path")
        if relative in seen:
            raise _blocked(
                "PROMO-IDEMPOTENCY-CONFLICT",
                f"signed-off snapshot repeats {relative}",
            )
        seen.add(relative)
        expected_sha = _sha(item["sha256"], f"snapshot file[{index}].sha256")
        hashes[relative] = expected_sha
        expected_size = item["size_bytes"]
        if not isinstance(expected_size, int) or isinstance(expected_size, bool) or expected_size < 0:
            raise _blocked(
                "PROMO-IDEMPOTENCY-CONFLICT",
                f"snapshot file[{index}].size_bytes is invalid",
            )
        path = (snapshot_root / Path(relative)).resolve()
        try:
            path.relative_to(snapshot_root.resolve())
        except ValueError as exc:
            raise _blocked(
                "PROMO-IDEMPOTENCY-CONFLICT",
                f"snapshot file escapes snapshot root: {relative}",
            ) from exc
        if not path.is_file():
            raise _blocked(
                "PROMO-IDEMPOTENCY-CONFLICT",
                f"signed-off snapshot file is missing: {relative}",
            )
        content = path.read_bytes()
        if len(content) != expected_size or _sha256_bytes(content) != expected_sha:
            raise _blocked(
                "PROMO-IDEMPOTENCY-CONFLICT",
                f"signed-off snapshot file is stale: {relative}",
            )
    yaml_manifest = snapshot_root / "snapshot-manifest.yaml"
    if not yaml_manifest.is_file():
        raise _blocked(
            "PROMO-IDEMPOTENCY-CONFLICT",
            "signed-off snapshot YAML manifest is missing",
        )
    mismatches = {
        path: {"expected": sha, "actual": hashes.get(path)}
        for path, sha in expected_hashes.items()
        if hashes.get(path) != sha
    }
    unexpected = sorted(set(hashes) - set(expected_hashes))
    if mismatches or unexpected:
        raise _blocked(
            "PROMO-IDEMPOTENCY-CONFLICT",
            "signed-off snapshot entries do not match committed bindings: "
            f"mismatches={mismatches}, unexpected={unexpected}",
        )
    expected_yaml = _snapshot_yaml_bytes(cycle_id=cycle_id, entries=files)
    if yaml_manifest.read_bytes() != expected_yaml:
        raise _blocked(
            "PROMO-IDEMPOTENCY-CONFLICT",
            "signed-off snapshot YAML manifest differs from its committed JSON entries",
        )


def _completed_promotion_result(
    *,
    repo_root: Path,
    ft_root: Path,
    cycle_dir: Path,
    basis: PromotionBasis,
    basis_path: Path,
    basis_sha: str,
    canonical_path: Path,
    prior_elapsed_ms: int,
    monotonic_ns: Callable[[], int],
) -> PromotionResult | None:
    started = monotonic_ns()
    transaction_id = f"{basis.candidate.sha256[:16]}-{basis.reviewer.sha256[:8]}"
    transaction_root = cycle_dir / "promotion" / transaction_id
    snapshot_root = cycle_dir / "versions" / "signed-off"
    cycle_update = next(
        item for item in basis.state_updates if item.role == "cycle-state"
    )
    cycle_state_path = _resolve(repo_root, cycle_update.path, "cycle-state")
    cycle_state_promoted = False
    if cycle_state_path.is_file():
        try:
            cycle_state_promoted = (
                _top_level_yaml(cycle_state_path.read_text(encoding="utf-8")).get(
                    "final_promoted"
                )
                is True
            )
        except UnicodeError:
            cycle_state_promoted = True
    markers = (
        transaction_root.exists(),
        snapshot_root.exists(),
        cycle_state_promoted,
    )
    if not any(markers):
        return None
    if not all(markers):
        raise _blocked(
            "PROMO-IDEMPOTENCY-CONFLICT",
            "partial completed-promotion markers exist; refusing replay",
        )
    receipt_path = transaction_root / "promotion-receipt.json"
    metrics_path = transaction_root / "promotion-metrics.json"
    receipt = _load_json(
        receipt_path,
        missing_code="PROMO-IDEMPOTENCY-CONFLICT",
        label="committed promotion receipt",
    )
    metrics = _load_json(
        metrics_path,
        missing_code="PROMO-IDEMPOTENCY-CONFLICT",
        label="committed promotion metrics",
    )
    if receipt.get("schema_version") != 1 or receipt.get("status") != "signed-off":
        raise _blocked(
            "PROMO-IDEMPOTENCY-CONFLICT",
            "committed promotion receipt has an unsupported status",
        )
    if receipt.get("transaction_id") != transaction_id:
        raise _blocked(
            "PROMO-IDEMPOTENCY-CONFLICT",
            "committed promotion receipt transaction id does not match",
        )
    basis_binding = receipt.get("promotion_basis")
    if not isinstance(basis_binding, Mapping) or dict(basis_binding) != {
        "path": _relative(basis_path, repo_root),
        "sha256": basis_sha,
    }:
        raise _blocked(
            "PROMO-IDEMPOTENCY-CONFLICT",
            "committed promotion receipt does not bind the exact promotion basis",
        )
    expected_hash_chain = {
        "candidate_sha256": basis.candidate.sha256,
        "writer_draft_sha256": basis.writer.sha256,
        "source_basis_sha256": basis.source_basis.sha256,
        "obligation_set_sha256": basis.obligation_set.sha256,
        "prepared_package_sha256": basis.prepared_package.sha256,
        "review_result_sha256": basis.reviewer.sha256,
    }
    if receipt.get("hash_chain") != expected_hash_chain:
        raise _blocked(
            "PROMO-IDEMPOTENCY-CONFLICT",
            "committed promotion receipt hash chain does not match the current basis",
        )
    if receipt.get("review") != {"decision": "accepted", "contract_version": 4}:
        raise _blocked(
            "PROMO-IDEMPOTENCY-CONFLICT",
            "committed promotion receipt reviewer decision is invalid",
        )
    expected_gate_reports = [
        {
            "path": item.path,
            "sha256": item.sha256,
            "validator": item.validator,
        }
        for item in basis.gate_reports
    ]
    if receipt.get("gate_reports") != expected_gate_reports:
        raise _blocked(
            "PROMO-IDEMPOTENCY-CONFLICT",
            "committed promotion receipt gate bindings do not match the basis",
        )
    expected_state_updates = [
        {
            "role": item.role,
            "path": item.path,
            "before_sha256": item.before_sha256,
            "after_sha256": item.after_sha256,
            "schema": item.schema,
        }
        for item in sorted(basis.state_updates, key=lambda value: value.role)
    ]
    if receipt.get("state_updates") != expected_state_updates:
        raise _blocked(
            "PROMO-IDEMPOTENCY-CONFLICT",
            "committed promotion receipt state bindings do not match the basis",
        )
    publication = receipt.get("publication")
    if not isinstance(publication, Mapping) or any(
        publication.get(key) != value
        for key, value in {
            "candidate_path": basis.candidate.path,
            "canonical_path": basis.publication.canonical_path,
            "sha256": basis.candidate.sha256,
            "byte_identical": True,
        }.items()
    ):
        raise _blocked(
            "PROMO-IDEMPOTENCY-CONFLICT",
            "committed promotion publication binding is invalid",
        )
    if not canonical_path.is_file() or _sha256_file(canonical_path) != basis.candidate.sha256:
        raise _blocked(
            "PROMO-IDEMPOTENCY-CONFLICT",
            "canonical bytes no longer match the committed candidate",
        )
    for update in basis.state_updates:
        target = _resolve(repo_root, update.path, f"{update.role} target")
        replacement = _resolve(repo_root, update.after_path, f"{update.role} replacement")
        if (
            not target.is_file()
            or not replacement.is_file()
            or _sha256_file(target) != update.after_sha256
            or _sha256_file(replacement) != update.after_sha256
        ):
            raise _blocked(
                "PROMO-IDEMPOTENCY-CONFLICT",
                f"{update.role} is not the committed signed-off replacement",
            )
    expected_snapshot_hashes = {
        _relative(canonical_path, ft_root): basis.candidate.sha256,
    }
    for update in basis.state_updates:
        relative = _relative(
            _resolve(repo_root, update.path, f"{update.role} target"),
            ft_root,
        )
        prior = expected_snapshot_hashes.get(relative)
        if prior is not None and prior != update.after_sha256:
            raise _blocked(
                "PROMO-IDEMPOTENCY-CONFLICT",
                f"conflicting signed-off snapshot binding for {relative}",
            )
        expected_snapshot_hashes[relative] = update.after_sha256
    for name, binding in basis.final_aliases.items():
        if binding is None:
            continue
        relative = _relative(
            _resolve(repo_root, binding.path, f"final alias {name}"),
            ft_root,
        )
        prior = expected_snapshot_hashes.get(relative)
        if prior is not None and prior != binding.sha256:
            raise _blocked(
                "PROMO-IDEMPOTENCY-CONFLICT",
                f"conflicting signed-off snapshot binding for {relative}",
            )
        expected_snapshot_hashes[relative] = binding.sha256
    _verify_signed_off_snapshot(
        snapshot_root,
        cycle_id=cycle_dir.name,
        expected_hashes=expected_snapshot_hashes,
    )
    if metrics.get("schema_version") != 1 or metrics.get("transaction_id") != transaction_id:
        raise _blocked(
            "PROMO-IDEMPOTENCY-CONFLICT",
            "committed promotion metrics do not match the transaction",
        )
    stored_timings = metrics.get("phase_timings_ms")
    if not isinstance(stored_timings, Mapping) or any(
        not isinstance(value, int) or isinstance(value, bool) or value < 0
        for value in stored_timings.values()
    ):
        raise _blocked(
            "PROMO-IDEMPOTENCY-CONFLICT",
            "committed promotion metrics contain invalid timings",
        )
    production_gate = receipt.get("production_tc_gate")
    if (
        not isinstance(production_gate, Mapping)
        or production_gate.get("passed") is not True
        or production_gate.get("validator")
        not in {"production-tc-runtime-gate-v1", "production-tc-runtime-gate-v2"}
        or production_gate.get("checked_paths") != [basis.candidate.path]
        or not isinstance(production_gate.get("findings"), list)
    ):
        raise _blocked(
            "PROMO-IDEMPOTENCY-CONFLICT",
            "committed receipt has no valid candidate-bound production gate",
        )
    receipt_slo = receipt.get("slo")
    if (
        not isinstance(receipt_slo, Mapping)
        or not isinstance(receipt_slo.get("target_ms"), int)
        or isinstance(receipt_slo.get("target_ms"), bool)
        or not isinstance(receipt_slo.get("hard_ms"), int)
        or isinstance(receipt_slo.get("hard_ms"), bool)
        or receipt_slo["target_ms"] <= 0
        or receipt_slo["hard_ms"] < receipt_slo["target_ms"]
        or metrics.get("target_slo_ms") != receipt_slo["target_ms"]
        or metrics.get("hard_slo_ms") != receipt_slo["hard_ms"]
        or "total" not in stored_timings
        or metrics.get("slo_status")
        != (
            "within-target"
            if stored_timings["total"] <= receipt_slo["target_ms"]
            else "above-target"
        )
    ):
        raise _blocked(
            "PROMO-IDEMPOTENCY-CONFLICT",
            "committed promotion SLO receipt/metrics are inconsistent",
        )
    replay_ms = prior_elapsed_ms + _phase_ms(started, monotonic_ns())
    return PromotionResult(
        status="already-promoted",
        transaction_id=transaction_id,
        basis_path=basis_path,
        canonical_path=canonical_path,
        receipt_path=receipt_path,
        metrics_path=metrics_path,
        byte_identical=True,
        production_gate=dict(production_gate),
        phase_timings_ms={
            "basis_build": prior_elapsed_ms,
            "idempotency_check": max(0, replay_ms - prior_elapsed_ms),
            "total": replay_ms,
        },
    )


def promote_review_cycle(
    *,
    repo_root: Path,
    cycle_dir: Path,
    basis_path: Path | None = None,
    validate_only: bool = False,
    dry_run: bool = False,
    allow_overwrite: bool = False,
    target_slo_ms: int = TARGET_SLO_MS,
    hard_slo_ms: int = HARD_SLO_MS,
    prior_elapsed_ms: int = 0,
    include_basis_build_timing: bool = False,
    production_gate: Callable[..., ProductionTcGateResult] = validate_production_tc_draft,
    fault_injector: Callable[[str], None] | None = None,
    monotonic_ns: Callable[[], int] = time.monotonic_ns,
) -> PromotionResult:
    """Validate and atomically promote one accepted source-first review cycle.

    The function never reconstructs reviewer or handoff state from logs. Both the
    normalized review receipt and hash-bound state replacements must already exist.
    """

    if validate_only and dry_run:
        raise _blocked("PROMO-INVALID-OPTIONS", "validate_only and dry_run are mutually exclusive")
    if target_slo_ms <= 0 or hard_slo_ms < target_slo_ms:
        raise _blocked("PROMO-INVALID-OPTIONS", "SLO values must satisfy 0 < target <= hard")
    if prior_elapsed_ms < 0:
        raise _blocked("PROMO-INVALID-OPTIONS", "prior_elapsed_ms must be non-negative")
    repo_root = repo_root.resolve()
    cycle_dir = cycle_dir.resolve()
    try:
        cycle_rel = cycle_dir.relative_to(repo_root).as_posix()
    except ValueError as exc:
        raise _blocked("PROMO-UNSAFE-PATH", "cycle_dir must be inside repo_root") from exc
    if not cycle_dir.is_dir():
        raise _blocked("PROMO-MISSING-CYCLE", f"cycle directory is missing: {cycle_dir}")
    basis_path = (basis_path or cycle_dir / "promotion-basis.json").resolve()
    try:
        basis_path.relative_to(cycle_dir)
    except ValueError as exc:
        raise _blocked(
            "PROMO-UNSAFE-PATH", "promotion-basis.json must be inside cycle_dir"
        ) from exc

    total_start = monotonic_ns()
    timings: dict[str, int] = (
        {"basis_build": prior_elapsed_ms}
        if include_basis_build_timing or prior_elapsed_ms
        else {}
    )
    preliminary_basis, _, preliminary_basis_sha = _read_basis_snapshot(basis_path)
    if preliminary_basis.cycle_dir != cycle_rel:
        raise _blocked(
            "PROMO-CYCLE-MISMATCH",
            f"promotion basis cycle_dir={preliminary_basis.cycle_dir} does not match {cycle_rel}",
        )
    preliminary_canonical = _resolve(
        repo_root,
        preliminary_basis.publication.canonical_path,
        "canonical path",
    )
    preliminary_workflow_updates = [
        item
        for item in preliminary_basis.state_updates
        if item.role == "workflow-state"
    ]
    if len(preliminary_workflow_updates) != 1:
        raise _blocked(
            "PROMO-INVALID-BASIS",
            "promotion basis must bind exactly one workflow-state update",
        )
    preliminary_workflow_state = _resolve(
        repo_root,
        preliminary_workflow_updates[0].path,
        "workflow-state",
    )
    preliminary_transaction_id = (
        f"{preliminary_basis.candidate.sha256[:16]}-"
        f"{preliminary_basis.reviewer.sha256[:8]}"
    )
    _recover_expected_locks(
        repo_root=repo_root,
        cycle_dir=cycle_dir,
        canonical_path=preliminary_canonical,
        workflow_state_path=preliminary_workflow_state,
    )
    _reconcile_interrupted_promotion(
        repo_root=repo_root,
        cycle_dir=cycle_dir,
        basis=preliminary_basis,
        basis_path=basis_path,
        basis_sha=preliminary_basis_sha,
    )
    owner_token = uuid.uuid4().hex
    with _promotion_target_lock(
        repo_root,
        preliminary_canonical,
        preliminary_workflow_state,
        cycle_dir=cycle_dir,
        basis_path=basis_path,
        basis_sha=preliminary_basis_sha,
        transaction_id=preliminary_transaction_id,
        owner_token=owner_token,
    ), _promotion_lock(
        cycle_dir,
        repo_root=repo_root,
        basis_path=basis_path,
        basis_sha=preliminary_basis_sha,
        transaction_id=preliminary_transaction_id,
        owner_token=owner_token,
    ):
        phase_start = monotonic_ns()
        basis, basis_bytes, basis_sha = _read_basis_snapshot(basis_path)
        if basis.cycle_dir != cycle_rel:
            raise _blocked(
                "PROMO-CYCLE-MISMATCH",
                f"promotion basis cycle_dir={basis.cycle_dir} does not match {cycle_rel}",
            )
        locked_canonical = _resolve(
            repo_root, basis.publication.canonical_path, "canonical path"
        )
        locked_workflow_updates = [
            item for item in basis.state_updates if item.role == "workflow-state"
        ]
        if (
            len(locked_workflow_updates) != 1
            or basis_sha != preliminary_basis_sha
            or locked_canonical != preliminary_canonical
            or _resolve(
                repo_root,
                locked_workflow_updates[0].path,
                "workflow-state",
            )
            != preliminary_workflow_state
            or (
                f"{basis.candidate.sha256[:16]}-{basis.reviewer.sha256[:8]}"
                != preliminary_transaction_id
            )
        ):
            raise _blocked(
                "PROMO-BASIS-CONFLICT",
                "promotion target bindings changed while acquiring the target lock",
            )
        _call_fault(fault_injector, "after-locks-acquired")
        ft_root = _resolve(repo_root, basis.ft_root, "ft_root")
        if not ft_root.is_dir():
            raise _blocked("PROMO-MISSING-FT-ROOT", f"ft_root is missing: {ft_root}")
        try:
            cycle_dir.relative_to(ft_root)
        except ValueError as exc:
            raise _blocked("PROMO-CYCLE-MISMATCH", "cycle directory must be inside ft_root") from exc

        candidate_path, candidate_bytes = _verify_and_capture_binding(
            basis.candidate,
            repo_root,
            label="accepted candidate",
            parent=cycle_dir,
        )
        if basis.writer.sha256 != basis.candidate.sha256:
            raise _blocked(
                "PROMO-HASH-MISMATCH", "accepted candidate is not byte-identical to reviewed writer draft"
            )
        writer_path = (
            candidate_path
            if basis.writer == basis.candidate
            else _verify_binding(
                basis.writer,
                repo_root,
                label="writer draft",
                parent=cycle_dir,
            )
        )
        source_path = _verify_binding(basis.source_basis, repo_root, label="source basis")
        obligation_path = _verify_binding(basis.obligation_set, repo_root, label="obligation set")
        prepared_package_path = _verify_binding(
            basis.prepared_package,
            repo_root,
            label="prepared package",
            parent=ft_root,
        )
        try:
            prepared_package = load_prepared_package(
                prepared_package_path,
                repo_root,
            )
        except StageRuntimeError as exc:
            raise _blocked(
                "PROMO-PREPARED-PACKAGE-INVALID",
                f"prepared package is stale or invalid: {exc}",
            ) from exc
        if prepared_package.package_version != PACKAGE_VERSION:
            raise _blocked(
                "PROMO-PREPARED-PACKAGE-LEGACY",
                "promotion requires the current prepared package contract; "
                f"found v{prepared_package.package_version}, required v{PACKAGE_VERSION}",
            )
        release_status = prepared_package.release_status
        if release_status is None:
            raise _blocked(
                "PROMO-PREPARED-PACKAGE-INVALID",
                "current prepared package has no structured release_status",
            )
        if release_status.execution_dependency_registry:
            raise _blocked(
                "PROMO-BLOCKED-EXECUTION-DEPENDENCIES",
                "execution dependencies remain unresolved: "
                + ", ".join(release_status.blocking_gap_ids),
            )
        if release_status.blocking_gap_ids:
            raise _blocked(
                "PROMO-BLOCKING-SOURCE-GAPS",
                "blocking source gaps remain unresolved: "
                + ", ".join(release_status.blocking_gap_ids),
            )
        if (
            release_status.output_mode != "release"
            or not release_status.release_eligible
            or release_status.unsigned_status != "none"
        ):
            raise _blocked(
                "PROMO-PREPARED-PACKAGE-NOT-RELEASE-ELIGIBLE",
                "prepared package release_status is unsigned or draft-only",
            )
        if prepared_package.scope_slug != basis.scope_slug:
            raise _blocked(
                "PROMO-SCOPE-MISMATCH",
                "prepared package scope_slug does not match promotion basis",
            )
        package_artifacts = {
            item.kind: ArtifactBinding(item.path, item.sha256)
            for item in prepared_package.package_artifacts
        }
        if package_artifacts.get("source-evidence") != basis.source_basis:
            raise _blocked(
                "PROMO-PACKAGE-BINDING-MISMATCH",
                "promotion source_basis is not the exact prepared-package source-evidence artifact",
            )
        if package_artifacts.get("atomic-obligations") != basis.obligation_set:
            raise _blocked(
                "PROMO-PACKAGE-BINDING-MISMATCH",
                "promotion obligation_set is not the exact prepared-package obligations artifact",
            )
        reviewer_path = _verify_binding(
            basis.reviewer,
            repo_root,
            label="review-result",
            parent=cycle_dir,
            missing_code="PROMO-MISSING-REVIEW-RESULT",
        )
        review_result = _validate_review_result(
            reviewer_path,
            repo_root=repo_root,
            scope_slug=basis.scope_slug,
            writer=basis.writer,
            source_basis=basis.source_basis,
            source_basis_path=source_path,
            obligation_set=basis.obligation_set,
            obligation_path=obligation_path,
        )
        gates = _validate_gate_reports(
            basis.gate_reports, repo_root, cycle_dir, basis.candidate.sha256
        )

        canonical_path = _resolve(repo_root, basis.publication.canonical_path, "canonical path")
        production_root = (ft_root / "test-cases").resolve()
        try:
            canonical_path.relative_to(production_root)
        except ValueError as exc:
            raise _blocked(
                "PROMO-PRODUCTION-BOUNDARY",
                f"canonical path must be inside {production_root}",
            ) from exc
        if canonical_path.suffix.lower() != ".md":
            raise _blocked("PROMO-PRODUCTION-BOUNDARY", "canonical publication must be Markdown")

        alias_material: dict[str, tuple[Path, bytes]] = {}
        for name, binding in basis.final_aliases.items():
            if binding is None:
                continue
            alias_path, alias_bytes = _verify_and_capture_binding(
                binding,
                repo_root,
                label=f"final alias {name}",
                parent=ft_root,
            )
            alias_material[name] = (alias_path, alias_bytes)

        immutable_inputs: dict[Path, tuple[str, str]] = {}

        def bind_immutable(path: Path, sha256: str, label: str) -> None:
            existing = immutable_inputs.get(path)
            if existing is not None and existing[0] != sha256:
                raise _blocked(
                    "PROMO-BASIS-CONFLICT",
                    f"conflicting immutable bindings for {path}",
                )
            immutable_inputs[path] = (sha256, label)

        for path, binding, label in (
            (basis_path, ArtifactBinding(_relative(basis_path, repo_root), basis_sha), "promotion basis"),
            (candidate_path, basis.candidate, "accepted candidate"),
            (writer_path, basis.writer, "writer draft"),
            (source_path, basis.source_basis, "source basis"),
            (obligation_path, basis.obligation_set, "obligation set"),
            (prepared_package_path, basis.prepared_package, "prepared package"),
            (reviewer_path, basis.reviewer, "review result"),
        ):
            bind_immutable(path, binding.sha256, label)
        for index, gate in enumerate(basis.gate_reports):
            bind_immutable(
                _resolve(repo_root, gate.path, f"gate report[{index}]"),
                gate.sha256,
                f"gate report[{index}]",
            )
        for name, binding in basis.final_aliases.items():
            if binding is not None:
                bind_immutable(
                    alias_material[name][0],
                    binding.sha256,
                    f"final alias {name}",
                )

        def assert_immutable_inputs() -> None:
            for path, (sha256, label) in immutable_inputs.items():
                _assert_file_sha(path, sha256, label=label)

        def assert_basis_snapshot() -> None:
            if not basis_path.is_file():
                raise _blocked(
                    "PROMO-HASH-MISMATCH",
                    "promotion basis disappeared after its locked snapshot was read",
                )
            try:
                current_basis_bytes = basis_path.read_bytes()
            except OSError as exc:
                raise _blocked(
                    "PROMO-HASH-MISMATCH",
                    f"cannot recheck promotion basis before publication: {exc}",
                ) from exc
            if current_basis_bytes != basis_bytes:
                raise _blocked(
                    "PROMO-HASH-MISMATCH",
                    "promotion basis bytes changed after its locked snapshot was read",
                )

        assert_basis_snapshot()
        completed = _completed_promotion_result(
            repo_root=repo_root,
            ft_root=ft_root,
            cycle_dir=cycle_dir,
            basis=basis,
            basis_path=basis_path,
            basis_sha=basis_sha,
            canonical_path=canonical_path,
            prior_elapsed_ms=prior_elapsed_ms,
            monotonic_ns=monotonic_ns,
        )
        if completed is not None:
            return completed

        if canonical_path.exists():
            prior_sha = _sha256_file(canonical_path)
            if not allow_overwrite:
                raise _blocked(
                    "PROMO-TARGET-EXISTS",
                    f"canonical target exists and overwrite is disabled: {canonical_path}",
                )
            if basis.publication.expected_prior_sha256 is None:
                raise _blocked(
                    "PROMO-OVERWRITE-BASIS-MISSING",
                    "overwrite requires expected_prior_sha256 in promotion basis",
                )
            if prior_sha != basis.publication.expected_prior_sha256:
                raise _blocked(
                    "PROMO-HASH-MISMATCH", "canonical target changed after promotion basis was created"
                )
        elif basis.publication.expected_prior_sha256 is not None:
            raise _blocked(
                "PROMO-HASH-MISMATCH", "promotion basis expected an existing canonical target"
            )

        updates_by_role = {item.role: item for item in basis.state_updates}
        cycle_update = updates_by_role["cycle-state"]
        workflow_update = updates_by_role["workflow-state"]
        cycle_state_path = _resolve(repo_root, cycle_update.path, "cycle-state")
        if cycle_state_path != cycle_dir / "cycle-state.yaml":
            raise _blocked(
                "PROMO-STATE-PATH-MISMATCH", "cycle-state update must target cycle_dir/cycle-state.yaml"
            )
        workflow_state_path = _resolve(repo_root, workflow_update.path, "workflow-state")
        handoff_root = (ft_root / "work" / "stage-handoffs").resolve()
        try:
            workflow_state_path.relative_to(handoff_root)
        except ValueError as exc:
            raise _blocked(
                "PROMO-STATE-PATH-MISMATCH", "workflow-state update must target ft_root/work/stage-handoffs"
            ) from exc
        if workflow_state_path.name != "workflow-state.yaml":
            raise _blocked("PROMO-STATE-PATH-MISMATCH", "workflow-state target name is invalid")

        state_material: dict[str, tuple[StateUpdate, Path, bytes, bytes]] = {}
        for update in basis.state_updates:
            target = _resolve(repo_root, update.path, f"{update.role} target")
            after = _resolve(repo_root, update.after_path, f"{update.role} replacement")
            try:
                after.relative_to(cycle_dir)
            except ValueError as exc:
                raise _blocked(
                    "PROMO-STATE-PATH-MISMATCH", "state replacement artifacts must be inside cycle_dir"
                ) from exc
            before_bytes = target.read_bytes() if target.is_file() else b""
            if not target.is_file():
                raise _blocked("PROMO-MISSING-STATE", f"state target is missing: {target}")
            if _sha256_bytes(before_bytes) != update.before_sha256:
                raise _blocked(
                    "PROMO-HASH-MISMATCH", f"{update.role} changed after promotion basis was created"
                )
            after_bytes = after.read_bytes() if after.is_file() else b""
            if not after.is_file():
                raise _blocked("PROMO-MISSING-STATE", f"state replacement is missing: {after}")
            if _sha256_bytes(after_bytes) != update.after_sha256:
                raise _blocked(
                    "PROMO-HASH-MISMATCH", f"{update.role} replacement SHA-256 mismatch"
                )
            try:
                before_text = before_bytes.decode("utf-8")
                after_text = after_bytes.decode("utf-8")
            except UnicodeError as exc:
                raise _blocked("PROMO-INVALID-STATE", f"{update.role} must be UTF-8") from exc
            state_material[update.role] = (update, target, before_bytes, after_bytes)
            if update.role == "cycle-state":
                if update.schema != "codex-exec-cycle-state-v1":
                    raise _blocked(
                        "PROMO-STATE-SCHEMA-UNSUPPORTED", "cycle-state uses the wrong schema"
                    )
                _validate_cycle_state_before(before_text, basis.candidate.sha256)
                _validate_cycle_state_after(
                    after_text,
                    candidate_sha=basis.candidate.sha256,
                    canonical_path=_relative(canonical_path, ft_root),
                )
            elif update.schema != "workflow-state-final-aliases-v1":
                raise _blocked(
                    "PROMO-STATE-SCHEMA-UNSUPPORTED", "workflow-state uses the wrong schema"
                )
            else:
                _validate_workflow_state_before(
                    before_text,
                    scope_slug=basis.scope_slug,
                )

        transaction_id = (
            f"{basis.candidate.sha256[:16]}-{basis.reviewer.sha256[:8]}"
        )
        transaction_root = cycle_dir / "promotion" / transaction_id
        signed_off_snapshot = cycle_dir / "versions" / "signed-off"
        signed_off_snapshot_rel = _relative(signed_off_snapshot, ft_root)
        _validate_workflow_state_after(
            state_material["workflow-state"][3].decode("utf-8"),
            basis=basis,
            repo_root=repo_root,
            ft_root=ft_root,
            cycle_state_path=cycle_state_path,
            signed_off_snapshot_rel=signed_off_snapshot_rel,
        )
        if transaction_root.exists():
            raise _blocked(
                "PROMO-TRANSACTION-EXISTS", f"promotion transaction already exists: {transaction_root}"
            )
        if signed_off_snapshot.exists():
            raise _blocked(
                "PROMO-SNAPSHOT-EXISTS",
                f"signed-off snapshot already exists: {signed_off_snapshot}",
            )
        timings["preflight"] = _phase_ms(phase_start, monotonic_ns())

        phase_start = monotonic_ns()
        with _materialized_candidate(cycle_dir, candidate_bytes) as materialized_candidate:
            production_result = production_gate(draft_path=materialized_candidate)
        timings["production_tc_gate"] = _phase_ms(phase_start, monotonic_ns())
        if not production_result.passed:
            raise _blocked(
                "PROMO-PRODUCTION-GATE-FAILED",
                f"production_tc_gate found {len(production_result.findings)} blocking findings",
            )
        production_payload = production_result.as_dict()
        production_payload["checked_paths"] = [basis.candidate.path]

        validation_elapsed_ms = prior_elapsed_ms + _phase_ms(
            total_start, monotonic_ns()
        )
        if validation_elapsed_ms > hard_slo_ms:
            raise _blocked(
                "PROMO-HARD-SLO-EXCEEDED",
                (
                    f"promotion validation took {validation_elapsed_ms} ms, "
                    f"exceeding hard SLO {hard_slo_ms} ms"
                ),
            )

        if validate_only or dry_run:
            timings["total"] = prior_elapsed_ms + _phase_ms(
                total_start, monotonic_ns()
            )
            return PromotionResult(
                status="validated" if validate_only else "dry-run-passed",
                transaction_id=transaction_id,
                basis_path=basis_path,
                canonical_path=canonical_path,
                receipt_path=None,
                metrics_path=None,
                byte_identical=True,
                production_gate=production_payload,
                phase_timings_ms=timings,
            )

        prior_target_bytes = canonical_path.read_bytes() if canonical_path.is_file() else None
        journal_root = _recovery_journal_root(cycle_dir, transaction_id)
        prepare_root = _recovery_prepare_root(
            cycle_dir,
            transaction_id,
            owner_token,
        )
        if journal_root.exists() or prepare_root.exists():
            raise _blocked(
                "PROMO-TEMP-CONFLICT",
                "promotion recovery journal/staging root already exists",
            )
        mutated = False
        mutated_state_targets: set[Path] = set()
        transaction_committed = False
        snapshot_committed = False
        try:
            phase_start = monotonic_ns()
            before_entries: list[dict[str, Any]] = []
            if prior_target_bytes is not None:
                before_entries.append(
                    _snapshot_write(
                        prepare_root / "before",
                        f"repo/{basis.publication.canonical_path}",
                        prior_target_bytes,
                    )
                )
            else:
                before_entries.append(
                    {"path": f"repo/{basis.publication.canonical_path}", "status": "absent"}
                )
            for update, _, before_bytes, _ in state_material.values():
                before_entries.append(
                    _snapshot_write(
                        prepare_root / "before", f"repo/{update.path}", before_bytes
                    )
                )
            _write_json(
                prepare_root / "before" / "snapshot-manifest.json",
                {
                    "schema_version": 1,
                    "snapshot_id": "prior-canonical-state",
                    "files": sorted(before_entries, key=lambda item: item["path"]),
                },
            )
            _write_json(
                prepare_root / "recovery-journal.json",
                _recovery_journal_payload(
                    basis=basis,
                    basis_path=basis_path,
                    basis_sha=basis_sha,
                    repo_root=repo_root,
                    transaction_id=transaction_id,
                    owner_token=owner_token,
                ),
            )
            journal_root.parent.mkdir(parents=True, exist_ok=True)
            os.replace(prepare_root, journal_root)
            _call_fault(fault_injector, "after-journal-prepared")
            timings["snapshot_before"] = _phase_ms(phase_start, monotonic_ns())

            phase_start = monotonic_ns()
            for update, target, _, _ in state_material.values():
                _assert_file_sha(
                    target,
                    update.before_sha256,
                    label=f"{update.role} target",
                )
                _assert_file_sha(
                    _resolve(repo_root, update.after_path, update.role),
                    update.after_sha256,
                    label=f"{update.role} replacement",
                )
            if prior_target_bytes is None:
                if canonical_path.exists():
                    raise _blocked(
                        "PROMO-HASH-MISMATCH",
                        "canonical target appeared after promotion preflight",
                    )
            elif (
                not canonical_path.is_file()
                or _sha256_file(canonical_path)
                != basis.publication.expected_prior_sha256
            ):
                raise _blocked(
                    "PROMO-HASH-MISMATCH",
                    "canonical target changed after promotion preflight",
                )
            assert_basis_snapshot()
            assert_immutable_inputs()
            _atomic_write(canonical_path, candidate_bytes)
            mutated = True
            _call_fault(fault_injector, "after-publication")
            if _sha256_file(canonical_path) != basis.candidate.sha256:
                raise _blocked(
                    "PROMO-BYTE-IDENTITY-FAILED", "published canonical bytes differ from accepted candidate"
                )
            timings["publication"] = _phase_ms(phase_start, monotonic_ns())

            phase_start = monotonic_ns()
            final_snapshot_temp = journal_root / "final-snapshot"
            after_entries = [
                _snapshot_write(
                    final_snapshot_temp,
                    _relative(canonical_path, ft_root),
                    candidate_bytes,
                )
            ]
            for update, _, _, after_bytes in state_material.values():
                after_entries.append(
                    _snapshot_write(
                        final_snapshot_temp,
                        _relative(_resolve(repo_root, update.path, update.role), ft_root),
                        after_bytes,
                    )
                )
            for name, binding in basis.final_aliases.items():
                if binding is None:
                    continue
                alias_path, alias_bytes = alias_material[name]
                after_entries.append(
                    _snapshot_write(
                        final_snapshot_temp,
                        _relative(alias_path, ft_root),
                        alias_bytes,
                    )
                )
            canonical_snapshot_entries = [
                item
                for item in after_entries
                if item["path"] == _relative(canonical_path, ft_root)
            ]
            if (
                len(canonical_snapshot_entries) != 1
                or canonical_snapshot_entries[0]["sha256"]
                != basis.candidate.sha256
            ):
                raise _blocked(
                    "PROMO-SNAPSHOT-HASH-MISMATCH",
                    "signed-off snapshot canonical entry is not bound to the accepted candidate",
                )
            snapshot_manifest = {
                "schema_version": 1,
                "transaction_id": transaction_id,
                "snapshot_id": "signed-off",
                "files": sorted(after_entries, key=lambda item: item["path"]),
            }
            _write_json(final_snapshot_temp / "snapshot-manifest.json", snapshot_manifest)
            _write_snapshot_yaml(
                final_snapshot_temp / "snapshot-manifest.yaml",
                cycle_id=cycle_dir.name,
                entries=after_entries,
            )

            receipt = {
                "schema_version": 1,
                "status": "signed-off",
                "transaction_id": transaction_id,
                "promotion_basis": {
                    "path": _relative(basis_path, repo_root),
                    "sha256": basis_sha,
                },
                "hash_chain": {
                    "candidate_sha256": basis.candidate.sha256,
                    "writer_draft_sha256": basis.writer.sha256,
                    "source_basis_sha256": basis.source_basis.sha256,
                    "obligation_set_sha256": basis.obligation_set.sha256,
                    "prepared_package_sha256": basis.prepared_package.sha256,
                    "review_result_sha256": basis.reviewer.sha256,
                },
                "review": {
                    "decision": review_result["decision"],
                    "contract_version": review_result["contract_version"],
                },
                "gate_reports": list(gates),
                "production_tc_gate": production_payload,
                "publication": {
                    "candidate_path": basis.candidate.path,
                    "canonical_path": basis.publication.canonical_path,
                    "sha256": basis.candidate.sha256,
                    "byte_identical": True,
                },
                "state_updates": [
                    {
                        "role": item.role,
                        "path": item.path,
                        "before_sha256": item.before_sha256,
                        "after_sha256": item.after_sha256,
                        "schema": item.schema,
                    }
                    for item in sorted(basis.state_updates, key=lambda value: value.role)
                ],
                "prior_snapshot_manifest": (
                    f"{_relative(transaction_root, ft_root)}/before/snapshot-manifest.json"
                ),
                "signed_off_snapshot_manifest": (
                    f"{signed_off_snapshot_rel}/snapshot-manifest.yaml"
                ),
                "slo": {"target_ms": target_slo_ms, "hard_ms": hard_slo_ms},
            }
            _write_json(journal_root / "promotion-receipt.json", receipt)
            timings["snapshot_and_receipt"] = _phase_ms(phase_start, monotonic_ns())

            phase_start = monotonic_ns()
            for update, target, _, _ in state_material.values():
                _assert_file_sha(
                    target,
                    update.before_sha256,
                    label=f"{update.role} target",
                )
                _assert_file_sha(
                    _resolve(repo_root, update.after_path, update.role),
                    update.after_sha256,
                    label=f"{update.role} replacement",
                )
            for update, target, _, after_bytes in (
                state_material["cycle-state"],
                state_material["workflow-state"],
            ):
                _atomic_write(target, after_bytes)
                mutated_state_targets.add(target)
                _call_fault(fault_injector, f"after-{update.role}-update")
            _call_fault(fault_injector, "after-state-updates")
            timings["state_updates"] = _phase_ms(phase_start, monotonic_ns())

            total_ms = prior_elapsed_ms + _phase_ms(total_start, monotonic_ns())
            timings["total"] = total_ms
            if total_ms > hard_slo_ms:
                raise _blocked(
                    "PROMO-HARD-SLO-EXCEEDED",
                    f"promotion took {total_ms} ms, exceeding hard SLO {hard_slo_ms} ms",
                )
            metrics = {
                "schema_version": 1,
                "transaction_id": transaction_id,
                "phase_timings_ms": timings,
                "target_slo_ms": target_slo_ms,
                "hard_slo_ms": hard_slo_ms,
                "slo_status": "within-target" if total_ms <= target_slo_ms else "above-target",
            }
            _write_json(journal_root / "promotion-metrics.json", metrics)
            _call_fault(fault_injector, "after-metrics")

            assert_immutable_inputs()
            transaction_root.parent.mkdir(parents=True, exist_ok=True)
            os.replace(journal_root, transaction_root)
            transaction_committed = True
            _call_fault(fault_injector, "after-transaction-commit")
            signed_off_snapshot.parent.mkdir(parents=True, exist_ok=True)
            os.replace(transaction_root / "final-snapshot", signed_off_snapshot)
            snapshot_committed = True
            _call_fault(fault_injector, "after-snapshot-commit")
            _assert_file_sha(
                canonical_path,
                basis.candidate.sha256,
                label="published canonical target",
            )
            for update, target, _, _ in state_material.values():
                _assert_file_sha(
                    target,
                    update.after_sha256,
                    label=f"committed {update.role} target",
                )
            metrics_path = transaction_root / "promotion-metrics.json"
            return PromotionResult(
                status="promoted",
                transaction_id=transaction_id,
                basis_path=basis_path,
                canonical_path=canonical_path,
                receipt_path=transaction_root / "promotion-receipt.json",
                metrics_path=metrics_path,
                byte_identical=_sha256_file(canonical_path) == basis.candidate.sha256,
                production_gate=production_payload,
                phase_timings_ms=timings,
            )
        except Exception:
            if mutated:
                canonical_state = _recovery_target_state(
                    canonical_path,
                    before_sha=basis.publication.expected_prior_sha256,
                    after_sha=basis.candidate.sha256,
                )
                if canonical_state in {"after", "both"}:
                    if prior_target_bytes is None:
                        canonical_path.unlink(missing_ok=True)
                    else:
                        _atomic_write(canonical_path, prior_target_bytes)
                for update, target, before_bytes, _ in state_material.values():
                    state = _recovery_target_state(
                        target,
                        before_sha=update.before_sha256,
                        after_sha=update.after_sha256,
                    )
                    if target in mutated_state_targets and state in {"after", "both"}:
                        _atomic_write(target, before_bytes)
            if transaction_committed:
                shutil.rmtree(transaction_root, ignore_errors=True)
                try:
                    transaction_root.parent.rmdir()
                except OSError:
                    pass
            if snapshot_committed:
                shutil.rmtree(signed_off_snapshot, ignore_errors=True)
                try:
                    signed_off_snapshot.parent.rmdir()
                except OSError:
                    pass
            shutil.rmtree(journal_root, ignore_errors=True)
            raise
        finally:
            shutil.rmtree(prepare_root, ignore_errors=True)
            try:
                prepare_root.parent.rmdir()
            except OSError:
                pass


def build_validate_promote_review_cycle(
    *,
    repo_root: Path,
    cycle_dir: Path,
    seed_path: Path | None = None,
    basis_path: Path | None = None,
    cycle_state_replacement_path: Path | None = None,
    workflow_state_replacement_path: Path | None = None,
    validate_only: bool = False,
    dry_run: bool = False,
    allow_overwrite: bool = False,
    target_slo_ms: int = TARGET_SLO_MS,
    hard_slo_ms: int = HARD_SLO_MS,
    production_gate: Callable[..., ProductionTcGateResult] = validate_production_tc_draft,
    fault_injector: Callable[[str], None] | None = None,
    monotonic_ns: Callable[[], int] = time.monotonic_ns,
) -> PromotionResult:
    """Build, revalidate and promote one cycle through a single deterministic call."""

    resolved_repo_root = repo_root.resolve()
    resolved_cycle_dir = cycle_dir.resolve()
    resolved_basis_path = (
        basis_path or resolved_cycle_dir / "promotion-basis.json"
    ).resolve()
    if resolved_basis_path.is_file():
        existing_basis, _, existing_basis_sha = _read_basis_snapshot(
            resolved_basis_path
        )
        expected_cycle = _relative(resolved_cycle_dir, resolved_repo_root)
        if existing_basis.cycle_dir != expected_cycle:
            raise _blocked(
                "PROMO-CYCLE-MISMATCH",
                "existing promotion basis belongs to another cycle",
            )
        existing_canonical = _resolve(
            resolved_repo_root,
            existing_basis.publication.canonical_path,
            "existing canonical path",
        )
        existing_workflow_update = next(
            item
            for item in existing_basis.state_updates
            if item.role == "workflow-state"
        )
        existing_workflow = _resolve(
            resolved_repo_root,
            existing_workflow_update.path,
            "existing workflow-state",
        )
        _recover_expected_locks(
            repo_root=resolved_repo_root,
            cycle_dir=resolved_cycle_dir,
            canonical_path=existing_canonical,
            workflow_state_path=existing_workflow,
        )
        _reconcile_interrupted_promotion(
            repo_root=resolved_repo_root,
            cycle_dir=resolved_cycle_dir,
            basis=existing_basis,
            basis_path=resolved_basis_path,
            basis_sha=existing_basis_sha,
        )

    build = build_full_promotion_basis(
        repo_root=resolved_repo_root,
        cycle_dir=resolved_cycle_dir,
        seed_path=seed_path,
        basis_path=resolved_basis_path,
        cycle_state_replacement_path=cycle_state_replacement_path,
        workflow_state_replacement_path=workflow_state_replacement_path,
        monotonic_ns=monotonic_ns,
    )
    return promote_review_cycle(
        repo_root=repo_root,
        cycle_dir=cycle_dir,
        basis_path=build.basis_path,
        validate_only=validate_only,
        dry_run=dry_run,
        allow_overwrite=allow_overwrite,
        target_slo_ms=target_slo_ms,
        hard_slo_ms=hard_slo_ms,
        prior_elapsed_ms=build.build_duration_ms,
        include_basis_build_timing=True,
        production_gate=production_gate,
        fault_injector=fault_injector,
        monotonic_ns=monotonic_ns,
    )
