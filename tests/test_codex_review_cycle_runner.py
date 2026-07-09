from __future__ import annotations

import json
import importlib.util
import os
import subprocess
import sys
import tempfile
import types
import unittest
from pathlib import Path


ROOT_DIR = Path(__file__).resolve().parents[1]
RUNNER = ROOT_DIR / "scripts" / "codex_review_cycle_runner.py"
TERMINAL_GATE_EVAL = ROOT_DIR / "evals" / "terminal-scoped-validator-gate-regression.md"
WAIVER_QUALITY_EVAL = ROOT_DIR / "evals" / "terminal-validator-waiver-quality-regression.md"
WRITER_SYNC_FORMAT_EVAL = ROOT_DIR / "evals" / "writer-cross-artifact-sync-and-format-regression.md"


def load_runner_module():
    spec = importlib.util.spec_from_file_location("codex_review_cycle_runner_under_test", RUNNER)
    assert spec is not None
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


class CodexReviewCycleRunnerTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tmp = tempfile.TemporaryDirectory()
        self.root = Path(self.tmp.name)
        self.ft_root = self.root / "fts" / "demo-ft"
        self.cycle_dir = self.ft_root / "work" / "review-cycles" / "section-scope"
        self.prompt_dir = self.cycle_dir / "prompts"
        self.test_design_dir = self.ft_root / "work" / "test-design" / "section-scope"
        self.test_case = self.ft_root / "test-cases" / "1-section-scope.md"
        self.state_path = self.cycle_dir / "cycle-state.yaml"

        self.prompt_dir.mkdir(parents=True)
        self.test_design_dir.mkdir(parents=True)
        self.test_case.parent.mkdir(parents=True)
        self.test_case.write_text("# Test cases\n\n## TC-001\n", encoding="utf-8")
        (self.prompt_dir / "next.md").write_text("Run next session", encoding="utf-8")
        (self.test_design_dir / "traceability.md").write_text("| atom | tc |\n", encoding="utf-8")
        (self.cycle_dir / "codex-session-map.yaml").write_text("sessions: []\n", encoding="utf-8")

    def tearDown(self) -> None:
        self.tmp.cleanup()

    def write_state(
        self,
        *,
        status: str = "writer-draft-ready",
        semantic_round: int = 0,
        prompt: str = "work/review-cycles/section-scope/prompts/next.md",
    ) -> None:
        self.state_path.write_text(
            "\n".join(
                [
                    "cycle_id: cycle-demo",
                    "ft_slug: demo-ft",
                    "scope_slug: section-scope",
                    "canonical_test_cases: test-cases/1-section-scope.md",
                    "test_design_dir: work/test-design/section-scope",
                    "current_stage: writer-r1",
                    f"stage_status: {status}",
                    f"semantic_round: {semantic_round}",
                    "max_semantic_rounds: 2",
                    f"active_transition_prompt: {prompt}",
                    "",
                ]
            ),
            encoding="utf-8",
        )

    def write_valid_structure_preflight_artifacts(self) -> None:
        self.test_case.write_text(
            "\n".join(
                [
                    "# Test cases",
                    "",
                    "| metric | value |",
                    "| --- | --- |",
                    "| executable_test_cases | TC-001..TC-002 |",
                    "",
                    "## TC-001 Create document",
                    "",
                    "**Title:** Create document",
                    "**Type:** positive",
                    "**Priority:** high",
                    "**package_id:** WP-01",
                    "**Traceability:** REQ-001",
                    "",
                    "### Preconditions",
                    "- User is authorized.",
                    "",
                    "### Test Data",
                    "- Valid document.",
                    "",
                    "### Steps",
                    "1. Create the document.",
                    "",
                    "### Expected Result",
                    "Document is created.",
                    "",
                    "### Postconditions",
                    "- Document can be removed.",
                    "",
                    "## TC-002 Update document",
                    "",
                    "**Title:** Update document",
                    "**Type:** positive",
                    "**Priority:** medium",
                    "**package_id:** WP-01",
                    "**Traceability:** REQ-002",
                    "",
                    "### Preconditions",
                    "- Document exists.",
                    "",
                    "### Test Data",
                    "- Valid update payload.",
                    "",
                    "### Steps",
                    "1. Update the document.",
                    "",
                    "### Expected Result",
                    "Document is updated.",
                    "",
                    "### Postconditions",
                    "- Document remains available.",
                    "",
                ]
            ),
            encoding="utf-8",
        )
        for file_name in (
            "writer-quality-gate.md",
            "test-design-review.md",
            "package-test-design-plan.md",
            "atomic-requirements-ledger.md",
            "coverage-map.md",
            "test-design-decision-table.md",
            "test-design-applicability-matrix.md",
            "fixture-catalog.md",
        ):
            (self.test_design_dir / file_name).write_text(
                "\n".join(
                    [
                        f"# {file_name}",
                        "",
                        "| item | status |",
                        "| --- | --- |",
                        "| structure | pass |",
                        "",
                    ]
                ),
                encoding="utf-8",
            )
        self.write_scoped_validator_profile()

    def bounded_semantic_response(
        self,
        *,
        findings: list[dict[str, object]] | None = None,
        recommended_stage_status: str = "format-review-ready",
    ) -> str:
        payload = {
            "coverage_summary": {
                "status": "complete" if not findings else "revision-needed",
                "covered_requirements": 2,
                "blocked_requirements": 0 if not findings else 1,
                "notes": "Bounded semantic review summary.",
            },
            "traceability_matrix_rows": [
                {
                    "requirement_id": "REQ-001",
                    "requirement_summary": "Create document.",
                    "covered_by_tc_ids": "TC-001",
                    "coverage_status": "covered",
                    "evidence": "TC-001 expected result.",
                    "finding_ids": "",
                },
                {
                    "requirement_id": "REQ-002",
                    "requirement_summary": "Update document.",
                    "covered_by_tc_ids": "TC-002",
                    "coverage_status": "covered",
                    "evidence": "TC-002 expected result.",
                    "finding_ids": findings[0]["id"] if findings else "",
                },
            ],
            "findings": findings or [],
            "human_summary": "No blocking semantic findings." if not findings else "Blocking semantic findings found.",
            "recommended_stage_status": recommended_stage_status,
        }
        return "```json\n" + json.dumps(payload, ensure_ascii=False) + "\n```"

    def bounded_reviewer_response(
        self,
        *,
        findings: list[dict[str, object]] | None = None,
        recommended_stage_status: str,
    ) -> str:
        payload = {
            "findings": findings or [],
            "human_summary": "No blocking reviewer findings." if not findings else "Blocking reviewer findings found.",
            "recommended_stage_status": recommended_stage_status,
        }
        return "```json\n" + json.dumps(payload, ensure_ascii=False) + "\n```"

    def install_fake_codex(
        self,
        runner,
        *,
        final_response: str,
        turn_status: str = "completed",
        turn_id: str = "fake-turn-bounded",
        thread_id: str = "fake-thread-bounded",
    ):
        class FakeSandbox:
            read_only = "read_only"
            workspace_write = "workspace_write"
            full_access = "full_access"

        class FakeApprovalMode:
            auto_review = "auto_review"
            deny_all = "deny_all"

        class FakeTurn:
            def __init__(self) -> None:
                self.id = turn_id
                self.status = turn_status
                self.duration_ms = 123
                self.final_response = final_response

        class FakeThread:
            id = thread_id

            def set_name(self, name: str) -> None:
                self.name = name

            def run(self, *args, **kwargs):
                return FakeTurn()

        class FakeCodex:
            def thread_start(self, *args, **kwargs):
                self.thread_start_kwargs = kwargs
                return FakeThread()

            def close(self) -> None:
                pass

        previous_module = sys.modules.get("openai_codex")
        sys.modules["openai_codex"] = type(
            "FakeOpenAICodex",
            (),
            {
                "Codex": FakeCodex,
                "Sandbox": FakeSandbox,
                "ApprovalMode": FakeApprovalMode,
            },
        )
        return previous_module

    def restore_openai_codex_module(self, previous_module) -> None:
        if previous_module is None:
            sys.modules.pop("openai_codex", None)
        else:
            sys.modules["openai_codex"] = previous_module

    def write_scoped_validator_profile(
        self,
        *,
        command: str = "python scripts/validate_agent_artifacts.py --root fts/demo-ft --json",
        scope_slug: str = "section-scope",
        current_stage: str = "writer-r1",
        unresolved_warning_error_count: int = 0,
    ) -> Path:
        output_dir = self.cycle_dir / "outputs"
        output_dir.mkdir(parents=True, exist_ok=True)
        profile_path = output_dir / "scoped-validator-profile.writer-r1.json"
        profile_path.write_text(
            json.dumps(
                {
                    "version": 1,
                    "generated_by": "codex_review_cycle_runner",
                    "command": command,
                    "scope_slug": scope_slug,
                    "canonical_test_cases": "test-cases/1-section-scope.md",
                    "test_design_dir": "work/test-design/section-scope",
                    "current_stage": current_stage,
                    "current_scope_findings": [],
                    "unresolved_warning_error_count": unresolved_warning_error_count,
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )
        return profile_path

    def run_runner(self, *args: str) -> subprocess.CompletedProcess[str]:
        return subprocess.run(
            [sys.executable, str(RUNNER), *args],
            cwd=ROOT_DIR,
            capture_output=True,
            text=True,
            check=False,
        )

    def test_validate_reports_next_structure_preflight_session(self) -> None:
        self.write_state()

        result = self.run_runner("validate", "--state", str(self.state_path))

        self.assertEqual(result.returncode, 0, result.stderr)
        payload = json.loads(result.stdout)
        self.assertTrue(payload["valid"])
        self.assertEqual("reviewer.structure_preflight", payload["next"]["scenario"])
        self.assertEqual("structure-preflight-r1", payload["next"]["stage"])
        self.assertEqual("read_only", payload["next"]["sandbox_policy"])
        self.assertEqual(
            "reviewer.structure_preflight",
            payload["next"]["instruction_context"]["scenario"],
        )
        context_paths = {
            item["path"] for item in payload["next"]["instruction_context"]["files"]
        }
        self.assertIn("skills/ft-test-case-reviewer/SKILL.md", context_paths)
        self.assertIn("references/agent/session-based-review-cycle-format.md", context_paths)

    def test_start_dry_run_selects_semantic_revision_writer(self) -> None:
        self.write_state(status="semantic-revision-needed", semantic_round=1)

        result = self.run_runner("start", "--state", str(self.state_path), "--dry-run")

        self.assertEqual(result.returncode, 0, result.stderr)
        payload = json.loads(result.stdout)
        self.assertEqual("start-session", payload["action"])
        self.assertEqual("writer", payload["role"])
        self.assertEqual("writer-r2", payload["stage"])
        self.assertEqual("writer.session_semantic_revision", payload["scenario"])
        self.assertEqual("workspace_write", payload["sandbox_policy"])
        self.assertEqual(
            "writer.session_semantic_revision",
            payload["instruction_context"]["scenario"],
        )

    def test_semantic_review_next_session_uses_read_only_sandbox(self) -> None:
        self.write_state(status="semantic-review-ready", semantic_round=1)
        runner = load_runner_module()

        next_session = runner.next_session_for_state(runner.load_simple_yaml(self.state_path))

        self.assertEqual("semantic-review-r1", next_session.stage)
        self.assertEqual("reviewer.semantic_traceability_test_design", next_session.scenario)
        self.assertEqual("read_only", next_session.sandbox_policy)

    def test_session_sandbox_policy_matrix_is_explicit(self) -> None:
        runner = load_runner_module()

        self.assertEqual(
            {
                "reviewer.scope_gap_review",
                "reviewer.structure_preflight",
                "reviewer.semantic_traceability_test_design",
                "reviewer.structure_format_final",
                "reviewer.semantic_regression",
            },
            runner.REVIEWER_READ_ONLY_SCENARIOS,
        )
        self.assertEqual(
            {
                "reviewer.scope_gap_review",
                "reviewer.structure_format_final",
                "reviewer.semantic_regression",
            },
            runner.BOUNDED_REVIEWER_SCENARIOS,
        )

        cases = [
            ("scope-ready-for-gap-review", 0, "reviewer", "reviewer.scope_gap_review", "read_only"),
            ("scope-gap-review-passed", 0, "writer", "writer.session_initial_draft", "workspace_write"),
            ("scope-ready-for-writer", 0, "writer", "writer.session_initial_draft", "workspace_write"),
            ("writer-draft-ready", 0, "reviewer", "reviewer.structure_preflight", "read_only"),
            ("structure-preflight-blocked", 0, "writer", "writer.remediation.style", "workspace_write"),
            ("semantic-review-ready", 1, "reviewer", "reviewer.semantic_traceability_test_design", "read_only"),
            ("semantic-revision-needed", 1, "writer", "writer.session_semantic_revision", "workspace_write"),
            ("semantic-review-passed", 1, "reviewer", "reviewer.structure_format_final", "read_only"),
            ("format-review-ready", 1, "reviewer", "reviewer.structure_format_final", "read_only"),
            ("format-revision-needed", 1, "writer", "writer.session_format_revision", "workspace_write"),
            ("final-regression-ready", 1, "reviewer", "reviewer.semantic_regression", "read_only"),
        ]

        for status, semantic_round, role, scenario, sandbox_policy in cases:
            with self.subTest(status=status):
                self.write_state(status=status, semantic_round=semantic_round)
                next_session = runner.next_session_for_state(runner.load_simple_yaml(self.state_path))
                self.assertEqual(role, next_session.role)
                self.assertEqual(scenario, next_session.scenario)
                self.assertEqual(sandbox_policy, next_session.sandbox_policy)

    def test_completion_manifest_marks_reviewer_read_only_without_write_exception(self) -> None:
        self.write_state(status="final-regression-ready", semantic_round=1)
        runner = load_runner_module()
        state_before = runner.load_simple_yaml(self.state_path)
        state_after = dict(state_before)
        state_after["stage_status"] = "signed-off"
        final_response = self.cycle_dir / "outputs" / "semantic-regression-final-response.md"
        final_response.parent.mkdir(exist_ok=True)
        final_response.write_text("No regression findings.", encoding="utf-8")
        next_session = runner.NextSession(
            stage="semantic-regression-final",
            role="reviewer",
            scenario="reviewer.semantic_regression",
            prompt_path="work/review-cycles/section-scope/prompts/next.md",
            sandbox_policy="read_only",
        )

        manifest_path = runner.write_completion_manifest(
            self.state_path,
            state_before=state_before,
            state_after=state_after,
            next_session=next_session,
            thread_id="thread-legacy-reviewer",
            turn_id="turn-legacy-reviewer",
            turn_status="completed",
            session_status="completed",
            state_advanced=True,
            input_snapshot="work/review-cycles/section-scope/versions/before",
            output_snapshot="work/review-cycles/section-scope/versions/after",
            final_response=final_response,
            started_at_epoch=1,
            completed_at_epoch=2,
            duration_ms=1000,
        )

        manifest = runner.load_simple_yaml(manifest_path)
        self.assertEqual("read_only", manifest["sandbox_policy"])
        self.assertFalse(manifest["reviewer_write_exception"])

    def test_diagnose_sdk_turn_dry_run_exports_prompt_without_cycle_mutation(self) -> None:
        self.write_state()
        state_before = self.state_path.read_text(encoding="utf-8")
        session_map = self.cycle_dir / "codex-session-map.yaml"
        session_map_before = session_map.read_text(encoding="utf-8")
        output_dir = self.root / "sdk-diagnostic"

        result = self.run_runner(
            "diagnose-sdk-turn",
            "--state",
            str(self.state_path),
            "--prompt-source",
            "runner-composed",
            "--dry-run",
            "--timeout-seconds",
            "7",
            "--output-dir",
            str(output_dir),
        )

        self.assertEqual(result.returncode, 0, result.stderr)
        payload = json.loads(result.stdout)
        self.assertEqual("rendered-sdk-diagnostic-prompt", payload["action"])
        self.assertEqual("rendered-prompt", payload["status"])
        self.assertEqual("runner-composed", payload["prompt_source"])
        self.assertEqual("structure-preflight-r1", payload["stage"])
        self.assertEqual("reviewer.structure_preflight", payload["scenario"])
        self.assertEqual(state_before, self.state_path.read_text(encoding="utf-8"))
        self.assertEqual(session_map_before, session_map.read_text(encoding="utf-8"))
        self.assertFalse((self.cycle_dir / "runner.lock.yaml").exists())
        self.assertFalse((self.cycle_dir / "runner-events.ndjson").exists())
        self.assertFalse((self.cycle_dir / "outputs").exists())
        self.assertFalse((self.cycle_dir / "snapshots").exists())
        prompt_text = (output_dir / "prompt.md").read_text(encoding="utf-8")
        self.assertIn("Standalone SDK diagnostic safety contract.", prompt_text)
        self.assertIn("Session-based structure-preflight stage.", prompt_text)
        self.assertIn("Do not edit codex-session-map.yaml", prompt_text)
        self.assertIn("Slim runtime contract:", prompt_text)
        self.assertIn("Do not recursively read directories", prompt_text)
        self.assertIn("validator-report*.json", prompt_text)
        run_payload = json.loads((output_dir / "run.json").read_text(encoding="utf-8"))
        self.assertEqual(payload["prompt_sha256"], run_payload["prompt_sha256"])
        self.assertEqual("reviewer", run_payload["role"])
        self.assertEqual("structure-preflight-r1", run_payload["stage"])
        self.assertEqual("section-scope", run_payload["scope_slug"])
        self.assertEqual("0", str(run_payload["current_round"]))
        self.assertIn("session_id", run_payload)
        events = (output_dir / "events.ndjson").read_text(encoding="utf-8")
        self.assertIn("prompt_rendered", events)

    def test_diagnose_sdk_turn_dry_run_can_export_blocked_current_stage_prompt(self) -> None:
        self.write_state(status="blocked-input")
        self.state_path.write_text(
            self.state_path.read_text(encoding="utf-8").replace(
                "current_stage: writer-r1",
                "current_stage: structure-preflight-r1",
            )
            + "blocking_reasons:\n"
            + "- infrastructure timeout\n",
            encoding="utf-8",
        )
        state_before = self.state_path.read_text(encoding="utf-8")
        output_dir = self.root / "blocked-sdk-diagnostic"

        result = self.run_runner(
            "diagnose-sdk-turn",
            "--state",
            str(self.state_path),
            "--prompt-source",
            "runner-composed",
            "--dry-run",
            "--output-dir",
            str(output_dir),
        )

        self.assertEqual(result.returncode, 0, result.stderr)
        payload = json.loads(result.stdout)
        self.assertEqual("structure-preflight-r1", payload["stage"])
        self.assertEqual("reviewer.structure_preflight", payload["scenario"])
        self.assertEqual(state_before, self.state_path.read_text(encoding="utf-8"))
        self.assertFalse((self.cycle_dir / "runner.lock.yaml").exists())
        self.assertFalse((self.cycle_dir / "runner-events.ndjson").exists())
        self.assertFalse((self.cycle_dir / "outputs").exists())
        prompt_text = (output_dir / "prompt.md").read_text(encoding="utf-8")
        self.assertIn("stage: structure-preflight-r1", prompt_text)
        self.assertIn("Standalone SDK diagnostic safety contract.", prompt_text)
        self.assertIn("Slim runtime contract:", prompt_text)
        self.assertIn("Do not recursively read directories", prompt_text)

    def test_validate_accepts_unindented_top_level_list_items(self) -> None:
        self.write_state(status="semantic-revision-needed", semantic_round=1)
        self.state_path.write_text(
            self.state_path.read_text(encoding="utf-8")
            + "\n".join(
                [
                    "latest_artifacts:",
                    "- test-cases/1-section-scope.md",
                    "blocking_reasons:",
                    "- 'FINDING-001: generic expected result'",
                    "",
                ]
            ),
            encoding="utf-8",
        )

        result = self.run_runner("validate", "--state", str(self.state_path))

        self.assertEqual(result.returncode, 0, result.stderr)
        payload = json.loads(result.stdout)
        self.assertTrue(payload["valid"])
        self.assertEqual("writer-r2", payload["next"]["stage"])

    def test_load_simple_yaml_preserves_colon_string_list_items(self) -> None:
        self.write_state(status="blocked-input", semantic_round=0)
        self.state_path.write_text(
            self.state_path.read_text(encoding="utf-8")
            + "\n".join(
                [
                    "open_questions:",
                    "- GAP-001: exact SMS notification text is not confirmed.",
                    "accepted_risks:",
                    "- source-quality-oversized-blocks: selected rows are split defensively.",
                    "blocking_reasons:",
                    "- 'structure-preflight-r1: Codex SDK turn timed out after 600 seconds'",
                    "",
                ]
            ),
            encoding="utf-8",
        )
        runner = load_runner_module()

        state = runner.load_simple_yaml(self.state_path)

        self.assertEqual(
            ["GAP-001: exact SMS notification text is not confirmed."],
            state["open_questions"],
        )
        self.assertEqual(
            ["source-quality-oversized-blocks: selected rows are split defensively."],
            state["accepted_risks"],
        )
        self.assertEqual(
            ["structure-preflight-r1: Codex SDK turn timed out after 600 seconds"],
            state["blocking_reasons"],
        )

    def test_load_simple_yaml_recovers_nested_state_list_leaves(self) -> None:
        self.write_state(status="scope-ready-for-writer", semantic_round=0)
        self.state_path.write_text(
            self.state_path.read_text(encoding="utf-8")
            + "\n".join(
                [
                    "latest_artifacts:",
                    "  canonical_test_cases: test-cases/1-section-scope.md",
                    "  review_outputs:",
                    "    - work/review-cycles/section-scope/outputs/review.md",
                    "blocking_reasons:",
                    "  validator:",
                    "    - current-scope warning remains",
                    "",
                ]
            ),
            encoding="utf-8",
        )
        runner = load_runner_module()

        state = runner.load_simple_yaml(self.state_path)

        self.assertEqual(
            [
                "test-cases/1-section-scope.md",
                "work/review-cycles/section-scope/outputs/review.md",
            ],
            state["latest_artifacts"],
        )
        self.assertEqual(["current-scope warning remains"], state["blocking_reasons"])

    def test_start_dry_run_selects_scope_gap_review_before_writer(self) -> None:
        self.test_case.unlink()
        self.write_state(status="scope-ready-for-gap-review", semantic_round=0)

        result = self.run_runner("start", "--state", str(self.state_path), "--dry-run")

        self.assertEqual(result.returncode, 0, result.stderr)
        payload = json.loads(result.stdout)
        self.assertEqual("reviewer", payload["role"])
        self.assertEqual("scope-gap-review", payload["stage"])
        self.assertEqual("reviewer.scope_gap_review", payload["scenario"])
        self.assertEqual("read_only", payload["sandbox_policy"])

    def test_start_dry_run_selects_initial_writer_after_scope_gap_review_passed(self) -> None:
        self.test_case.unlink()
        self.write_state(status="scope-gap-review-passed", semantic_round=0)

        result = self.run_runner("start", "--state", str(self.state_path), "--dry-run")

        self.assertEqual(result.returncode, 0, result.stderr)
        payload = json.loads(result.stdout)
        self.assertEqual("writer", payload["role"])
        self.assertEqual("writer-r1", payload["stage"])
        self.assertEqual("writer.session_initial_draft", payload["scenario"])
        self.assertEqual("workspace_write", payload["sandbox_policy"])

    def test_start_dry_run_routes_structure_preflight_blocker_to_writer_remediation(self) -> None:
        self.write_state(status="structure-preflight-blocked", semantic_round=1)

        result = self.run_runner("start", "--state", str(self.state_path), "--dry-run")

        self.assertEqual(result.returncode, 0, result.stderr)
        payload = json.loads(result.stdout)
        self.assertEqual("start-session", payload["action"])
        self.assertEqual("writer", payload["role"])
        self.assertEqual("writer-structure-r1", payload["stage"])
        self.assertEqual("writer.remediation.style", payload["scenario"])
        self.assertEqual("workspace_write", payload["sandbox_policy"])

    def test_start_dry_run_blocks_writer_after_second_semantic_round(self) -> None:
        self.write_state(status="semantic-revision-needed", semantic_round=2)

        result = self.run_runner("start", "--state", str(self.state_path), "--dry-run")

        self.assertEqual(result.returncode, 2)
        self.assertIn("semantic round cap reached", result.stderr)
        self.assertIn("do not start writer-r3", result.stderr)

    def test_semantic_review_may_route_directly_to_format_review(self) -> None:
        runner = load_runner_module()
        next_session = runner.NextSession(
            stage="semantic-review-r2",
            role="reviewer",
            scenario="reviewer.semantic_traceability_test_design",
            prompt_path="work/review-cycles/section-scope/prompts/next.md",
            sandbox_policy="workspace_write",
        )

        runner.validate_post_session_state_transition(
            {"stage_status": "semantic-review-ready"},
            {"stage_status": "format-review-ready"},
            next_session,
            self.state_path,
        )

    def test_snapshot_copies_canonical_state_and_design_artifacts(self) -> None:
        self.write_state(status="semantic-review-ready", semantic_round=1)

        result = self.run_runner(
            "snapshot",
            "--state",
            str(self.state_path),
            "--snapshot-id",
            "r1-before-review",
        )

        self.assertEqual(result.returncode, 0, result.stderr)
        payload = json.loads(result.stdout)
        snapshot_dir = self.cycle_dir / "versions" / "r1-before-review"
        snapshot_paths = {item["snapshot_path"] for item in payload["files"]}

        self.assertTrue((snapshot_dir / "snapshot-manifest.yaml").exists())
        self.assertIn("test-cases/1-section-scope.md", snapshot_paths)
        self.assertIn("work/review-cycles/section-scope/cycle-state.yaml", snapshot_paths)
        self.assertIn("work/review-cycles/section-scope/codex-session-map.yaml", snapshot_paths)
        self.assertIn("work/test-design/section-scope/traceability.md", snapshot_paths)

    def test_snapshot_excludes_artifact_write_scratch_files(self) -> None:
        scratch_file = self.test_design_dir / "_artifact_write" / "traceability" / "draft.md"
        scratch_file.parent.mkdir(parents=True)
        scratch_file.write_text("scratch draft", encoding="utf-8")
        normal_file = self.test_design_dir / "coverage-metrics.md"
        normal_file.write_text("coverage", encoding="utf-8")
        self.write_state(status="semantic-review-ready", semantic_round=1)
        self.state_path.write_text(
            self.state_path.read_text(encoding="utf-8")
            + "\n".join(
                [
                    "latest_artifacts:",
                    "  - work/test-design/section-scope/_artifact_write/traceability/draft.md",
                    "  - work/test-design/section-scope/coverage-metrics.md",
                    "",
                ]
            ),
            encoding="utf-8",
        )

        result = self.run_runner(
            "snapshot",
            "--state",
            str(self.state_path),
            "--snapshot-id",
            "r1-before-review",
        )

        self.assertEqual(result.returncode, 0, result.stderr)
        payload = json.loads(result.stdout)
        snapshot_paths = {item["snapshot_path"] for item in payload["files"]}
        self.assertIn("work/test-design/section-scope/coverage-metrics.md", snapshot_paths)
        self.assertNotIn(
            "work/test-design/section-scope/_artifact_write/traceability/draft.md",
            snapshot_paths,
        )
        self.assertFalse(
            (
                self.cycle_dir
                / "versions"
                / "r1-before-review"
                / "work"
                / "test-design"
                / "section-scope"
                / "_artifact_write"
            ).exists()
        )

    def test_snapshot_includes_runner_output_artifacts_when_present(self) -> None:
        self.write_state(status="final-regression-ready", semantic_round=2)
        output_dir = self.cycle_dir / "outputs"
        output_dir.mkdir()
        (output_dir / "semantic-regression-final-final-response.md").write_text(
            "review response",
            encoding="utf-8",
        )

        result = self.run_runner(
            "snapshot",
            "--state",
            str(self.state_path),
            "--snapshot-id",
            "after-final-regression",
        )

        self.assertEqual(result.returncode, 0, result.stderr)
        payload = json.loads(result.stdout)
        snapshot_paths = {item["snapshot_path"] for item in payload["files"]}
        self.assertIn(
            "work/review-cycles/section-scope/outputs/semantic-regression-final-final-response.md",
            snapshot_paths,
        )

    def test_pre_writer_snapshot_skips_missing_canonical_test_case_file(self) -> None:
        self.test_case.unlink()
        self.write_state(status="scope-ready-for-gap-review", semantic_round=0)

        result = self.run_runner(
            "snapshot",
            "--state",
            str(self.state_path),
            "--snapshot-id",
            "before-scope-gap-review",
        )

        self.assertEqual(result.returncode, 0, result.stderr)
        payload = json.loads(result.stdout)
        snapshot_paths = {item["snapshot_path"] for item in payload["files"]}
        self.assertNotIn("test-cases/1-section-scope.md", snapshot_paths)
        self.assertIn("work/review-cycles/section-scope/cycle-state.yaml", snapshot_paths)

    def test_validate_rejects_snapshot_as_canonical_test_case(self) -> None:
        snapshot_case = (
            "work/review-cycles/section-scope/versions/r1/test-cases/1-section-scope.md"
        )
        (self.ft_root / snapshot_case).parent.mkdir(parents=True)
        (self.ft_root / snapshot_case).write_text("snapshot copy", encoding="utf-8")
        self.write_state(prompt="work/review-cycles/section-scope/prompts/next.md")
        content = self.state_path.read_text(encoding="utf-8").replace(
            "canonical_test_cases: test-cases/1-section-scope.md",
            f"canonical_test_cases: {snapshot_case}",
        )
        self.state_path.write_text(content, encoding="utf-8")

        result = self.run_runner("validate", "--state", str(self.state_path))

        self.assertEqual(result.returncode, 2)
        self.assertIn("must not point to a version snapshot", result.stderr)

    def test_terminal_scoped_validator_gate_regression_eval_is_tracked(self) -> None:
        content = TERMINAL_GATE_EVAL.read_text(encoding="utf-8")

        for token in (
            "Terminal Scoped Validator Gate Regression",
            "terminal signed-off has unwaived current-scope validator findings",
            "Validator Warning Waivers",
            "waiver_class",
            "versions/",
            "_artifact_write/",
            "unrelated scope",
        ):
            self.assertIn(token, content)

    def test_terminal_validator_waiver_quality_regression_eval_is_tracked(self) -> None:
        content = WAIVER_QUALITY_EVAL.read_text(encoding="utf-8")

        for token in (
            "Terminal Validator Waiver Quality Regression",
            "ui-employment-canary-v6-terminal-gate-regression",
            "waiver_class",
            "pre-existing",
            "accepted-source-gap",
            "validator-schema-lag",
        ):
            self.assertIn(token, content)

    def test_writer_cross_artifact_sync_and_format_regression_eval_is_tracked(self) -> None:
        content = WRITER_SYNC_FORMAT_EVAL.read_text(encoding="utf-8")

        for token in (
            "Writer Cross-Artifact Sync And Format Regression",
            "ui-employment-canary-v7-waiver-gate-regression",
            "test-case-mixed-schema-duplicate-fields",
            "test-case-runtime-field-duplicated",
            "writer-draft-ready",
            "semantic-review-ready",
            "versions/",
            "_artifact_write/",
        ):
            self.assertIn(token, content)

    def test_terminal_signed_off_gate_rejects_unwaived_current_scope_warning(self) -> None:
        self.write_state(status="signed-off")
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {
            "findings": [
                {
                    "id": "test-case-requiredness-without-empty-or-marker-check",
                    "severity": "warning",
                    "path": "test-cases/1-section-scope.md",
                    "evidence": ["TC-001"],
                }
            ]
        }
        state = runner.load_simple_yaml(self.state_path)

        with self.assertRaisesRegex(
            runner.RunnerError,
            "unwaived current-scope validator findings",
        ):
            runner.enforce_terminal_signed_off_validator_gate(state, self.state_path)

    def test_writer_ready_gate_rejects_scoped_validator_pass_contradiction(self) -> None:
        self.write_state(status="writer-draft-ready")
        (self.test_design_dir / "writer-quality-gate.md").write_text(
            "\n".join(
                [
                    "## Writer Quality Gate",
                    "",
                    "| gate_item | status | evidence | affected_package | required_action | blocks_ready_for_review |",
                    "| --- | --- | --- | --- | --- | --- |",
                    "| `scoped-validator-findings` | `pass` | `outputs/scoped-validator-profile.writer-r1.json`: unresolved_warning_error_count=0 | `WP-01` | - | `no` |",
                    "| `package-ready` | `pass` | self-check passed | `WP-01` | - | `no` |",
                    "",
                ]
            ),
            encoding="utf-8",
        )
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {
            "findings": [
                {
                    "id": "some-current-scope-warning",
                    "severity": "warning",
                    "path": "test-cases/1-section-scope.md",
                    "evidence": ["actual warning"],
                }
            ]
        }
        state = runner.load_simple_yaml(self.state_path)

        with self.assertRaisesRegex(
            runner.RunnerError,
            "writer-quality-gate-contradicts-validator-profile",
        ):
            runner.enforce_session_ready_validator_gate(state, self.state_path)
        profile = json.loads(
            (self.cycle_dir / "outputs" / "scoped-validator-profile.writer-r1.json").read_text(
                encoding="utf-8"
            )
        )
        self.assertEqual(
            "writer-quality-gate-contradicts-validator-profile",
            profile["current_scope_findings"][0]["id"],
        )
        self.assertEqual(2, profile["unresolved_warning_error_count"])

    def test_writer_ready_gate_writes_runner_scoped_validator_profile(self) -> None:
        self.write_state(status="writer-draft-ready")
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {
            "findings": [
                {
                    "id": "some-current-scope-warning",
                    "severity": "warning",
                    "path": "test-cases/1-section-scope.md",
                    "evidence": ["actual warning"],
                }
            ]
        }
        state = runner.load_simple_yaml(self.state_path)

        gate = runner.session_ready_validator_gate(state, self.state_path)

        profile_path = self.cycle_dir / "outputs" / "scoped-validator-profile.writer-r1.json"
        self.assertTrue(profile_path.exists())
        profile = json.loads(profile_path.read_text(encoding="utf-8"))
        self.assertEqual("codex_review_cycle_runner", profile["generated_by"])
        self.assertEqual(1, profile["unresolved_warning_error_count"])
        self.assertEqual(
            "work/review-cycles/section-scope/outputs/scoped-validator-profile.writer-r1.json",
            gate["scoped_validator_profile"],
        )

    def test_deterministic_structure_preflight_pass_advances_to_semantic_review(self) -> None:
        self.write_state(status="writer-draft-ready")
        self.write_valid_structure_preflight_artifacts()
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {"findings": []}
        runner.load_openai_codex_runtime = lambda *args, **kwargs: self.fail("SDK must not be loaded")
        state = runner.load_simple_yaml(self.state_path)

        payload = runner.run_real_session(
            state,
            self.state_path,
            cwd=str(self.root),
            approval_mode="auto_review",
            model=None,
            session_timeout_seconds=1,
            process_supervised=True,
        )

        updated_state = runner.load_simple_yaml(self.state_path)
        self.assertEqual("completed-deterministic-session", payload["action"])
        self.assertEqual("deterministic", payload["execution_mode"])
        self.assertEqual("semantic-review-ready", updated_state["stage_status"])
        self.assertEqual("structure-preflight-r1", updated_state["current_stage"])
        self.assertEqual(
            "work/review-cycles/section-scope/prompts/prompt.semantic-review-r1.md",
            updated_state["active_transition_prompt"],
        )
        for artifact in (
            self.cycle_dir / "outputs" / "structure-preflight-r1-findings.md",
            self.cycle_dir / "outputs" / "reviewer-session-log.structure-preflight-r1.md",
            self.cycle_dir / "outputs" / "agent-decision-log.structure-preflight-r1.md",
            self.cycle_dir / "outputs" / "structure-preflight-r1-completion.yaml",
            self.cycle_dir / "versions" / "after-structure-preflight-r1" / "snapshot-manifest.yaml",
        ):
            self.assertTrue(artifact.exists(), str(artifact))
        manifest = runner.load_simple_yaml(
            self.cycle_dir / "outputs" / "structure-preflight-r1-completion.yaml"
        )
        self.assertEqual("deterministic", manifest["execution_mode"])
        self.assertEqual("deterministic-structure-preflight", manifest["thread_id"])
        session_map = (self.cycle_dir / "codex-session-map.yaml").read_text(encoding="utf-8")
        self.assertIn("execution_mode: deterministic", session_map)
        self.assertIn("thread_id: deterministic-structure-preflight", session_map)
        latest_artifacts = updated_state["latest_artifacts"]
        self.assertIn(
            "work/review-cycles/section-scope/outputs/structure-preflight-r1-findings.md",
            latest_artifacts,
        )
        self.assertIn(
            "work/review-cycles/section-scope/outputs/structure-preflight-r1-completion.yaml",
            latest_artifacts,
        )

    def test_deterministic_structure_preflight_blocks_bootstrap_profile(self) -> None:
        self.write_state(status="writer-draft-ready")
        self.write_valid_structure_preflight_artifacts()
        self.write_scoped_validator_profile(
            command="bootstrap before runner validate - must be overwritten"
        )
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {"findings": []}
        state = runner.load_simple_yaml(self.state_path)

        payload = runner.run_real_session(
            state,
            self.state_path,
            cwd=str(self.root),
            approval_mode="auto_review",
            model=None,
            session_timeout_seconds=1,
            process_supervised=True,
        )

        updated_state = runner.load_simple_yaml(self.state_path)
        self.assertEqual("structure-preflight-blocked", updated_state["stage_status"])
        self.assertIn(
            "structure-preflight-validator-profile-bootstrap",
            updated_state["blocking_findings"],
        )
        self.assertEqual(1, payload["finding_count"])
        self.assertTrue((self.prompt_dir / "prompt.writer-structure-r1.md").exists())
        findings = (self.cycle_dir / "outputs" / "structure-preflight-r1-findings.md").read_text(
            encoding="utf-8"
        )
        self.assertIn("bootstrap evidence", findings)

    def test_deterministic_structure_preflight_blocks_stale_coverage_summary(self) -> None:
        self.write_state(status="writer-draft-ready")
        self.write_valid_structure_preflight_artifacts()
        self.test_case.write_text(
            self.test_case.read_text(encoding="utf-8").replace(
                "| executable_test_cases | TC-001..TC-002 |",
                "| executable_test_cases | TC-001..TC-001 |",
            ),
            encoding="utf-8",
        )
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {"findings": []}
        state = runner.load_simple_yaml(self.state_path)

        runner.run_real_session(
            state,
            self.state_path,
            cwd=str(self.root),
            approval_mode="auto_review",
            model=None,
            session_timeout_seconds=1,
            process_supervised=True,
        )

        updated_state = runner.load_simple_yaml(self.state_path)
        self.assertEqual("structure-preflight-blocked", updated_state["stage_status"])
        self.assertIn(
            "structure-preflight-coverage-summary-range-stale",
            updated_state["blocking_findings"],
        )

    def test_deterministic_structure_preflight_blocks_missing_runtime_field_with_tc_id(self) -> None:
        self.write_state(status="writer-draft-ready")
        self.write_valid_structure_preflight_artifacts()
        self.test_case.write_text(
            self.test_case.read_text(encoding="utf-8").replace(
                "**package_id:** WP-01\n**Traceability:** REQ-002",
                "**Traceability:** REQ-002",
            ),
            encoding="utf-8",
        )
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {"findings": []}
        state = runner.load_simple_yaml(self.state_path)

        runner.run_real_session(
            state,
            self.state_path,
            cwd=str(self.root),
            approval_mode="auto_review",
            model=None,
            session_timeout_seconds=1,
            process_supervised=True,
        )

        updated_state = runner.load_simple_yaml(self.state_path)
        self.assertEqual("structure-preflight-blocked", updated_state["stage_status"])
        self.assertIn(
            "structure-preflight-test-case-runtime-field-missing",
            updated_state["blocking_findings"],
        )
        findings = (self.cycle_dir / "outputs" / "structure-preflight-r1-findings.md").read_text(
            encoding="utf-8"
        )
        self.assertIn("TC-002", findings)
        self.assertIn("package_id", findings)

    def test_run_until_terminal_can_stop_after_deterministic_structure_preflight_pass(self) -> None:
        self.write_state(status="writer-draft-ready")
        self.write_valid_structure_preflight_artifacts()
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {"findings": []}

        payload = runner.run_until_terminal(
            self.state_path,
            cwd=str(self.root),
            approval_mode="auto_review",
            model=None,
            max_sessions=1,
            session_timeout_seconds=1,
        )

        updated_state = runner.load_simple_yaml(self.state_path)
        self.assertEqual("max-sessions-reached", payload["action"])
        self.assertEqual("semantic-review-ready", updated_state["stage_status"])
        self.assertEqual("completed-deterministic-session", payload["sessions"][0]["action"])

    def test_bounded_semantic_pass_advances_to_format_review(self) -> None:
        self.write_state(status="semantic-review-ready", semantic_round=1)
        self.write_valid_structure_preflight_artifacts()
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {"findings": []}
        recorded_turn_kwargs = {}

        def fake_turn(thread, prompt, **kwargs):
            recorded_turn_kwargs.update(kwargs)
            return thread.run(prompt, **kwargs)

        runner.run_codex_turn_with_timeout = fake_turn
        previous_module = self.install_fake_codex(
            runner,
            final_response=self.bounded_semantic_response(),
        )
        try:
            payload = runner.run_real_session(
                runner.load_simple_yaml(self.state_path),
                self.state_path,
                cwd=str(self.root),
                approval_mode="auto_review",
                model=None,
                session_timeout_seconds=1,
                process_supervised=False,
            )
        finally:
            self.restore_openai_codex_module(previous_module)

        updated_state = runner.load_simple_yaml(self.state_path)
        self.assertEqual("completed-bounded-semantic-session", payload["action"])
        self.assertEqual("bounded-sdk", payload["execution_mode"])
        self.assertEqual("format-review-ready", updated_state["stage_status"])
        self.assertEqual("semantic-review-r1", updated_state["current_stage"])
        self.assertEqual(1, updated_state["semantic_round"])
        self.assertEqual(
            "work/review-cycles/section-scope/prompts/prompt.structure-format-final.md",
            updated_state["active_transition_prompt"],
        )
        self.assertEqual("read_only", recorded_turn_kwargs["sandbox"])
        for artifact in (
            self.cycle_dir / "outputs" / "round-1-findings.md",
            self.cycle_dir / "outputs" / "round-1-traceability-matrix.md",
            self.cycle_dir / "outputs" / "round-1-traceability-matrix.xlsx",
            self.cycle_dir / "outputs" / "reviewer-session-log.semantic-review-r1.md",
            self.cycle_dir / "outputs" / "agent-decision-log.semantic-review-r1.md",
            self.cycle_dir / "outputs" / "semantic-review-r1-response.json",
            self.cycle_dir / "outputs" / "semantic-review-r1-completion.yaml",
            self.cycle_dir / "versions" / "after-semantic-review-r1" / "snapshot-manifest.yaml",
        ):
            self.assertTrue(artifact.exists(), str(artifact))
        completion = runner.load_simple_yaml(
            self.cycle_dir / "outputs" / "semantic-review-r1-completion.yaml"
        )
        self.assertEqual("read_only", completion["sandbox_policy"])
        self.assertFalse(completion["reviewer_write_exception"])
        prompt_text = (self.prompt_dir / "prompt.semantic-review-r1.bounded.md").read_text(
            encoding="utf-8"
        )
        self.assertNotIn("Run reviewer.semantic_traceability_test_design", prompt_text)
        self.assertIn("Do not recursively read directories", prompt_text)
        self.assertIn("fts/demo-ft/work/test-design/section-scope/coverage-map.md", prompt_text)
        from openpyxl import load_workbook

        matrix_md = (self.cycle_dir / "outputs" / "round-1-traceability-matrix.md").read_text(
            encoding="utf-8"
        )
        workbook = load_workbook(self.cycle_dir / "outputs" / "round-1-traceability-matrix.xlsx")
        sheet = workbook.active
        xlsx_rows = [tuple(cell.value or "" for cell in row) for row in sheet.iter_rows()]
        self.assertEqual(
            (
                "atom_id",
                "source_ref",
                "coverage_status",
                "covered_by_tc",
                "notes",
            ),
            xlsx_rows[0],
        )
        self.assertIn("| REQ-001 | Create document. | covered | TC-001 | TC-001 expected result. |", matrix_md)
        self.assertIn(("REQ-001", "Create document.", "covered", "TC-001", "TC-001 expected result."), xlsx_rows)
        session_map = (self.cycle_dir / "codex-session-map.yaml").read_text(encoding="utf-8")
        self.assertIn("execution_mode: bounded-sdk", session_map)
        self.assertIn("sandbox: read_only", session_map)
        self.assertIn("thread_id: fake-thread-bounded", session_map)
        self.assertIn(
            "work/review-cycles/section-scope/outputs/semantic-review-r1-completion.yaml",
            updated_state["latest_artifacts"],
        )

    def test_bounded_structure_format_pass_advances_to_final_regression(self) -> None:
        self.write_state(status="format-review-ready", semantic_round=2)
        self.write_valid_structure_preflight_artifacts()
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {"findings": []}
        previous_module = self.install_fake_codex(
            runner,
            final_response=self.bounded_reviewer_response(
                recommended_stage_status="final-regression-ready",
            ),
            thread_id="fake-thread-format-review",
            turn_id="fake-turn-format-review",
        )
        try:
            payload = runner.run_real_session(
                runner.load_simple_yaml(self.state_path),
                self.state_path,
                cwd=str(self.root),
                approval_mode="auto_review",
                model=None,
                session_timeout_seconds=1,
                process_supervised=False,
            )
        finally:
            self.restore_openai_codex_module(previous_module)

        updated_state = runner.load_simple_yaml(self.state_path)
        completion = runner.load_simple_yaml(
            self.cycle_dir / "outputs" / "format-review-final-completion.yaml"
        )
        self.assertEqual("completed-bounded-reviewer-session", payload["action"])
        self.assertEqual("final-regression-ready", updated_state["stage_status"])
        self.assertEqual("format-review-final", updated_state["current_stage"])
        self.assertEqual("read_only", completion["sandbox_policy"])
        self.assertFalse(completion["reviewer_write_exception"])
        session_map = (self.cycle_dir / "codex-session-map.yaml").read_text(encoding="utf-8")
        self.assertIn("thread_id: fake-thread-format-review", session_map)
        self.assertIn("sandbox: read_only", session_map)

    def test_bounded_semantic_regression_pass_signs_off_read_only(self) -> None:
        self.write_state(status="final-regression-ready", semantic_round=2)
        self.write_valid_structure_preflight_artifacts()
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {"findings": []}
        previous_module = self.install_fake_codex(
            runner,
            final_response=self.bounded_reviewer_response(
                recommended_stage_status="signed-off",
            ),
            thread_id="fake-thread-regression-review",
            turn_id="fake-turn-regression-review",
        )
        try:
            payload = runner.run_real_session(
                runner.load_simple_yaml(self.state_path),
                self.state_path,
                cwd=str(self.root),
                approval_mode="auto_review",
                model=None,
                session_timeout_seconds=1,
                process_supervised=False,
            )
        finally:
            self.restore_openai_codex_module(previous_module)

        updated_state = runner.load_simple_yaml(self.state_path)
        completion = runner.load_simple_yaml(
            self.cycle_dir / "outputs" / "semantic-regression-final-completion.yaml"
        )
        self.assertEqual("completed-bounded-reviewer-session", payload["action"])
        self.assertEqual("signed-off", updated_state["stage_status"])
        self.assertEqual("semantic-regression-final", updated_state["current_stage"])
        self.assertEqual("read_only", completion["sandbox_policy"])
        self.assertFalse(completion["reviewer_write_exception"])
        session_map = (self.cycle_dir / "codex-session-map.yaml").read_text(encoding="utf-8")
        self.assertIn("thread_id: fake-thread-regression-review", session_map)
        self.assertIn("sandbox: read_only", session_map)

    def test_bounded_semantic_findings_force_semantic_revision(self) -> None:
        self.write_state(status="semantic-review-ready", semantic_round=1)
        self.write_valid_structure_preflight_artifacts()
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {"findings": []}
        finding = {
            "id": "SEM-001",
            "severity": "warning",
            "title": "REQ-002 lacks negative coverage",
            "requirement_id": "REQ-002",
            "tc_ids": ["TC-002"],
            "evidence": "Only positive update flow is covered.",
            "recommendation": "Add a negative authorization or validation test.",
        }
        previous_module = self.install_fake_codex(
            runner,
            final_response=self.bounded_semantic_response(
                findings=[finding],
                recommended_stage_status="format-review-ready",
            ),
        )
        try:
            runner.run_real_session(
                runner.load_simple_yaml(self.state_path),
                self.state_path,
                cwd=str(self.root),
                approval_mode="auto_review",
                model=None,
                session_timeout_seconds=1,
                process_supervised=False,
            )
        finally:
            self.restore_openai_codex_module(previous_module)

        updated_state = runner.load_simple_yaml(self.state_path)
        self.assertEqual("semantic-revision-needed", updated_state["stage_status"])
        self.assertEqual(["SEM-001"], updated_state["blocking_findings"])
        self.assertEqual(
            "work/review-cycles/section-scope/prompts/prompt.writer-r2.md",
            updated_state["active_transition_prompt"],
        )
        self.assertTrue((self.prompt_dir / "prompt.writer-r2.md").exists())
        findings = (self.cycle_dir / "outputs" / "round-1-findings.md").read_text(
            encoding="utf-8"
        )
        self.assertIn("SEM-001", findings)
        self.assertIn("REQ-002 lacks negative coverage", findings)

    def test_bounded_semantic_invalid_json_blocks_with_diagnostic_artifact(self) -> None:
        self.write_state(status="semantic-review-ready", semantic_round=1)
        self.write_valid_structure_preflight_artifacts()
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {"findings": []}
        previous_module = self.install_fake_codex(
            runner,
            final_response="This is not fenced JSON.",
        )
        try:
            payload = runner.run_real_session(
                runner.load_simple_yaml(self.state_path),
                self.state_path,
                cwd=str(self.root),
                approval_mode="auto_review",
                model=None,
                session_timeout_seconds=1,
                process_supervised=False,
            )
        finally:
            self.restore_openai_codex_module(previous_module)

        updated_state = runner.load_simple_yaml(self.state_path)
        self.assertEqual("failed", payload["session_status"])
        self.assertEqual("blocked-input", updated_state["stage_status"])
        self.assertIn(
            "semantic-review-r1: bounded semantic reviewer returned invalid JSON",
            updated_state["blocking_reasons"],
        )
        self.assertTrue((self.cycle_dir / "outputs" / "semantic-review-r1-invalid-response.md").exists())
        self.assertTrue((self.cycle_dir / "outputs" / "semantic-review-r1-response.json").exists())

    def test_bounded_semantic_timeout_recovers_to_blocked_input(self) -> None:
        self.write_state(status="semantic-review-ready", semantic_round=1)
        self.write_valid_structure_preflight_artifacts()
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {"findings": []}

        def fake_turn(*args, **kwargs):
            raise runner.concurrent.futures.TimeoutError()

        runner.run_codex_turn_with_timeout = fake_turn
        previous_module = self.install_fake_codex(
            runner,
            final_response=self.bounded_semantic_response(),
        )
        try:
            payload = runner.run_real_session(
                runner.load_simple_yaml(self.state_path),
                self.state_path,
                cwd=str(self.root),
                approval_mode="auto_review",
                model=None,
                session_timeout_seconds=1,
                process_supervised=False,
            )
        finally:
            self.restore_openai_codex_module(previous_module)

        updated_state = runner.load_simple_yaml(self.state_path)
        self.assertEqual("recovered-timeout-session", payload["action"])
        self.assertEqual("blocked-input", updated_state["stage_status"])
        self.assertTrue((self.cycle_dir / "outputs" / "semantic-review-r1-timeout-recovery.md").exists())
        self.assertTrue((self.cycle_dir / "outputs" / "semantic-review-r1-completion.yaml").exists())

    def test_bounded_semantic_diagnostic_renders_from_blocked_input_active_prompt(self) -> None:
        self.write_state(
            status="blocked-input",
            semantic_round=1,
            prompt="work/review-cycles/section-scope/prompts/prompt.semantic-review-r1.md",
        )
        self.state_path.write_text(
            self.state_path.read_text(encoding="utf-8").replace(
                "current_stage: writer-r1",
                "current_stage: structure-preflight-r1",
            ),
            encoding="utf-8",
        )
        self.write_valid_structure_preflight_artifacts()
        (self.prompt_dir / "prompt.semantic-review-r1.md").write_text(
            "old lifecycle prompt",
            encoding="utf-8",
        )

        result = self.run_runner(
            "diagnose-sdk-turn",
            "--state",
            str(self.state_path),
            "--prompt-source",
            "bounded-semantic",
            "--dry-run",
        )

        self.assertEqual(result.returncode, 0, result.stderr)
        payload = json.loads(result.stdout)
        self.assertEqual("rendered-sdk-diagnostic-prompt", payload["action"])
        self.assertEqual("bounded-semantic", payload["prompt_source"])
        self.assertEqual("semantic-review-r1", payload["stage"])
        self.assertEqual("reviewer.semantic_traceability_test_design", payload["scenario"])
        self.assertEqual("read_only", payload["sandbox_policy"])

    def test_doctor_reports_stale_scoped_validator_profile_event(self) -> None:
        self.write_state(status="blocked-input")
        output_dir = self.cycle_dir / "outputs"
        output_dir.mkdir(parents=True, exist_ok=True)
        profile_path = output_dir / "scoped-validator-profile.writer-r1.json"
        profile_path.write_text(
            json.dumps(
                {
                    "version": 1,
                    "generated_by": "codex_review_cycle_runner",
                    "command": "python scripts/validate_agent_artifacts.py --root fts/demo-ft --json",
                    "scope_slug": "section-scope",
                    "canonical_test_cases": "test-cases/1-section-scope.md",
                    "test_design_dir": "work/test-design/section-scope",
                    "current_stage": "writer-r1",
                    "current_scope_findings": [
                        {
                            "id": "current-scope-warning",
                            "severity": "warning",
                            "path": "test-cases/1-section-scope.md",
                        }
                    ],
                    "unresolved_warning_error_count": 2,
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )
        runner = load_runner_module()
        runner.append_runner_event(
            self.cycle_dir,
            "scoped_validator_profile_written",
            stage="writer-r1",
            profile="work/review-cycles/section-scope/outputs/scoped-validator-profile.writer-r1.json",
            unresolved_warning_error_count=0,
        )

        payload = runner.build_doctor_payload(self.state_path, stale_lock_seconds=60)

        consistency = payload["events"]["scoped_validator_profile_consistency"]
        self.assertTrue(consistency["checked"])
        self.assertFalse(consistency["consistent"])
        self.assertEqual(0, consistency["event_unresolved_warning_error_count"])
        self.assertEqual(2, consistency["profile_unresolved_warning_error_count"])
        self.assertIn("differs from profile", consistency["issues"][0])
        self.assertEqual("inspect-event-profile-drift", payload["recommendation"])

    def test_run_real_session_timeout_recovers_to_blocked_input(self) -> None:
        self.write_state(status="scope-ready-for-writer")
        runner = load_runner_module()
        runner.prompt_text_for_session = lambda state, state_path, next_session: "prompt"
        runner.sdk_sandbox = lambda policy: policy
        runner.sdk_approval_mode = lambda mode: mode

        class FakeThread:
            id = "thread-timeout"

            def set_name(self, name: str) -> None:
                self.name = name

        class FakeCodex:
            def thread_start(self, **kwargs):
                return FakeThread()

            def close(self) -> None:
                pass

        def fake_turn(*args, **kwargs):
            raise runner.concurrent.futures.TimeoutError()

        previous_module = sys.modules.get("openai_codex")
        sys.modules["openai_codex"] = types.SimpleNamespace(Codex=FakeCodex)
        runner.run_codex_turn_with_timeout = fake_turn
        state = runner.load_simple_yaml(self.state_path)
        try:
            payload = runner.run_real_session(
                state,
                self.state_path,
                cwd=str(self.root),
                approval_mode="auto_review",
                model=None,
                session_timeout_seconds=1,
                process_supervised=False,
            )
        finally:
            if previous_module is None:
                sys.modules.pop("openai_codex", None)
            else:
                sys.modules["openai_codex"] = previous_module

        updated_state = runner.load_simple_yaml(self.state_path)
        self.assertEqual("recovered-timeout-session", payload["action"])
        self.assertEqual("timeout", payload["turn_status"])
        self.assertEqual("failed", payload["session_status"])
        self.assertEqual("blocked-input", updated_state["stage_status"])
        self.assertTrue((self.cycle_dir / "outputs" / "writer-r1-completion.yaml").exists())
        self.assertTrue((self.cycle_dir / "outputs" / "writer-r1-timeout-recovery.md").exists())

    def test_start_cli_timeout_recovers_to_blocked_input_with_zero_exit(self) -> None:
        self.write_state(status="scope-ready-for-writer")
        fake_sdk_dir = self.root / "fake-sdk"
        fake_sdk_dir.mkdir()
        (fake_sdk_dir / "openai_codex.py").write_text(
            "\n".join(
                [
                    "import time",
                    "",
                    "class Sandbox:",
                    "    read_only = 'read_only'",
                    "    workspace_write = 'workspace_write'",
                    "    full_access = 'full_access'",
                    "",
                    "class ApprovalMode:",
                    "    AUTO_REVIEW = 'auto_review'",
                    "    DENY_ALL = 'deny_all'",
                    "",
                    "class FakeThread:",
                    "    id = 'fake-thread-timeout'",
                    "    def set_name(self, name):",
                    "        self.name = name",
                    "    def run(self, *args, **kwargs):",
                    "        time.sleep(10)",
                    "",
                    "class Codex:",
                    "    def thread_start(self, **kwargs):",
                    "        return FakeThread()",
                    "    def close(self):",
                    "        pass",
                    "",
                ]
            ),
            encoding="utf-8",
        )
        env = os.environ.copy()
        env["PYTHONIOENCODING"] = "utf-8"
        env["PYTHONPATH"] = str(fake_sdk_dir)

        result = subprocess.run(
            [
                sys.executable,
                str(RUNNER),
                "start",
                "--state",
                str(self.state_path),
                "--cwd",
                str(self.root),
                "--session-timeout-seconds",
                "1",
            ],
            cwd=ROOT_DIR,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            env=env,
            check=False,
        )

        self.assertEqual(result.returncode, 0, result.stderr)
        payload = json.loads(result.stdout)
        updated_state = load_runner_module().load_simple_yaml(self.state_path)
        self.assertEqual("recovered-timeout-session", payload["action"])
        self.assertEqual("timeout", payload["turn_status"])
        self.assertEqual("failed", payload["session_status"])
        self.assertEqual("blocked-input", updated_state["stage_status"])
        self.assertFalse((self.cycle_dir / "runner.lock.yaml").exists())
        self.assertTrue((self.cycle_dir / "outputs" / "writer-r1-completion.yaml").exists())
        self.assertTrue((self.cycle_dir / "outputs" / "writer-r1-timeout-recovery.md").exists())

    def test_timeout_recovery_reports_artifacts_dirty_profile_and_missing_next_prompt(self) -> None:
        self.write_state(status="scope-ready-for-writer")
        self.state_path.write_text(
            self.state_path.read_text(encoding="utf-8")
            + "\n".join(
                [
                    "latest_artifacts:",
                    "  - test-cases/1-section-scope.md",
                    "open_questions:",
                    "  - GAP-001: exact UI notification text is not confirmed.",
                    "accepted_risks:",
                    "  - source-quality-oversized-blocks: selected rows are split defensively.",
                    "",
                ]
            ),
            encoding="utf-8",
        )
        runner = load_runner_module()
        state = runner.load_simple_yaml(self.state_path)
        next_session = runner.next_session_for_state(state)
        self.assertIsNotNone(next_session)
        output_dir = self.cycle_dir / "outputs"
        output_dir.mkdir(parents=True, exist_ok=True)
        (output_dir / "writer-r1-response.md").write_text("# Writer response\n", encoding="utf-8")
        (output_dir / "scoped-validator-profile.writer-r1.json").write_text(
            json.dumps(
                {
                    "generated_by": "codex_review_cycle_runner",
                    "command": "python scripts/validate_agent_artifacts.py --root fts/demo-ft --json",
                    "scope_slug": "section-scope",
                    "canonical_test_cases": "test-cases/1-section-scope.md",
                    "test_design_dir": "work/test-design/section-scope",
                    "current_stage": "writer-r1",
                    "current_scope_findings": [
                        {"id": "writer-quality-gate-scoped-validator-profile-invalid", "severity": "warning"}
                    ],
                    "unresolved_warning_error_count": 1,
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )

        payload = runner.recover_timed_out_session(
            self.state_path,
            state_before=state,
            next_session=next_session,
            thread_id="thread-timeout",
            approval_mode="auto_review",
            model=None,
            before_snapshot_id="before-writer-r1",
            after_snapshot_id="after-writer-r1",
            started_at_epoch=100,
            timeout_seconds=1800,
        )

        updated_state = runner.load_simple_yaml(self.state_path)
        timeout_output = self.cycle_dir / "outputs" / "writer-r1-timeout-recovery.md"
        timeout_text = timeout_output.read_text(encoding="utf-8")
        self.assertEqual("recovered-timeout-session", payload["action"])
        self.assertEqual("blocked-input", updated_state["stage_status"])
        self.assertIn("writer response: present", timeout_text)
        self.assertIn("unresolved_warning_error_count=1", timeout_text)
        self.assertIn("next review prompt: missing", timeout_text)
        self.assertTrue(
            any("current-stage scoped validator profile unresolved=1" in reason for reason in updated_state["blocking_reasons"])
        )
        self.assertIn("test-cases/1-section-scope.md", updated_state["latest_artifacts"])
        self.assertIn(
            "work/review-cycles/section-scope/outputs/writer-r1-timeout-recovery.md",
            updated_state["latest_artifacts"],
        )
        self.assertIn(
            "work/review-cycles/section-scope/outputs/writer-r1-completion.yaml",
            updated_state["latest_artifacts"],
        )
        self.assertEqual(
            ["GAP-001: exact UI notification text is not confirmed."],
            updated_state["open_questions"],
        )
        self.assertEqual(
            ["source-quality-oversized-blocks: selected rows are split defensively."],
            updated_state["accepted_risks"],
        )
        snapshot_manifest = (
            self.cycle_dir / "versions" / "after-writer-r1" / "snapshot-manifest.yaml"
        ).read_text(encoding="utf-8")
        self.assertIn(
            "work/review-cycles/section-scope/outputs/writer-r1-timeout-recovery.md",
            snapshot_manifest,
        )
        self.assertIn(
            "work/review-cycles/section-scope/outputs/writer-r1-completion.yaml",
            snapshot_manifest,
        )

    def test_process_supervised_timeout_accepts_completed_child_payload(self) -> None:
        self.write_state(status="scope-ready-for-writer")
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {"findings": []}
        state = runner.load_simple_yaml(self.state_path)

        class FakeProcess:
            pid = 4242
            returncode = None

            def __init__(self, command, **kwargs):
                self.command = command
                payload_path = Path(command[command.index("--payload-out") + 1])
                payload_path.parent.mkdir(parents=True, exist_ok=True)
                updated_state = self.state_path.read_text(encoding="utf-8").replace(
                    "stage_status: scope-ready-for-writer",
                    "stage_status: writer-draft-ready",
                )
                self.state_path.write_text(updated_state, encoding="utf-8")
                payload_path.write_text(
                    json.dumps(
                        {
                            "action": "completed-session",
                            "stage": "writer-r1",
                            "role": "writer",
                            "scenario": "writer.session_initial_draft",
                            "thread_id": "fake-thread",
                            "turn_id": "fake-turn",
                            "turn_status": "completed",
                            "session_status": "completed",
                            "state_advanced": True,
                        }
                    ),
                    encoding="utf-8",
                )

            def communicate(self, timeout=None):
                if timeout == 10:
                    return "", ""
                raise subprocess.TimeoutExpired(cmd=self.command, timeout=timeout)

            def kill(self):
                self.returncode = -9

        FakeProcess.state_path = self.state_path
        previous_popen = runner.subprocess.Popen
        try:
            runner.subprocess.Popen = FakeProcess
            payload = runner.run_real_session_process_supervised(
                state,
                self.state_path,
                cwd=str(self.root),
                approval_mode="auto_review",
                model=None,
                runner_lock=None,
                session_timeout_seconds=1,
            )
        finally:
            runner.subprocess.Popen = previous_popen

        updated_state = runner.load_simple_yaml(self.state_path)
        events = (self.cycle_dir / "runner-events.ndjson").read_text(encoding="utf-8")
        self.assertEqual("completed-session", payload["action"])
        self.assertEqual("writer-draft-ready", updated_state["stage_status"])
        self.assertFalse((self.cycle_dir / "outputs" / "writer-r1-timeout-recovery.md").exists())
        self.assertIn("session_child_timeout_after_payload", events)

    def test_process_supervised_timeout_rereads_child_payload_after_kill(self) -> None:
        self.write_state(status="scope-ready-for-writer")
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {"findings": []}
        state = runner.load_simple_yaml(self.state_path)

        class FakeProcess:
            pid = 4243
            returncode = None

            def __init__(self, command, **kwargs):
                self.command = command
                self.payload_path = Path(command[command.index("--payload-out") + 1])

            def write_payload(self):
                self.payload_path.parent.mkdir(parents=True, exist_ok=True)
                updated_state = self.state_path.read_text(encoding="utf-8").replace(
                    "stage_status: scope-ready-for-writer",
                    "stage_status: writer-draft-ready",
                )
                self.state_path.write_text(updated_state, encoding="utf-8")
                self.payload_path.write_text(
                    json.dumps(
                        {
                            "action": "completed-session",
                            "stage": "writer-r1",
                            "role": "writer",
                            "scenario": "writer.session_initial_draft",
                            "thread_id": "fake-thread",
                            "turn_id": "fake-turn",
                            "turn_status": "completed",
                            "session_status": "completed",
                            "state_advanced": True,
                        }
                    ),
                    encoding="utf-8",
                )

            def communicate(self, timeout=None):
                if timeout == 10:
                    self.write_payload()
                    return "", ""
                raise subprocess.TimeoutExpired(cmd=self.command, timeout=timeout)

            def kill(self):
                self.returncode = -9

        FakeProcess.state_path = self.state_path
        previous_popen = runner.subprocess.Popen
        try:
            runner.subprocess.Popen = FakeProcess
            payload = runner.run_real_session_process_supervised(
                state,
                self.state_path,
                cwd=str(self.root),
                approval_mode="auto_review",
                model=None,
                runner_lock=None,
                session_timeout_seconds=1,
            )
        finally:
            runner.subprocess.Popen = previous_popen

        updated_state = runner.load_simple_yaml(self.state_path)
        events = (self.cycle_dir / "runner-events.ndjson").read_text(encoding="utf-8")
        self.assertEqual("completed-session", payload["action"])
        self.assertEqual("writer-draft-ready", updated_state["stage_status"])
        self.assertFalse((self.cycle_dir / "outputs" / "writer-r1-timeout-recovery.md").exists())
        self.assertIn("session_child_timeout_after_payload", events)

    def test_process_supervised_timeout_preserves_valid_state_progress_without_payload(self) -> None:
        self.write_state(status="scope-ready-for-writer")
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {"findings": []}
        state = runner.load_simple_yaml(self.state_path)

        class FakeProcess:
            pid = 4244
            returncode = None

            def __init__(self, command, **kwargs):
                self.command = command

            def communicate(self, timeout=None):
                if timeout == 10:
                    return "", ""
                raise subprocess.TimeoutExpired(cmd=self.command, timeout=timeout)

            def kill(self):
                self.returncode = -9
                updated_state = self.state_path.read_text(encoding="utf-8").replace(
                    "stage_status: scope-ready-for-writer",
                    "stage_status: writer-draft-ready",
                )
                self.state_path.write_text(updated_state, encoding="utf-8")

        FakeProcess.state_path = self.state_path
        previous_popen = runner.subprocess.Popen
        try:
            runner.subprocess.Popen = FakeProcess
            payload = runner.run_real_session_process_supervised(
                state,
                self.state_path,
                cwd=str(self.root),
                approval_mode="auto_review",
                model=None,
                runner_lock=None,
                session_timeout_seconds=1,
            )
        finally:
            runner.subprocess.Popen = previous_popen

        updated_state = runner.load_simple_yaml(self.state_path)
        events = (self.cycle_dir / "runner-events.ndjson").read_text(encoding="utf-8")
        self.assertEqual("completed-session-after-timeout", payload["action"])
        self.assertEqual("completed-with-progress", payload["session_status"])
        self.assertEqual("writer-draft-ready", updated_state["stage_status"])
        self.assertFalse((self.cycle_dir / "outputs" / "writer-r1-timeout-recovery.md").exists())
        self.assertTrue((self.cycle_dir / "outputs" / "writer-r1-timeout-after-progress.md").exists())
        self.assertIn("session_child_timeout_after_state_progress", events)

    def test_process_supervised_timeout_recovers_writer_artifact_progress_without_state_progress(self) -> None:
        self.write_state(status="semantic-revision-needed", semantic_round=1)
        (self.prompt_dir / "prompt.semantic-review-r2.md").write_text(
            "Review writer-r2 changes",
            encoding="utf-8",
        )
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {"findings": []}
        state = runner.load_simple_yaml(self.state_path)

        class FakeProcess:
            pid = 4245
            returncode = None

            def __init__(self, command, **kwargs):
                self.command = command

            def communicate(self, timeout=None):
                if timeout == 10:
                    return "", ""
                raise subprocess.TimeoutExpired(cmd=self.command, timeout=timeout)

            def kill(self):
                self.returncode = -9
                output_dir = self.state_path.parent / "outputs"
                output_dir.mkdir(parents=True, exist_ok=True)
                (output_dir / "writer-r2-response.md").write_text(
                    "# Writer Response\n\nFINDING-001 fixed.\n",
                    encoding="utf-8",
                )
                (output_dir / "writer-session-log.md").write_text(
                    "# Writer Session Log\n\n## Session Metadata\n\nskill = ft-test-case-writer\n",
                    encoding="utf-8",
                )
                (output_dir / "scoped-validator-profile.writer-r2.json").write_text(
                    json.dumps(
                        {
                            "version": 1,
                            "current_stage": "writer-r2",
                            "current_scope_findings": [],
                            "unresolved_warning_error_count": 0,
                        }
                    ),
                    encoding="utf-8",
                )

        FakeProcess.state_path = self.state_path
        previous_popen = runner.subprocess.Popen
        try:
            runner.subprocess.Popen = FakeProcess
            payload = runner.run_real_session_process_supervised(
                state,
                self.state_path,
                cwd=str(self.root),
                approval_mode="auto_review",
                model=None,
                runner_lock=None,
                session_timeout_seconds=1,
            )
        finally:
            runner.subprocess.Popen = previous_popen

        updated_state = runner.load_simple_yaml(self.state_path)
        events = (self.cycle_dir / "runner-events.ndjson").read_text(encoding="utf-8")
        self.assertEqual("completed-session-after-timeout", payload["action"])
        self.assertEqual("completed-with-progress", payload["session_status"])
        self.assertEqual("semantic-review-r2", updated_state["current_stage"])
        self.assertEqual("semantic-review-ready", updated_state["stage_status"])
        self.assertEqual(2, updated_state["semantic_round"])
        self.assertIn(
            "work/review-cycles/section-scope/outputs/writer-session-log.md",
            updated_state["latest_artifacts"],
        )
        self.assertFalse((self.cycle_dir / "outputs" / "writer-r2-timeout-recovery.md").exists())
        self.assertTrue((self.cycle_dir / "outputs" / "writer-r2-timeout-after-artifacts.md").exists())
        self.assertIn("session_child_timeout_after_artifact_progress", events)

    def test_process_supervised_timeout_keeps_blocked_input_when_writer_artifacts_are_ambiguous(self) -> None:
        self.write_state(status="semantic-revision-needed", semantic_round=1)
        (self.prompt_dir / "prompt.semantic-review-r2.md").write_text(
            "Review writer-r2 changes",
            encoding="utf-8",
        )
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {
            "findings": [
                {
                    "id": "writer-quality-gate-failed",
                    "severity": "warning",
                    "path": "test-cases/1-section-scope.md",
                }
            ]
        }
        state = runner.load_simple_yaml(self.state_path)

        class FakeProcess:
            pid = 4246
            returncode = None

            def __init__(self, command, **kwargs):
                self.command = command

            def communicate(self, timeout=None):
                if timeout == 10:
                    return "", ""
                raise subprocess.TimeoutExpired(cmd=self.command, timeout=timeout)

            def kill(self):
                self.returncode = -9
                output_dir = self.state_path.parent / "outputs"
                output_dir.mkdir(parents=True, exist_ok=True)
                (output_dir / "writer-r2-response.md").write_text(
                    "# Writer Response\n\nIncomplete.\n",
                    encoding="utf-8",
                )

        FakeProcess.state_path = self.state_path
        previous_popen = runner.subprocess.Popen
        try:
            runner.subprocess.Popen = FakeProcess
            payload = runner.run_real_session_process_supervised(
                state,
                self.state_path,
                cwd=str(self.root),
                approval_mode="auto_review",
                model=None,
                runner_lock=None,
                session_timeout_seconds=1,
            )
        finally:
            runner.subprocess.Popen = previous_popen

        updated_state = runner.load_simple_yaml(self.state_path)
        events = (self.cycle_dir / "runner-events.ndjson").read_text(encoding="utf-8")
        self.assertEqual("recovered-timeout-session", payload["action"])
        self.assertEqual("failed", payload["session_status"])
        self.assertEqual("blocked-input", updated_state["stage_status"])
        self.assertTrue((self.cycle_dir / "outputs" / "writer-r2-timeout-recovery.md").exists())
        self.assertIn("session_timeout_recovered", events)

    def test_process_supervised_timeout_uses_thread_id_from_child_events(self) -> None:
        self.write_state(status="writer-draft-ready")
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {"findings": []}
        state = runner.load_simple_yaml(self.state_path)

        class FakeProcess:
            pid = 4247
            returncode = None

            def __init__(self, command, **kwargs):
                self.command = command
                runner.append_runner_event(
                    self.cycle_dir,
                    "thread_started",
                    stage="structure-preflight-r1",
                    scenario="reviewer.structure_preflight",
                    thread_id="real-reviewer-thread",
                )
                runner.append_runner_event(
                    self.cycle_dir,
                    "turn_started",
                    stage="structure-preflight-r1",
                    thread_id="real-reviewer-thread",
                )

            def communicate(self, timeout=None):
                if timeout == 10:
                    return "", ""
                raise subprocess.TimeoutExpired(cmd=self.command, timeout=timeout)

            def kill(self):
                self.returncode = -9

        FakeProcess.cycle_dir = self.cycle_dir
        previous_popen = runner.subprocess.Popen
        try:
            runner.subprocess.Popen = FakeProcess
            payload = runner.run_real_session_process_supervised(
                state,
                self.state_path,
                cwd=str(self.root),
                approval_mode="auto_review",
                model=None,
                runner_lock=None,
                session_timeout_seconds=1,
            )
        finally:
            runner.subprocess.Popen = previous_popen

        events = (self.cycle_dir / "runner-events.ndjson").read_text(encoding="utf-8")
        completion = (
            self.cycle_dir / "outputs" / "structure-preflight-r1-completion.yaml"
        ).read_text(encoding="utf-8")
        self.assertEqual("recovered-timeout-session", payload["action"])
        self.assertEqual("real-reviewer-thread", payload["thread_id"])
        self.assertIn("thread_id: real-reviewer-thread", completion)
        self.assertIn('"thread_id": "real-reviewer-thread"', events)
        self.assertIn('"child_timeout_phase": "sdk-turn-timeout-after-turn-started"', events)

    def write_stale_writer_lock(self) -> None:
        (self.cycle_dir / "runner.lock.yaml").write_text(
            "\n".join(
                [
                    "pid: 999999",
                    "command: resume",
                    f"state: {self.state_path}",
                    "started_at_epoch: 1",
                    "last_heartbeat_epoch: 1",
                    "stage: writer-r1",
                    "scenario: writer.session_initial_draft",
                    "thread_id: stale-writer-thread",
                    "status: running-session-child",
                    "",
                ]
            ),
            encoding="utf-8",
        )

    def test_resume_recovered_stale_writer_lock_advances_from_complete_artifacts_without_new_sdk_session(self) -> None:
        self.write_state(status="scope-ready-for-writer")
        output_dir = self.cycle_dir / "outputs"
        output_dir.mkdir(parents=True, exist_ok=True)
        (output_dir / "writer-r1-response.md").write_text(
            "# Writer R1 Response\n\n- result: `writer-draft-ready`\n",
            encoding="utf-8",
        )
        self.write_scoped_validator_profile()
        (self.prompt_dir / "prompt.structure-preflight-r1.md").write_text(
            "Run deterministic structure preflight.",
            encoding="utf-8",
        )
        self.write_stale_writer_lock()

        result = self.run_runner(
            "resume",
            "--state",
            str(self.state_path),
            "--recover-stale-lock",
            "--stale-lock-seconds",
            "1",
        )

        self.assertEqual(result.returncode, 0, result.stderr)
        payload = json.loads(result.stdout)
        runner = load_runner_module()
        updated_state = runner.load_simple_yaml(self.state_path)
        self.assertEqual("completed-session-after-incomplete-run", payload["action"])
        self.assertEqual("completed-with-progress", payload["session_status"])
        self.assertEqual("interrupted", payload["turn_status"])
        self.assertEqual("structure-preflight-r1", updated_state["current_stage"])
        self.assertEqual("writer-draft-ready", updated_state["stage_status"])
        self.assertFalse((self.cycle_dir / "runner.lock.yaml").exists())
        self.assertTrue((output_dir / "writer-r1-completion.yaml").exists())
        events = (self.cycle_dir / "runner-events.ndjson").read_text(encoding="utf-8")
        self.assertIn("session_incomplete_recovered_after_artifact_progress", events)
        self.assertNotIn("session_child_started", events)

    def test_resume_recovered_stale_writer_lock_blocks_when_completion_and_artifacts_missing(self) -> None:
        self.write_state(status="scope-ready-for-writer")
        self.test_case.unlink()
        self.write_stale_writer_lock()

        result = self.run_runner(
            "resume",
            "--state",
            str(self.state_path),
            "--recover-stale-lock",
            "--stale-lock-seconds",
            "1",
        )

        self.assertEqual(result.returncode, 0, result.stderr)
        payload = json.loads(result.stdout)
        runner = load_runner_module()
        updated_state = runner.load_simple_yaml(self.state_path)
        self.assertEqual("blocked-runner-completion-missing", payload["action"])
        self.assertEqual("failed", payload["session_status"])
        self.assertEqual("interrupted", payload["turn_status"])
        self.assertEqual("blocked-input", updated_state["stage_status"])
        self.assertTrue(
            any(
                reason.startswith("blocked-runner-completion-missing:")
                for reason in updated_state["blocking_reasons"]
            )
        )
        self.assertFalse((self.cycle_dir / "runner.lock.yaml").exists())
        self.assertTrue((self.cycle_dir / "outputs" / "writer-r1-completion-missing.md").exists())
        self.assertTrue((self.cycle_dir / "outputs" / "writer-r1-completion.yaml").exists())
        events = (self.cycle_dir / "runner-events.ndjson").read_text(encoding="utf-8")
        self.assertIn("session_completion_missing_blocked", events)
        self.assertNotIn("session_child_started", events)

    def test_doctor_reports_expected_completion_contract_for_next_stage(self) -> None:
        self.write_state(status="scope-ready-for-writer")
        result = self.run_runner("doctor", "--state", str(self.state_path))

        self.assertEqual(result.returncode, 0, result.stderr)
        payload = json.loads(result.stdout)
        contract = payload["completion_contract"]
        self.assertEqual("writer-r1", contract["stage"])
        self.assertEqual("writer.session_initial_draft", contract["scenario"])
        self.assertEqual(
            "work/review-cycles/section-scope/outputs/writer-r1-completion.yaml",
            contract["expected_completion_manifest"],
        )
        self.assertFalse(contract["completion_manifest_exists"])
        self.assertIn("completion_manifest", contract["missing"])
        self.assertEqual("structure-preflight-r1", contract["recoverable_writer_next_state"]["current_stage"])

    def test_effective_session_timeout_uses_stage_defaults_and_allows_override_or_disable(self) -> None:
        runner = load_runner_module()
        writer = runner.NextSession(
            stage="writer-r2",
            role="writer",
            scenario="writer.session_semantic_revision",
            prompt_path="prompt.md",
            sandbox_policy="workspace_write",
        )
        reviewer = runner.NextSession(
            stage="semantic-review-r2",
            role="reviewer",
            scenario="reviewer.semantic_traceability_test_design",
            prompt_path="prompt.md",
            sandbox_policy="workspace_write",
        )

        self.assertEqual(
            3600,
            runner.effective_session_timeout_seconds(
                writer,
                runner.AUTO_SESSION_TIMEOUT_SENTINEL,
            ),
        )
        self.assertEqual(
            1800,
            runner.effective_session_timeout_seconds(
                reviewer,
                runner.AUTO_SESSION_TIMEOUT_SENTINEL,
            ),
        )
        self.assertEqual(0, runner.effective_session_timeout_seconds(writer, 0))
        self.assertEqual(2400, runner.effective_session_timeout_seconds(writer, 2400))

    def test_terminal_signed_off_gate_accepts_structured_validator_waiver(self) -> None:
        self.write_state(status="signed-off")
        output_dir = self.cycle_dir / "outputs"
        output_dir.mkdir()
        (output_dir / "semantic-regression-final-response.md").write_text(
            "\n".join(
                [
                    "## Validator Warning Waivers",
                    "",
                    "| finding_id | path | waiver_status | waiver_class | rationale | evidence |",
                    "| --- | --- | --- | --- | --- | --- |",
                    "| `test-case-requiredness-without-empty-or-marker-check` | `test-cases/1-section-scope.md` | `false-positive` | `false-positive` | TC-001 verifies the required marker through a field-level precondition. | Evidence is the explicit TC-001 precondition and expected result. |",
                    "",
                ]
            ),
            encoding="utf-8",
        )
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {
            "findings": [
                {
                    "id": "test-case-requiredness-without-empty-or-marker-check",
                    "severity": "warning",
                    "path": "test-cases/1-section-scope.md",
                    "evidence": ["TC-001"],
                }
            ]
        }
        state = runner.load_simple_yaml(self.state_path)

        gate = runner.enforce_terminal_signed_off_validator_gate(state, self.state_path)

        self.assertEqual(0, gate["blocking_unwaived_count"])
        self.assertEqual(0, gate["invalid_waivers_count"])
        self.assertEqual(1, gate["waived_count"])
        self.assertEqual({"false-positive": 1}, gate["waived_by_class"])

    def test_terminal_signed_off_gate_rejects_waiver_without_class(self) -> None:
        self.write_state(status="signed-off")
        output_dir = self.cycle_dir / "outputs"
        output_dir.mkdir()
        (output_dir / "semantic-regression-final-response.md").write_text(
            "\n".join(
                [
                    "## Validator Warning Waivers",
                    "",
                    "| finding_id | path | waiver_status | rationale | evidence |",
                    "| --- | --- | --- | --- | --- |",
                    "| `validator-warning` | `test-cases/1-section-scope.md` | `waived` | Reviewer accepted this warning. | TC-001 evidence. |",
                    "",
                ]
            ),
            encoding="utf-8",
        )
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {
            "findings": [
                {
                    "id": "validator-warning",
                    "severity": "warning",
                    "path": "test-cases/1-section-scope.md",
                    "evidence": ["TC-001"],
                }
            ]
        }
        state = runner.load_simple_yaml(self.state_path)

        with self.assertRaisesRegex(
            runner.RunnerError,
            "invalid current-scope validator waivers",
        ):
            runner.enforce_terminal_signed_off_validator_gate(state, self.state_path)

    def test_terminal_signed_off_gate_rejects_process_only_nonblocking_waiver(self) -> None:
        self.write_state(status="signed-off")
        output_dir = self.cycle_dir / "outputs"
        output_dir.mkdir()
        (output_dir / "semantic-regression-final-response.md").write_text(
            "\n".join(
                [
                    "## Validator Warning Waivers",
                    "",
                    "| finding_id | path | waiver_status | waiver_class | rationale | evidence |",
                    "| --- | --- | --- | --- | --- | --- |",
                    "| `validator-warning` | `test-cases/1-section-scope.md` | `waived` | `accepted-nonblocking-risk` | Pre-existing warning, no regression after format-only changes. | Hash unchanged from previous round. |",
                    "",
                ]
            ),
            encoding="utf-8",
        )
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {
            "findings": [
                {
                    "id": "validator-warning",
                    "severity": "warning",
                    "path": "test-cases/1-section-scope.md",
                    "evidence": ["TC-001"],
                }
            ]
        }
        state = runner.load_simple_yaml(self.state_path)

        with self.assertRaisesRegex(
            runner.RunnerError,
            "pre-existing/no-regression/format-only",
        ):
            runner.enforce_terminal_signed_off_validator_gate(state, self.state_path)

    def test_terminal_signed_off_gate_requires_existing_gap_for_source_gap_waiver(self) -> None:
        self.write_state(status="signed-off")
        output_dir = self.cycle_dir / "outputs"
        output_dir.mkdir()
        (output_dir / "semantic-regression-final-response.md").write_text(
            "\n".join(
                [
                    "## Validator Warning Waivers",
                    "",
                    "| finding_id | path | waiver_status | waiver_class | rationale | evidence |",
                    "| --- | --- | --- | --- | --- | --- |",
                    "| `validator-warning` | `test-cases/1-section-scope.md` | `waived` | `accepted-source-gap` | Deferred because GAP-404 is unresolved. | GAP-404 tracks the missing source oracle. |",
                    "",
                ]
            ),
            encoding="utf-8",
        )
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {
            "findings": [
                {
                    "id": "validator-warning",
                    "severity": "warning",
                    "path": "test-cases/1-section-scope.md",
                    "evidence": ["TC-001"],
                }
            ]
        }
        state = runner.load_simple_yaml(self.state_path)

        with self.assertRaisesRegex(
            runner.RunnerError,
            "existing GAP",
        ):
            runner.enforce_terminal_signed_off_validator_gate(state, self.state_path)

        self.test_case.write_text(
            self.test_case.read_text(encoding="utf-8") + "\n\n## Coverage Gaps\n\n- GAP-404: missing source oracle.\n",
            encoding="utf-8",
        )
        gate = runner.enforce_terminal_signed_off_validator_gate(state, self.state_path)

        self.assertEqual(0, gate["invalid_waivers_count"])
        self.assertEqual({"accepted-source-gap": 1}, gate["waived_by_class"])

    def test_terminal_signed_off_gate_rejects_weak_schema_lag_waiver(self) -> None:
        self.write_state(status="signed-off")
        output_dir = self.cycle_dir / "outputs"
        output_dir.mkdir()
        (output_dir / "semantic-regression-final-response.md").write_text(
            "\n".join(
                [
                    "## Validator Warning Waivers",
                    "",
                    "| finding_id | path | waiver_status | waiver_class | rationale | evidence |",
                    "| --- | --- | --- | --- | --- | --- |",
                    "| `test-case-package-design-plan-missing-conditional-branch` | `test-cases/1-section-scope.md` | `waived` | `validator-schema-lag` | This is a validator schema lag. | Reviewer accepted it. |",
                    "",
                ]
            ),
            encoding="utf-8",
        )
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {
            "findings": [
                {
                    "id": "test-case-package-design-plan-missing-conditional-branch",
                    "severity": "warning",
                    "path": "test-cases/1-section-scope.md",
                    "evidence": ["PDP-041"],
                }
            ]
        }
        state = runner.load_simple_yaml(self.state_path)

        with self.assertRaisesRegex(
            runner.RunnerError,
            "validator-schema-lag waiver must cite affected",
        ):
            runner.enforce_terminal_signed_off_validator_gate(state, self.state_path)

    def test_terminal_signed_off_gate_accepts_strict_schema_lag_waiver(self) -> None:
        self.write_state(status="signed-off")
        output_dir = self.cycle_dir / "outputs"
        output_dir.mkdir()
        (output_dir / "semantic-regression-final-response.md").write_text(
            "\n".join(
                [
                    "## Validator Warning Waivers",
                    "",
                    "| finding_id | path | waiver_status | waiver_class | rationale | evidence |",
                    "| --- | --- | --- | --- | --- | --- |",
                    "| `test-case-package-design-plan-missing-conditional-branch` | `test-cases/1-section-scope.md` | `waived` | `validator-schema-lag` | Validator expects a separate inverse branch row, while the actual model represents PDP-041 as a no-blocking dependency row. | ATOM-041 is linked to TC-SAMPLE-001 with coverage_status = covered; traceability matrix has no open findings. |",
                    "",
                ]
            ),
            encoding="utf-8",
        )
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {
            "findings": [
                {
                    "id": "test-case-package-design-plan-missing-conditional-branch",
                    "severity": "warning",
                    "path": "test-cases/1-section-scope.md",
                    "evidence": ["PDP-041"],
                }
            ]
        }
        state = runner.load_simple_yaml(self.state_path)

        gate = runner.enforce_terminal_signed_off_validator_gate(state, self.state_path)

        self.assertEqual(0, gate["invalid_waivers_count"])
        self.assertEqual({"validator-schema-lag": 1}, gate["waived_by_class"])

    def test_terminal_signed_off_gate_excludes_snapshots_scratch_and_unrelated_scope(self) -> None:
        self.write_state(status="signed-off")
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {
            "findings": [
                {
                    "id": "old-snapshot-warning",
                    "severity": "warning",
                    "path": "work/review-cycles/section-scope/versions/r1/test-cases/1-section-scope.md",
                    "evidence": ["old copy"],
                },
                {
                    "id": "scratch-warning",
                    "severity": "warning",
                    "path": "work/test-design/section-scope/_artifact_write/draft.md",
                    "evidence": ["scratch"],
                },
                {
                    "id": "other-scope-warning",
                    "severity": "warning",
                    "path": "work/test-design/other-scope/traceability.md",
                    "evidence": ["other"],
                },
            ]
        }
        state = runner.load_simple_yaml(self.state_path)

        gate = runner.enforce_terminal_signed_off_validator_gate(state, self.state_path)

        self.assertEqual(0, gate["scoped_findings_count"])
        self.assertEqual(0, gate["blocking_unwaived_count"])

    def test_terminal_signed_off_gate_rejects_active_traceability_placeholder_warning(self) -> None:
        self.write_state(status="signed-off")
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {
            "findings": [
                {
                    "id": "traceability-placeholder-sentinel",
                    "severity": "warning",
                    "path": "work/review-cycles/section-scope/versions/r1/test-cases/1-section-scope.md",
                    "evidence": ["Atomic Requirements Ledger:ATOM-001:req_id=-"],
                },
                {
                    "id": "traceability-placeholder-sentinel",
                    "severity": "warning",
                    "path": "test-cases/1-section-scope.md",
                    "evidence": ["Atomic Requirements Ledger:ATOM-001:req_id=-"],
                },
            ]
        }
        state = runner.load_simple_yaml(self.state_path)

        with self.assertRaisesRegex(
            runner.RunnerError,
            "traceability-placeholder-sentinel",
        ):
            runner.enforce_terminal_signed_off_validator_gate(state, self.state_path)

    def test_terminal_signed_off_gate_excludes_snapshot_traceability_placeholder_warning(self) -> None:
        self.write_state(status="signed-off")
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {
            "findings": [
                {
                    "id": "traceability-placeholder-sentinel",
                    "severity": "warning",
                    "path": "work/review-cycles/section-scope/versions/r1/test-cases/1-section-scope.md",
                    "evidence": ["Atomic Requirements Ledger:ATOM-001:req_id=-"],
                }
            ]
        }
        state = runner.load_simple_yaml(self.state_path)

        gate = runner.enforce_terminal_signed_off_validator_gate(state, self.state_path)

        self.assertEqual(0, gate["scoped_findings_count"])
        self.assertEqual(0, gate["blocking_unwaived_count"])

    def test_writer_ready_gate_excludes_snapshot_validator_warnings(self) -> None:
        self.write_state(status="semantic-review-ready", semantic_round=2)
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {
            "findings": [
                {
                    "id": "test-case-package-design-plan-merged-numeric-class-row",
                    "severity": "warning",
                    "path": "work/review-cycles/section-scope/versions/r2/test-cases/1-section-scope.md",
                    "evidence": ["old snapshot"],
                },
                {
                    "id": "test-design-decision-table-merged-numeric-class-decision",
                    "severity": "warning",
                    "path": "work/review-cycles/section-scope/versions/r2/work/test-design/section-scope/test-design-decision-table.md",
                    "evidence": ["old snapshot"],
                },
            ]
        }
        state = runner.load_simple_yaml(self.state_path)

        gate = runner.enforce_session_ready_validator_gate(state, self.state_path)

        self.assertEqual(0, gate["scoped_findings_count"])
        self.assertEqual(0, gate["blocking_writer_ready_count"])

    def test_writer_ready_gate_rejects_current_scope_blocking_validator_warning(self) -> None:
        self.write_state(status="writer-draft-ready", semantic_round=1)
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {
            "findings": [
                {
                    "id": "test-case-mixed-schema-duplicate-fields",
                    "severity": "warning",
                    "path": "test-cases/1-section-scope.md",
                    "evidence": ["TC-SAMPLE-001"],
                }
            ]
        }
        state = runner.load_simple_yaml(self.state_path)

        with self.assertRaisesRegex(
            runner.RunnerError,
            "writer-ready state has current-scope blocking validator findings",
        ):
            runner.validate_state(state, self.state_path)

    def test_writer_ready_gate_rejects_unknown_current_scope_warning(self) -> None:
        self.write_state(status="writer-draft-ready", semantic_round=1)
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {
            "findings": [
                {
                    "id": "new-validator-warning-not-yet-in-runner-list",
                    "severity": "warning",
                    "path": "work/test-design/section-scope/new-artifact.md",
                    "evidence": ["new warning class"],
                }
            ]
        }
        state = runner.load_simple_yaml(self.state_path)

        with self.assertRaisesRegex(
            runner.RunnerError,
            "new-validator-warning-not-yet-in-runner-list",
        ):
            runner.validate_state(state, self.state_path)

    def test_writer_ready_gate_rejects_merged_numeric_class_artifact_warnings(self) -> None:
        self.write_state(status="semantic-review-ready", semantic_round=2)
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {
            "findings": [
                {
                    "id": "test-case-package-design-plan-merged-numeric-class-row",
                    "severity": "warning",
                    "path": "test-cases/1-section-scope.md",
                    "evidence": ["PD-007"],
                },
                {
                    "id": "test-design-decision-table-merged-numeric-class-decision",
                    "severity": "warning",
                    "path": "work/test-design/section-scope/test-design-decision-table.md",
                    "evidence": ["TDD-008"],
                },
            ]
        }
        state = runner.load_simple_yaml(self.state_path)

        with self.assertRaisesRegex(
            runner.RunnerError,
            "writer-ready state has current-scope blocking validator findings: 2 total",
        ):
            runner.validate_state(state, self.state_path)

    def test_writer_ready_gate_rejects_v8_quality_regression_validator_warning(self) -> None:
        self.write_state(status="semantic-review-ready", semantic_round=1)
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {
            "findings": [
                {
                    "id": "test-case-action-created-block-without-cleanup",
                    "severity": "warning",
                    "path": "test-cases/1-section-scope.md",
                    "evidence": ["TC-SAMPLE-001"],
                }
            ]
        }
        state = runner.load_simple_yaml(self.state_path)

        with self.assertRaisesRegex(
            runner.RunnerError,
            "writer-ready state has current-scope blocking validator findings",
        ):
            runner.validate_state(state, self.state_path)

    def test_writer_ready_gate_rejects_v9_design_and_fixture_regression_warnings(self) -> None:
        self.write_state(status="writer-draft-ready", semantic_round=1)
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {
            "findings": [
                {
                    "id": "test-case-negative-transition-without-valid-fixture-smell",
                    "severity": "warning",
                    "path": "test-cases/1-section-scope.md",
                    "evidence": ["TC-SAMPLE-014"],
                },
                {
                    "id": "test-case-requiredness-without-empty-or-marker-check",
                    "severity": "warning",
                    "path": "test-cases/1-section-scope.md",
                    "evidence": ["TC-SAMPLE-023"],
                },
                {
                    "id": "test-design-decision-table-metadata-behavior-smell",
                    "severity": "warning",
                    "path": "test-cases/1-section-scope.md",
                    "evidence": ["TDD-001"],
                },
                {
                    "id": "test-design-decision-table-metadata-cross-section-conflict",
                    "severity": "warning",
                    "path": "test-cases/1-section-scope.md",
                    "evidence": ["TDD-033"],
                },
            ]
        }
        state = runner.load_simple_yaml(self.state_path)

        with self.assertRaisesRegex(
            runner.RunnerError,
            "writer-ready state has current-scope blocking validator findings: 4 total",
        ):
            runner.validate_state(state, self.state_path)

    def test_writer_ready_gate_rejects_v10_artifact_shape_regression_warnings(self) -> None:
        self.write_state(status="writer-draft-ready", semantic_round=1)
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {
            "findings": [
                {
                    "id": "source-row-inventory-no-table",
                    "severity": "warning",
                    "path": "test-cases/1-section-scope.md",
                    "evidence": ["Source Row Inventory not parsed as table."],
                },
                {
                    "id": "coverage-obligation-table-no-table",
                    "severity": "warning",
                    "path": "test-cases/1-section-scope.md",
                    "evidence": ["Coverage Obligation Table not parsed as table."],
                },
                {
                    "id": "test-design-review-no-table",
                    "severity": "warning",
                    "path": "work/test-design/section-scope/test-design-review.md",
                    "evidence": ["Test Design Review not parsed as table."],
                },
                {
                    "id": "test-case-split-artifact-duplicated-sections",
                    "severity": "warning",
                    "path": "test-cases/1-section-scope.md",
                    "evidence": ["Canonical TC duplicates split artifact sections."],
                },
                {
                    "id": "writer-quality-gate-missing-columns",
                    "severity": "warning",
                    "path": "work/test-design/section-scope/writer-quality-gate.md",
                    "evidence": ["Missing gate_item/affected_package/required_action."],
                },
            ]
        }
        state = runner.load_simple_yaml(self.state_path)

        with self.assertRaisesRegex(
            runner.RunnerError,
            "writer-ready state has current-scope blocking validator findings: 5 total",
        ):
            runner.validate_state(state, self.state_path)

    def test_writer_ready_gate_rejects_v12_table_shape_and_parseability_warnings(self) -> None:
        self.write_state(status="writer-draft-ready", semantic_round=1)
        runner = load_runner_module()
        finding_ids = [
            "source-table-normalization-missing-columns",
            "test-design-decision-table-missing-columns",
            "test-case-package-design-plan-missing-columns",
            "source-row-completeness-matrix-missing",
            "source-row-inventory-invalid-in-scope",
            "source-row-gsr-count-mismatch",
            "source-normalization-unmapped-property",
            "test-case-missing-package-id",
            "test-case-action-treated-as-required-field-smell",
            "test-case-package-design-plan-merged-check-smell",
            "test-case-package-design-plan-missing-atoms",
            "test-case-package-design-plan-negative-without-positive-acceptance",
            "test-design-decision-table-invalid-decision",
            "test-design-review-failed",
            "test-design-review-missing-required-items",
        ]
        runner.run_agent_artifact_validator = lambda ft_root: {
            "findings": [
                {
                    "id": finding_id,
                    "severity": "warning",
                    "path": "work/test-design/section-scope/source-table-normalization.md",
                    "evidence": [finding_id],
                }
                for finding_id in finding_ids
            ]
        }
        state = runner.load_simple_yaml(self.state_path)

        with self.assertRaisesRegex(
            runner.RunnerError,
            "writer-ready state has current-scope blocking validator findings: 15 total",
        ):
            runner.validate_state(state, self.state_path)

    def test_writer_ready_gate_rejects_v13_cross_artifact_status_warnings(self) -> None:
        self.write_state(status="writer-draft-ready", semantic_round=1)
        runner = load_runner_module()
        finding_ids = [
            "coverage-obligation-table-duplicate-class",
            "coverage-obligation-table-invalid-status",
            "test-design-decision-table-overbroad-gap-smell",
            "test-design-decision-table-standalone-without-tc-or-oracle",
            "test-design-review-invalid-blocks-value",
            "test-design-review-invalid-severity",
            "test-design-review-invalid-status",
            "writer-self-check-empty-section",
        ]
        runner.run_agent_artifact_validator = lambda ft_root: {
            "findings": [
                {
                    "id": finding_id,
                    "severity": "warning",
                    "path": "work/test-design/section-scope/test-design-review.md",
                    "evidence": [finding_id],
                }
                for finding_id in finding_ids
            ]
        }
        state = runner.load_simple_yaml(self.state_path)

        with self.assertRaisesRegex(
            runner.RunnerError,
            "writer-ready state has current-scope blocking validator findings: 8 total",
        ):
            runner.validate_state(state, self.state_path)

    def test_writer_blocked_input_rejects_missing_post_write_validator_run(self) -> None:
        self.write_state(status="blocked-input", semantic_round=0)
        self.state_path.write_text(
            self.state_path.read_text(encoding="utf-8")
            + "\n".join(
                [
                    "blocking_reasons:",
                    "- Validator has not been run after initial artifact generation.",
                    "",
                ]
            ),
            encoding="utf-8",
        )
        runner = load_runner_module()
        state = runner.load_simple_yaml(self.state_path)

        with self.assertRaisesRegex(
            runner.RunnerError,
            "writer blocked-input cannot be caused by missing post-write validator run",
        ):
            runner.validate_state(state, self.state_path)

    def test_run_until_terminal_dry_run_reports_first_stage_and_loop_intent(self) -> None:
        self.write_state(status="scope-ready-for-writer", semantic_round=0)
        self.test_case.unlink()

        result = self.run_runner(
            "run-until-terminal",
            "--state",
            str(self.state_path),
            "--dry-run",
        )

        self.assertEqual(result.returncode, 0, result.stderr)
        payload = json.loads(result.stdout)
        self.assertEqual("dry-run-chain", payload["action"])
        self.assertEqual(12, payload["max_sessions"])
        self.assertEqual("writer-r1", payload["next"]["stage"])
        self.assertIn("first runnable stage", payload["note"])

    def test_run_until_terminal_reloads_state_until_terminal_status(self) -> None:
        self.write_state(status="scope-ready-for-writer", semantic_round=0)
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {"findings": []}
        calls: list[str] = []

        def fake_run_real_session(state, state_path, *, cwd, approval_mode, model, runner_lock=None):
            calls.append(state["stage_status"])
            updated = self.state_path.read_text(encoding="utf-8").replace(
                "stage_status: scope-ready-for-writer",
                "stage_status: signed-off",
            )
            self.state_path.write_text(updated, encoding="utf-8")
            return {
                "action": "completed-session",
                "stage": "writer-r1",
                "role": "writer",
                "scenario": "writer.session_initial_draft",
                "turn_status": "completed",
            }

        runner.run_real_session = fake_run_real_session
        payload = runner.run_until_terminal(
            self.state_path,
            cwd=None,
            approval_mode="auto_review",
            model=None,
            max_sessions=3,
        )

        self.assertEqual(["scope-ready-for-writer"], calls)
        self.assertEqual("completed-chain", payload["action"])
        self.assertEqual("signed-off", payload["final_status"])
        self.assertTrue(payload["terminal"])
        self.assertEqual(1, payload["sessions_started"])

    def test_run_until_terminal_accepts_terminal_after_last_allowed_session(self) -> None:
        self.write_state(status="scope-ready-for-writer", semantic_round=0)
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {"findings": []}

        def fake_run_real_session(state, state_path, *, cwd, approval_mode, model, runner_lock=None):
            updated = self.state_path.read_text(encoding="utf-8").replace(
                "stage_status: scope-ready-for-writer",
                "stage_status: signed-off",
            )
            self.state_path.write_text(updated, encoding="utf-8")
            return {
                "action": "completed-session",
                "stage": "writer-r1",
                "role": "writer",
                "scenario": "writer.session_initial_draft",
                "turn_status": "completed",
            }

        runner.run_real_session = fake_run_real_session
        payload = runner.run_until_terminal(
            self.state_path,
            cwd=None,
            approval_mode="auto_review",
            model=None,
            max_sessions=1,
        )

        self.assertEqual("completed-chain", payload["action"])
        self.assertEqual("signed-off", payload["final_status"])
        self.assertTrue(payload["terminal"])
        self.assertEqual(1, payload["sessions_started"])

    def test_run_until_terminal_reports_max_sessions_after_nonterminal_progress(self) -> None:
        self.write_state(status="scope-ready-for-writer", semantic_round=0)
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {"findings": []}

        def fake_run_real_session(state, state_path, *, cwd, approval_mode, model, runner_lock=None):
            updated = self.state_path.read_text(encoding="utf-8").replace(
                "stage_status: scope-ready-for-writer",
                "stage_status: writer-draft-ready",
            )
            updated = updated.replace(
                "active_transition_prompt: work/review-cycles/section-scope/prompts/next.md",
                "active_transition_prompt: work/review-cycles/section-scope/prompts/next.md",
            )
            self.state_path.write_text(updated, encoding="utf-8")
            return {
                "action": "completed-session",
                "stage": "writer-r1",
                "role": "writer",
                "scenario": "writer.session_initial_draft",
                "turn_status": "completed",
                "session_status": "completed",
            }

        runner.run_real_session = fake_run_real_session
        payload = runner.run_until_terminal(
            self.state_path,
            cwd=None,
            approval_mode="auto_review",
            model=None,
            max_sessions=1,
        )

        self.assertEqual("max-sessions-reached", payload["action"])
        self.assertEqual("writer-draft-ready", payload["final_status"])
        self.assertFalse(payload["terminal"])
        self.assertEqual(1, payload["sessions_started"])
        self.assertEqual("structure-preflight-r1", payload["next"]["stage"])

    def test_run_until_terminal_rejects_semantic_revision_writer_with_draft_ready_status(self) -> None:
        self.write_state(status="semantic-revision-needed", semantic_round=1)
        runner = load_runner_module()

        def fake_run_real_session(state, state_path, *, cwd, approval_mode, model, runner_lock=None):
            updated = self.state_path.read_text(encoding="utf-8").replace(
                "stage_status: semantic-revision-needed",
                "stage_status: writer-draft-ready",
            )
            self.state_path.write_text(updated, encoding="utf-8")
            return {
                "action": "completed-session",
                "stage": "writer-r2",
                "role": "writer",
                "scenario": "writer.session_semantic_revision",
                "turn_status": "completed",
                "session_status": "completed",
            }

        runner.run_real_session = fake_run_real_session
        with self.assertRaisesRegex(
            runner.RunnerError,
            "writer-r2.*invalid post-session stage_status='writer-draft-ready'",
        ):
            runner.run_until_terminal(
                self.state_path,
                cwd=None,
                approval_mode="auto_review",
                model=None,
                max_sessions=3,
            )

    def test_run_until_terminal_rejects_completed_stage_without_state_progress(self) -> None:
        self.write_state(status="scope-ready-for-writer", semantic_round=0)
        runner = load_runner_module()

        def fake_run_real_session(state, state_path, *, cwd, approval_mode, model, runner_lock=None):
            return {
                "action": "completed-session",
                "stage": "writer-r1",
                "role": "writer",
                "scenario": "writer.session_initial_draft",
                "turn_status": "completed",
            }

        runner.run_real_session = fake_run_real_session
        with self.assertRaisesRegex(runner.RunnerError, "did not advance"):
            runner.run_until_terminal(
                self.state_path,
                cwd=None,
                approval_mode="auto_review",
                model=None,
                max_sessions=3,
            )

    def test_run_until_terminal_accepts_interrupted_stage_when_state_advances(self) -> None:
        self.write_state(status="scope-ready-for-writer", semantic_round=0)
        runner = load_runner_module()
        runner.run_agent_artifact_validator = lambda ft_root: {"findings": []}

        def fake_run_real_session(state, state_path, *, cwd, approval_mode, model, runner_lock=None):
            updated = self.state_path.read_text(encoding="utf-8").replace(
                "stage_status: scope-ready-for-writer",
                "stage_status: signed-off",
            )
            self.state_path.write_text(updated, encoding="utf-8")
            return {
                "action": "completed-session",
                "stage": "writer-r1",
                "role": "writer",
                "scenario": "writer.session_initial_draft",
                "turn_status": "interrupted",
                "session_status": "completed-with-progress",
            }

        runner.run_real_session = fake_run_real_session
        payload = runner.run_until_terminal(
            self.state_path,
            cwd=None,
            approval_mode="auto_review",
            model=None,
            max_sessions=1,
        )

        self.assertEqual("completed-chain", payload["action"])
        self.assertEqual("signed-off", payload["final_status"])
        self.assertTrue(payload["terminal"])
        self.assertEqual("completed-with-progress", payload["sessions"][0]["session_status"])

    def test_run_until_terminal_rejects_interrupted_stage_without_state_progress(self) -> None:
        self.write_state(status="scope-ready-for-writer", semantic_round=0)
        runner = load_runner_module()

        def fake_run_real_session(state, state_path, *, cwd, approval_mode, model, runner_lock=None):
            return {
                "action": "completed-session",
                "stage": "writer-r1",
                "role": "writer",
                "scenario": "writer.session_initial_draft",
                "turn_status": "interrupted",
                "session_status": "failed",
            }

        runner.run_real_session = fake_run_real_session
        with self.assertRaisesRegex(runner.RunnerError, "did not advance"):
            runner.run_until_terminal(
                self.state_path,
                cwd=None,
                approval_mode="auto_review",
                model=None,
                max_sessions=1,
            )

    def test_classify_session_status_accepts_interrupted_only_with_progress(self) -> None:
        runner = load_runner_module()

        self.assertEqual(
            "completed-with-progress",
            runner.classify_session_status("interrupted", state_advanced=True),
        )
        self.assertEqual(
            "failed",
            runner.classify_session_status("interrupted", state_advanced=False),
        )
        self.assertEqual(
            "failed",
            runner.classify_session_status("failed", state_advanced=True),
        )

    def test_run_real_session_records_started_thread_before_sdk_run_exception(self) -> None:
        self.write_state(status="scope-ready-for-writer", semantic_round=0)
        runner = load_runner_module()

        class FakeSandbox:
            read_only = "read_only"
            workspace_write = "workspace_write"
            full_access = "full_access"

        class FakeApprovalMode:
            auto_review = "auto_review"
            deny_all = "deny_all"

        class FakeThread:
            id = "thread-started-before-error"

            def set_name(self, name: str) -> None:
                self.name = name

            def run(self, *args, **kwargs):
                raise RuntimeError("sdk run failed")

        class FakeCodex:
            def thread_start(self, *args, **kwargs):
                return FakeThread()

            def close(self) -> None:
                pass

        old_module = sys.modules.get("openai_codex")
        sys.modules["openai_codex"] = type(
            "FakeOpenAICodex",
            (),
            {
                "Codex": FakeCodex,
                "Sandbox": FakeSandbox,
                "ApprovalMode": FakeApprovalMode,
            },
        )
        try:
            with self.assertRaisesRegex(RuntimeError, "sdk run failed"):
                runner.run_real_session(
                    runner.load_simple_yaml(self.state_path),
                    self.state_path,
                    cwd=None,
                    approval_mode="auto_review",
                    model=None,
                )
        finally:
            if old_module is None:
                sys.modules.pop("openai_codex", None)
            else:
                sys.modules["openai_codex"] = old_module

        session_map = (self.cycle_dir / "codex-session-map.yaml").read_text(encoding="utf-8")
        self.assertIn("thread_id: thread-started-before-error", session_map)
        self.assertIn("status: started", session_map)
        self.assertIn("status: failed", session_map)
        self.assertIn("error: RuntimeError", session_map)

    def test_run_real_session_records_completed_with_progress_for_interrupted_turn(self) -> None:
        self.write_state(status="scope-ready-for-writer", semantic_round=0)
        runner = load_runner_module()

        class FakeSandbox:
            read_only = "read_only"
            workspace_write = "workspace_write"
            full_access = "full_access"

        class FakeApprovalMode:
            auto_review = "auto_review"
            deny_all = "deny_all"

        class FakeTurn:
            id = "turn-interrupted-after-progress"
            status = "interrupted"
            final_response = "stage wrote artifacts and advanced state"
            duration_ms = 123

        class FakeThread:
            id = "thread-interrupted-after-progress"

            def set_name(self, name: str) -> None:
                self.name = name

            def run(self, *args, **kwargs):
                updated = self_state_path.read_text(encoding="utf-8").replace(
                    "stage_status: scope-ready-for-writer",
                    "stage_status: signed-off",
                )
                self_state_path.write_text(updated, encoding="utf-8")
                return FakeTurn()

        class FakeCodex:
            def thread_start(self, *args, **kwargs):
                return FakeThread()

            def close(self) -> None:
                pass

        self_state_path = self.state_path
        old_module = sys.modules.get("openai_codex")
        sys.modules["openai_codex"] = type(
            "FakeOpenAICodex",
            (),
            {
                "Codex": FakeCodex,
                "Sandbox": FakeSandbox,
                "ApprovalMode": FakeApprovalMode,
            },
        )
        try:
            payload = runner.run_real_session(
                runner.load_simple_yaml(self.state_path),
                self.state_path,
                cwd=None,
                approval_mode="auto_review",
                model=None,
            )
        finally:
            if old_module is None:
                sys.modules.pop("openai_codex", None)
            else:
                sys.modules["openai_codex"] = old_module

        self.assertEqual("interrupted", payload["turn_status"])
        self.assertEqual("completed-with-progress", payload["session_status"])
        self.assertTrue(payload["state_advanced"])
        completion_manifest = self.cycle_dir / "outputs" / "writer-r1-completion.yaml"
        self.assertEqual(str(completion_manifest), payload["completion_manifest"])
        self.assertTrue(completion_manifest.exists())
        completion_content = completion_manifest.read_text(encoding="utf-8")
        self.assertIn("turn_status: interrupted", completion_content)
        self.assertIn("session_status: completed-with-progress", completion_content)
        self.assertIn("state_advanced: true", completion_content)

        event_path = self.cycle_dir / "runner-events.ndjson"
        self.assertTrue(event_path.exists())
        event_names = [
            json.loads(line)["event"]
            for line in event_path.read_text(encoding="utf-8").splitlines()
            if line.strip()
        ]
        self.assertIn("thread_started", event_names)
        self.assertIn("turn_finished", event_names)
        self.assertIn("stage_completed", event_names)

        snapshot_manifest = (
            self.cycle_dir / "versions" / "after-writer-r1" / "snapshot-manifest.yaml"
        ).read_text(encoding="utf-8")
        self.assertIn("work/review-cycles/section-scope/runner-events.ndjson", snapshot_manifest)
        self.assertIn(
            "work/review-cycles/section-scope/outputs/writer-r1-completion.yaml",
            snapshot_manifest,
        )
        session_map = (self.cycle_dir / "codex-session-map.yaml").read_text(encoding="utf-8")
        self.assertIn("turn_status: interrupted", session_map)
        self.assertIn("status: completed-with-progress", session_map)
        self.assertIn("state_advanced: true", session_map)
        self.assertIn("completion_manifest: work/review-cycles/section-scope/outputs/writer-r1-completion.yaml", session_map)

    def test_session_prompt_keeps_session_map_runner_owned(self) -> None:
        self.write_state(status="scope-ready-for-writer", semantic_round=0)
        runner = load_runner_module()
        state = runner.load_simple_yaml(self.state_path)
        next_session = runner.next_session_for_state(state)

        prompt = runner.prompt_text_for_session(state, self.state_path, next_session)

        self.assertIn("Do not edit codex-session-map.yaml", prompt)

    def test_session_prompt_requires_instruction_context_loading(self) -> None:
        self.write_state(status="semantic-review-ready", semantic_round=1)
        runner = load_runner_module()
        state = runner.load_simple_yaml(self.state_path)
        next_session = runner.next_session_for_state(state)

        prompt = runner.prompt_text_for_session(state, self.state_path, next_session)

        self.assertIn("Instruction loading contract:", prompt)
        self.assertIn(
            "python scripts/resolve_instruction_context.py --scenario reviewer.semantic_traceability_test_design --budget-report --fail-on-budget",
            prompt,
        )
        self.assertIn("Read every selected required file", prompt)
        self.assertIn("references/agent/test-design-defect-taxonomy.md", prompt)
        self.assertIn("references/agent/package-test-design-plan-format.md", prompt)
        self.assertIn("record the resolver command, budget status, and selected files", prompt)

    def test_runner_lock_blocks_parallel_run_on_same_state(self) -> None:
        self.write_state(status="scope-ready-for-writer", semantic_round=0)
        runner = load_runner_module()

        with runner.RunnerFileLock(self.state_path, command="run-until-terminal"):
            with self.assertRaisesRegex(runner.RunnerError, "active runner lock exists"):
                with runner.RunnerFileLock(self.state_path, command="run-until-terminal"):
                    pass

    def test_doctor_reports_terminal_state_without_lock(self) -> None:
        self.write_state(status="signed-off", semantic_round=2)

        result = self.run_runner("doctor", "--state", str(self.state_path))

        self.assertEqual(result.returncode, 0, result.stderr)
        payload = json.loads(result.stdout)
        self.assertEqual("doctor", payload["action"])
        self.assertTrue(payload["state_valid"])
        self.assertFalse(payload["lock"]["exists"])
        self.assertEqual("run-validate-terminal-gate", payload["recommendation"])
        self.assertFalse(payload["terminal_validator_gate"]["checked"])
        self.assertIn("run `validate`", payload["terminal_validator_gate"]["reason"])

    def test_resume_dry_run_reports_doctor_payload_without_instruction_resolution(self) -> None:
        self.write_state(status="scope-ready-for-writer", semantic_round=0)
        self.test_case.unlink()

        result = self.run_runner("resume", "--state", str(self.state_path), "--dry-run")

        self.assertEqual(result.returncode, 0, result.stderr)
        payload = json.loads(result.stdout)
        self.assertEqual("doctor", payload["action"])
        self.assertTrue(payload["state_valid"])
        self.assertEqual("writer-r1", payload["next"]["stage"])
        self.assertEqual("run-next-stage", payload["recommendation"])

    def test_doctor_reports_stale_dead_lock_as_recoverable(self) -> None:
        self.write_state(status="scope-ready-for-writer", semantic_round=0)
        lock_path = self.cycle_dir / "runner.lock.yaml"
        lock_path.write_text(
            "\n".join(
                [
                    "version: 1",
                    "pid: 99999999",
                    "command: run-until-terminal",
                    f"state: {self.state_path}",
                    "started_at_epoch: 1",
                    "last_heartbeat_epoch: 1",
                    "stage: semantic-review-r1",
                    "scenario: reviewer.semantic_traceability_test_design",
                    "thread_id: stale-thread-id",
                    "status: running-session",
                    "",
                ]
            ),
            encoding="utf-8",
        )

        result = self.run_runner(
            "doctor",
            "--state",
            str(self.state_path),
            "--stale-lock-seconds",
            "1",
        )

        self.assertEqual(result.returncode, 0, result.stderr)
        payload = json.loads(result.stdout)
        self.assertTrue(payload["lock"]["exists"])
        self.assertTrue(payload["lock"]["stale"])
        self.assertFalse(payload["lock"]["pid_alive"])
        self.assertTrue(payload["lock"]["safe_to_recover"])
        self.assertEqual("resume-with---recover-stale-lock", payload["recommendation"])

    def test_runner_lock_recover_stale_lock_archives_and_records_abort(self) -> None:
        self.write_state(status="scope-ready-for-writer", semantic_round=0)
        runner = load_runner_module()
        lock_path = self.cycle_dir / runner.RUNNER_LOCK_FILE
        lock_path.write_text(
            "\n".join(
                [
                    "version: 1",
                    "pid: 999999",
                    "command: run-until-terminal",
                    f"state: {self.state_path}",
                    "started_at_epoch: 1",
                    "last_heartbeat_epoch: 1",
                    "stage: semantic-review-r1",
                    "scenario: reviewer.semantic_traceability_test_design",
                    "thread_id: stale-thread-id",
                    "status: running-session",
                    "",
                ]
            ),
            encoding="utf-8",
        )

        with runner.RunnerFileLock(
            self.state_path,
            command="run-until-terminal",
            stale_lock_seconds=0,
            recover_stale_lock=True,
        ) as lock:
            self.assertEqual(1, lock.stale_lock_seconds)
            self.assertTrue(lock_path.exists())

        archives = list(self.cycle_dir.glob("runner.lock.recovered-*.yaml"))
        self.assertEqual(1, len(archives))
        session_map = (self.cycle_dir / "codex-session-map.yaml").read_text(encoding="utf-8")
        self.assertIn("thread_id: stale-thread-id", session_map)
        self.assertIn("status: aborted", session_map)
        self.assertIn("abort_reason: stale runner lock recovered", session_map)
        self.assertIn("recovered_pid: 999999", session_map)
        self.assertIn("recovered_command: run-until-terminal", session_map)
        self.assertIn("recovered_status: running-session", session_map)
        self.assertIn("recovered_last_heartbeat_epoch: 1", session_map)
        events = [
            json.loads(line)["event"]
            for line in (self.cycle_dir / "runner-events.ndjson").read_text(encoding="utf-8").splitlines()
            if line.strip()
        ]
        self.assertIn("lock_recovered", events)
        self.assertIn("lock_acquired", events)
        self.assertIn("lock_released", events)

    def test_resume_recover_only_archives_stale_lock_without_starting_session(self) -> None:
        self.write_state(status="scope-ready-for-writer", semantic_round=0)
        lock_path = self.cycle_dir / "runner.lock.yaml"
        lock_path.write_text(
            "\n".join(
                [
                    "version: 1",
                    "pid: 999999",
                    "command: run-until-terminal",
                    f"state: {self.state_path}",
                    "started_at_epoch: 1",
                    "last_heartbeat_epoch: 1",
                    "stage: semantic-review-r1",
                    "scenario: reviewer.semantic_traceability_test_design",
                    "thread_id: stale-thread-id",
                    "status: running-session",
                    "",
                ]
            ),
            encoding="utf-8",
        )

        result = self.run_runner(
            "resume",
            "--state",
            str(self.state_path),
            "--stale-lock-seconds",
            "1",
            "--recover-stale-lock",
            "--recover-only",
        )

        self.assertEqual(result.returncode, 0, result.stderr)
        payload = json.loads(result.stdout)
        self.assertEqual("recovered-stale-lock", payload["action"])
        self.assertFalse(payload["lock"]["exists"])
        self.assertEqual("run-next-stage", payload["recommendation"])
        self.assertIn("did not run validator or start the next session", payload["note"])
        self.assertFalse(lock_path.exists())
        archives = list(self.cycle_dir.glob("runner.lock.recovered-*.yaml"))
        self.assertEqual(1, len(archives))
        session_map = (self.cycle_dir / "codex-session-map.yaml").read_text(encoding="utf-8")
        self.assertIn("status: aborted", session_map)
        self.assertIn("abort_reason: stale runner lock recovered", session_map)
        self.assertNotIn("status: started", session_map)
        events = [
            json.loads(line)["event"]
            for line in (self.cycle_dir / "runner-events.ndjson").read_text(encoding="utf-8").splitlines()
            if line.strip()
        ]
        self.assertIn("lock_recovered", events)
        self.assertNotIn("scoped_validator_profile_written", events)

    def test_start_preflight_fails_before_lock_when_sdk_runtime_missing(self) -> None:
        self.write_state(status="scope-ready-for-writer", semantic_round=0)
        runner = load_runner_module()

        def missing_sdk() -> None:
            raise runner.RunnerError("openai-codex is not installed in this Python environment")

        args = types.SimpleNamespace(
            state=str(self.state_path),
            dry_run=False,
            command="start",
            stale_lock_seconds=1,
            recover_stale_lock=False,
            cwd=str(self.root),
            approval_mode="auto_review",
            model=None,
            session_timeout_seconds=runner.AUTO_SESSION_TIMEOUT_SENTINEL,
        )
        previous_preflight = runner.ensure_openai_codex_runtime_available
        runner.ensure_openai_codex_runtime_available = missing_sdk
        try:
            with self.assertRaisesRegex(runner.RunnerError, "openai-codex is not installed"):
                runner.cmd_start_or_continue(args)
        finally:
            runner.ensure_openai_codex_runtime_available = previous_preflight

        self.assertFalse((self.cycle_dir / "runner.lock.yaml").exists())
        self.assertFalse((self.cycle_dir / "runner-events.ndjson").exists())

    def test_resume_recover_only_skips_sdk_preflight(self) -> None:
        self.write_state(status="scope-ready-for-writer", semantic_round=0)
        lock_path = self.cycle_dir / "runner.lock.yaml"
        lock_path.write_text(
            "\n".join(
                [
                    "version: 1",
                    "pid: 999999",
                    "command: run-until-terminal",
                    f"state: {self.state_path}",
                    "started_at_epoch: 1",
                    "last_heartbeat_epoch: 1",
                    "stage: semantic-review-r1",
                    "scenario: reviewer.semantic_traceability_test_design",
                    "thread_id: stale-thread-id",
                    "status: running-session",
                    "",
                ]
            ),
            encoding="utf-8",
        )
        runner = load_runner_module()

        def missing_sdk() -> None:
            raise runner.RunnerError("openai-codex is not installed in this Python environment")

        args = types.SimpleNamespace(
            state=str(self.state_path),
            dry_run=False,
            command="resume",
            stale_lock_seconds=1,
            recover_stale_lock=True,
            recover_only=True,
            cwd=str(self.root),
            approval_mode="auto_review",
            model=None,
            max_sessions=1,
            session_timeout_seconds=runner.AUTO_SESSION_TIMEOUT_SENTINEL,
        )
        previous_preflight = runner.ensure_openai_codex_runtime_available
        runner.ensure_openai_codex_runtime_available = missing_sdk
        try:
            result = runner.cmd_resume(args)
        finally:
            runner.ensure_openai_codex_runtime_available = previous_preflight

        self.assertEqual(0, result)
        self.assertFalse(lock_path.exists())
        self.assertEqual(1, len(list(self.cycle_dir.glob("runner.lock.recovered-*.yaml"))))

    def test_runner_lock_updates_stage_and_thread_heartbeat(self) -> None:
        self.write_state(status="scope-ready-for-writer", semantic_round=0)
        runner = load_runner_module()
        lock_path = self.cycle_dir / runner.RUNNER_LOCK_FILE

        with runner.RunnerFileLock(self.state_path, command="run-until-terminal") as lock:
            lock.update(
                stage="semantic-review-r1",
                scenario="reviewer.semantic_traceability_test_design",
                thread_id="thread-123",
            )
            lock_data = runner.load_simple_yaml(lock_path)
            self.assertEqual("semantic-review-r1", lock_data["stage"])
            self.assertEqual("thread-123", lock_data["thread_id"])

        self.assertFalse(lock_path.exists())


if __name__ == "__main__":
    unittest.main()
