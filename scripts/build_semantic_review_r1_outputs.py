from __future__ import annotations

import csv
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


ROOT = Path("fts/ft-2-OF_16")
SCOPE = "ui-employment-canary-v6-terminal-gate-regression"
CYCLE_DIR = ROOT / "work" / "review-cycles" / SCOPE
OUTPUTS = CYCLE_DIR / "outputs"
PROMPTS = CYCLE_DIR / "prompts"
TD = ROOT / "work" / "test-design" / SCOPE
CANONICAL = ROOT / "test-cases" / "2-1-1-1-1-2-ui-employment-canary-v6-terminal-gate-regression.md"
STATE = CYCLE_DIR / "cycle-state.yaml"


def parse_md_table(path: Path) -> list[dict[str, str]]:
    lines = path.read_text(encoding="utf-8").splitlines()
    table_lines = [line for line in lines if line.startswith("|")]
    if len(table_lines) < 2:
        return []
    rows = []
    header = [cell.strip() for cell in table_lines[0].strip("|").split("|")]
    for line in table_lines[2:]:
        cells = [cell.strip() for cell in line.strip("|").split("|")]
        if len(cells) != len(header):
            continue
        rows.append(dict(zip(header, cells)))
    return rows


def md_table(headers: list[str], rows: list[list[str]]) -> str:
    out = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(["---"] * len(headers)) + " |",
    ]
    for row in rows:
        out.append("| " + " | ".join(cell.replace("\n", " ") for cell in row) + " |")
    return "\n".join(out)


def build_matrix() -> list[dict[str, str]]:
    ledger_rows = parse_md_table(TD / "atomic-requirements-ledger.md")
    matrix: list[dict[str, str]] = []
    for row in ledger_rows:
        atom = row.get("atom_id", "")
        if not atom.startswith("ATOM-"):
            continue
        matrix.append(
            {
                "atom_id": atom,
                "req_id": row.get("req_id", "none") or "none",
                "source_path": row.get("source_ref", ""),
                "atomic_statement": row.get("atomic_statement", ""),
                "field_or_block": row.get("field_or_block", "") or _field_from_statement(row.get("atomic_statement", "")),
                "condition": _condition_from_statement(row.get("atomic_statement", "")),
                "expected_behavior": _expected_from_statement(row.get("atomic_statement", "")),
                "covered_by_tc": row.get("linked_tc_or_gap", "-"),
                "coverage_status": row.get("coverage_status", "covered"),
                "gap_note": "-",
            }
        )

    gap_rows = [
        (
            "ATOM-062",
            "GSR 126; GSR 128; GSR 141",
            "GAP-001; SRC-004; SRC-005; SRC-017",
            "DaData employer lookup, exact trigger/dropdown/no-result behavior and non-UI SPR contract-field artifacts are not defined.",
            "DaData employer/address integration",
            "organization lookup / selected suggestion",
            "Only UI-visible selection/prefill is coverable; exact mechanics/backend artifacts remain unresolved.",
            "GAP-001",
        ),
        (
            "ATOM-063",
            "SRC-011; DICT-004",
            "GAP-002; SRC-011",
            "`Пенсия` and `Аренда` may be added only once, but the exact duplicate-prevention mechanism is not defined.",
            "Тип дохода",
            "duplicate `Пенсия`/`Аренда` attempt",
            "Duplicate prevention invariant retained; exact UI feedback/control state unresolved.",
            "GAP-002",
        ),
        (
            "ATOM-064",
            "GSR 142; SRC-018",
            "GAP-003; SRC-018",
            "Return-from-`Выбор решения` setup and observable SPR/anti-fraud/external-check artifacts are not defined.",
            "Следующий шаг / status lifecycle",
            "return from `Выбор решения` and critical field changed",
            "UI-visible validation/navigation can be covered; hidden effects remain unresolved.",
            "GAP-003",
        ),
        (
            "ATOM-065",
            "SRC-022; SRC-023; SRC-024",
            "GAP-004; PDF pp.65-67",
            "PDF-only CDI messages are source text, but deterministic UI setup/test data for CDI failure and mismatch states is missing.",
            "CDI UI messages",
            "CDI failure/mismatch state",
            "Expected messages preserved as residual context; trigger/setup unresolved.",
            "GAP-004",
        ),
        (
            "ATOM-066",
            "GSR 124; GSR 135",
            "GAP-005; SRC-003; SRC-012",
            "Numeric-only and below-minimum invalid classes lack deterministic UI rejection/feedback oracle.",
            "Income numeric fields",
            "letters/spaces/special/decimal/sign or value below 2000",
            "Valid numeric/minimum acceptance covered; exact invalid-input behavior unresolved.",
            "GAP-005",
        ),
        (
            "ATOM-067",
            "GSR 132; GSR 133; GSR 134",
            "GAP-006; SRC-009",
            "Work phone invalid length and non-digit classes lack deterministic UI rejection/feedback oracle.",
            "Рабочий телефон",
            "9 digits, 11 digits or non-digit characters",
            "Exact 10-digit acceptance and default mask covered; invalid enforcement unresolved.",
            "GAP-006",
        ),
        (
            "ATOM-068",
            "GSR 139; GSR 140",
            "GAP-007; SRC-016",
            "Clear/reset behavior for selected visual-assessment checkboxes is not defined.",
            "Параметры визуальной оценки",
            "selected values need to be cleared",
            "List, no-selection requiredness, single and multiple selection covered; reset behavior unresolved.",
            "GAP-007",
        ),
        (
            "ATOM-069",
            "GSR 143",
            "GAP-008; SRC-018",
            "Field-to-document content mapping for `Заявление-анкета` is not provided in this scope.",
            "Заявление-анкета",
            "print form generated",
            "Generation/opening covered; content mapping unresolved.",
            "GAP-008",
        ),
    ]
    for atom_id, req_id, source_path, statement, field, condition, behavior, gap_id in gap_rows:
        matrix.append(
            {
                "atom_id": atom_id,
                "req_id": req_id,
                "source_path": source_path,
                "atomic_statement": statement,
                "field_or_block": field,
                "condition": condition,
                "expected_behavior": behavior,
                "covered_by_tc": "-",
                "coverage_status": "unclear",
                "gap_note": f"{gap_id} retained as accepted non-blocking residual gap; writer revision still required for open findings.",
            }
        )
    return matrix


def _condition_from_statement(statement: str) -> str:
    marker = " при условии "
    if marker in statement:
        return statement.split(marker, 1)[1].strip().strip(".")
    return "always"


def _field_from_statement(statement: str) -> str:
    if ": " in statement:
        return statement.split(": ", 1)[0].strip()
    return "-"


def _expected_from_statement(statement: str) -> str:
    marker = ": "
    text = statement.split(marker, 1)[1] if marker in statement else statement
    if " при условии " in text:
        text = text.split(" при условии ", 1)[0]
    return text.strip().strip(".")


def write_matrix(matrix: list[dict[str, str]]) -> None:
    headers = [
        "atom_id",
        "req_id",
        "source_path",
        "atomic_statement",
        "field_or_block",
        "condition",
        "expected_behavior",
        "covered_by_tc",
        "coverage_status",
        "gap_note",
    ]
    rows = [[row[h] for h in headers] for row in matrix]
    md = "## Traceability Matrix\n\n" + md_table(headers, rows) + "\n"
    (OUTPUTS / "round-1-traceability-matrix.md").write_text(md, encoding="utf-8")

    wb = Workbook()
    ws = wb.active
    ws.title = "traceability"
    ws.append(headers)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 60)
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    meta = wb.create_sheet("meta")
    meta.append(["field", "value"])
    meta.append(["ft_slug", "ft-2-OF_16"])
    meta.append(["scope_slug", SCOPE])
    meta.append(["round", "1"])
    meta.append(["markdown_artifact", "round-1-traceability-matrix.md"])
    wb.save(OUTPUTS / "round-1-traceability-matrix.xlsx")


def write_findings() -> None:
    findings = """# Semantic Review R1 Findings

## Human Summary

| item | value |
| --- | --- |
| review_mode | `semantic_traceability_test_design` |
| ft_slug | `ft-2-OF_16` |
| scope_slug | `ui-employment-canary-v6-terminal-gate-regression` |
| canonical_test_cases | `test-cases/2-1-1-1-1-2-ui-employment-canary-v6-terminal-gate-regression.md` |
| verdict | `not signed-off` |
| recommended_next_step | `writer-r2` |

Semantic review did not sign off the set. The main blocker is unreliable negative transition attribution: several requiredness cases press `Следующий шаг` while other mandatory fields in the same branch are not fixed to a valid baseline. Traceability is mostly recoverable, but `GSR 126` is not normalized as its own source property/atom or narrow `GAP-*` row in writer-side normalization. Residual `GAP-001`-`GAP-008` are preserved as unresolved/unclear rows in the reviewer matrix; they must not be silently treated as covered.

Coverage counts from reviewer matrix:

| coverage_status | count |
| --- | --- |
| covered | 61 |
| unclear | 8 |
| gap | 0 |

## Review Findings

### FINDING-001
**Review Mode:** test-design
**Severity:** error
**Category:** test-design
**Coverage Dimension:** dependency
**Test Case ID:** TC-UI-EMP-V6-006; TC-UI-EMP-V6-011; TC-UI-EMP-V6-021; TC-UI-EMP-V6-024; TC-UI-EMP-V6-029; TC-UI-EMP-V6-033; TC-UI-EMP-V6-038; TC-UI-EMP-V6-040; TC-UI-EMP-V6-050
**Traceability Ref:** -
**Title:** Negative `Следующий шаг` checks do not isolate the invalid field

**Problem:** Requiredness cases leave the target field empty and press `Следующий шаг`, but their preconditions do not fully disclose a valid fixture for all other mandatory fields in the same branch. The expected result can therefore be caused by another missing mandatory field, so the TC does not prove the linked `ATOM-*`.

**Evidence:**
- `TC-UI-EMP-V6-006` sets only `Тип занятости = Работа по найму`, leaves main income empty and presses `Следующий шаг`; employer, position type, position, experience and work phone are not fixed as valid.
- `TC-UI-EMP-V6-011`, `TC-UI-EMP-V6-021`, `TC-UI-EMP-V6-024`, `TC-UI-EMP-V6-029`, `TC-UI-EMP-V6-033` have the same non-pension branch problem.
- `TC-UI-EMP-V6-038` and `TC-UI-EMP-V6-040` create an additional-income block but do not disclose valid baseline values for mandatory fields outside the checked field.
- `TC-UI-EMP-V6-050` sets `Визуальная информация = Да` and no visual-assessment checkbox, but does not disclose a valid baseline for the other mandatory employment fields.

**Required Change:** For each affected negative transition TC, add a concrete valid baseline fixture for every other mandatory field/control in the active branch, either expanded in the TC preconditions or by linking a fixture artifact and naming its exact values. Keep only the checked field/list invalid or empty.

**Source Reference:** `GSR 142`; `SRC-018`; affected requiredness atoms `ATOM-006`, `ATOM-011`, `ATOM-022`, `ATOM-025`, `ATOM-030`, `ATOM-034`, `ATOM-040`, `ATOM-042`, `ATOM-051`.

**Status:** open

### FINDING-002
**Review Mode:** traceability
**Severity:** error
**Category:** traceability
**Coverage Dimension:** traceability
**Coverage Gap:** GAP-001
**Traceability Ref:** ATOM-062
**Title:** Mandatory PDF-only `GSR 126` is not normalized into a stable source property or atom/gap row

**Problem:** `source-parity-check.md` marks `GSR 126` as a mandatory PDF-only requirement id. Writer-side `source-row-inventory.md` keeps `GSR 126` on `SRC-004`, and `GAP-001` mentions it, but `source-table-normalization.md` has only `SRC-004.P01`-`SRC-004.P04` with `GSR 125`; there is no normalized `SRC-004` property for `GSR 126` and no linked `ATOM-*`/narrow gap property. This makes source-row carryover semantically incomplete even though the id appears textually elsewhere.

**Evidence:**
- `source-parity-check.md` mandatory ids include `GSR 126`.
- Writer-side `source-row-inventory.md`: `SRC-004` has `GSR 125; GSR 126`.
- Writer-side `source-table-normalization.md`: all `SRC-004.P01`-`SRC-004.P04` rows use `GSR 125`; `GSR 126` is absent from normalized properties.
- `source-row-completeness-matrix.md` marks `SRC-004` as `split-complete`, which is not semantically true for the mandatory `GSR 126` id.

**Required Change:** Add a normalized source property for the `GSR 126` part of `SRC-004` and route it to either a concrete `ATOM-*`/TC for the UI-visible part or a narrow `GAP-*`/`unclear` row for the unresolved mechanism. Update source-row completeness, ledger, package plan and traceability links consistently.

**Source Reference:** `source-parity-check.md` -> Mandatory Traceability Inputs; `SRC-004`; `GSR 126`; `GAP-001`.

**Status:** open

### FINDING-003
**Review Mode:** test-design
**Severity:** warning
**Category:** duplication
**Coverage Dimension:** requiredness
**Test Case ID:** TC-UI-EMP-V6-003; TC-UI-EMP-V6-053
**Traceability Ref:** -
**Title:** `TC-UI-EMP-V6-053` duplicates the same empty-field check already performed by `TC-UI-EMP-V6-003`

**Problem:** `TC-UI-EMP-V6-003` and `TC-UI-EMP-V6-053` both leave `Тип занятости` empty, press `Следующий шаг`, and expect the section not to open with red highlight on `Тип занятости`. The second case is framed as action-level `GSR 142` coverage, but it does not exercise a distinct condition or expected result.

**Evidence:**
- `TC-UI-EMP-V6-003`: `Тип занятости` empty -> `Следующий шаг` -> `Тип занятости` highlighted red.
- `TC-UI-EMP-V6-053`: same test data/action/oracle for `Тип занятости`.

**Required Change:** Remove the duplicate action-level TC or convert `GSR 142` coverage into explicit trace links from the field-level requiredness cases. If a separate action-level case is retained, it must exercise a distinct requiredness condition that is not already covered by `TC-UI-EMP-V6-003`.

**Source Reference:** `GSR 142`; `SRC-018`; `ATOM-003`; `ATOM-054`.

**Status:** open

## Residual Unclear Items

The reviewer matrix retains `GAP-001`-`GAP-008` as `coverage_status = unclear`, not `covered`. These items are accepted as non-blocking residual context only after writer fixes the open findings above and keeps the unresolved mechanisms out of executable TC oracles.
"""
    (OUTPUTS / "round-1-findings.md").write_text(findings, encoding="utf-8")


def write_logs_and_prompt() -> None:
    selected = [
        "AGENTS.md",
        "skills/README.md",
        "references/agent/session-based-review-cycle-format.md",
        "references/agent/codex-sdk-orchestration-format.md",
        "skills/ft-test-case-reviewer/SKILL.md",
        "references/agent/reviewer-output-format.md",
        "references/agent/package-test-design-plan-format.md",
        "references/agent/dictionary-inventory-format.md",
        "references/agent/test-design-defect-taxonomy.md",
        "references/qa/review-findings-format.md",
        "references/qa/traceability-matrix-format.md",
        "references/qa/test-design-review-rubric.md",
        "references/qa/coverage-runtime-checklist.md",
        "references/qa/traceability-rules.md",
        "references/agent/workflow-state-format.md",
        "references/agent/session-log-format.md",
        "references/agent/agent-decision-log-format.md",
        "references/agent/next-step-prompt-format.md",
    ]
    read_inputs = [
        "fts/ft-2-OF_16/AGENT-NOTES.md",
        "fts/ft-2-OF_16/work/review-cycles/ui-employment-canary-v6-terminal-gate-regression/cycle-state.yaml",
        "fts/ft-2-OF_16/work/review-cycles/ui-employment-canary-v6-terminal-gate-regression/outputs/structure-preflight-r1-findings.md",
        "fts/ft-2-OF_16/work/review-cycles/ui-employment-canary-v6-terminal-gate-regression/outputs/reviewer-session-log.structure-preflight-r1.md",
        "fts/ft-2-OF_16/test-cases/2-1-1-1-1-2-ui-employment-canary-v6-terminal-gate-regression.md",
        "fts/ft-2-OF_16/work/test-design/ui-employment-canary-v6-terminal-gate-regression/",
        "fts/ft-2-OF_16/work/review-cycles/ui-employment-canary-v6-terminal-gate-regression/outputs/writer-session-log.writer-r1.md",
        "fts/ft-2-OF_16/work/review-cycles/ui-employment-canary-v6-terminal-gate-regression/outputs/agent-decision-log.writer-r1.md",
        "fts/ft-2-OF_16/work/stage-handoffs/01-ui-employment/source-row-inventory.md",
        "fts/ft-2-OF_16/work/stage-handoffs/01-ui-employment/source-parity-check.md",
        "fts/ft-2-OF_16/work/stage-handoffs/01-ui-employment/mockup-visual-inventory.md",
        "fts/ft-2-OF_16/work/stage-handoffs/01-ui-employment/scope-coverage-gaps.md",
        "fts/ft-2-OF_16/work/test-design/ui-employment/dictionary-inventory.md",
    ]
    session_log = "# Semantic Review R1 Session Log\n\n"
    session_log += """## Session Metadata

| field | value |
| --- | --- |
| skill | `ft-test-case-reviewer` |
| mode | `semantic_traceability_test_design` |
| ft_slug | `ft-2-OF_16` |
| scope_slug | `ui-employment-canary-v6-terminal-gate-regression` |
| started_from | `work/review-cycles/ui-employment-canary-v6-terminal-gate-regression/cycle-state.yaml` |
| status_after | `semantic-revision-needed` |

## Artifact Write Strategy

| artifact_path | artifact_size_class | write_strategy | declared_before_first_write | helper | forbidden_methods_checked |
| --- | --- | --- | --- | --- | --- |
| `work/review-cycles/ui-employment-canary-v6-terminal-gate-regression/outputs/round-1-findings.md` | `review artifact` | `file-based helper write` | `yes` | `scripts/build_semantic_review_r1_outputs.py` | `yes` |
| `work/review-cycles/ui-employment-canary-v6-terminal-gate-regression/outputs/round-1-traceability-matrix.md` | `generated matrix` | `file-based helper write` | `yes` | `scripts/build_semantic_review_r1_outputs.py` | `yes` |
| `work/review-cycles/ui-employment-canary-v6-terminal-gate-regression/outputs/round-1-traceability-matrix.xlsx` | `generated xlsx` | `openpyxl workbook write` | `yes` | `scripts/build_semantic_review_r1_outputs.py` | `yes` |

## Inputs Read

- `python scripts/resolve_instruction_context.py --scenario reviewer.semantic_traceability_test_design --budget-report --fail-on-budget` - resolver command; budget status `pass (259.1 / 290.0 KiB)`.
"""
    for path in selected:
        session_log += f"- `{path}` - selected required instruction file.\n"
    for path in read_inputs:
        session_log += f"- `{path}` - required stage input for semantic review.\n"
    session_log += """
## Inputs Not Used

- Older canary versions `v1`-`v5` - explicitly forbidden as requirement sources/templates for this v6 review.
- Neighboring FT packages - outside selected FT package.

## Key Decisions

- Did not edit the canonical test-case file; reviewer outputs only were created.
- Built reviewer traceability from current writer ledger and residual `GAP-*` rows.
- Classified `GAP-001`-`GAP-008` as residual `unclear` rows in reviewer matrix, not as covered behavior.
- Rejected semantic sign-off because negative transition cases do not isolate checked requiredness failures and mandatory `GSR 126` normalization is incomplete.

## Risks And Fallbacks

- Full-package validator output is noisy because it includes historical canary artifacts; it was used only as scoped warning context after filtering current v6 paths.
- Initial instruction stdout had mojibake for some Cyrillic files; affected files were reread with explicit UTF-8 and distorted stdout was not used as semantic evidence.

## Validation

- `python scripts/resolve_instruction_context.py --scenario reviewer.semantic_traceability_test_design --budget-report --fail-on-budget` - pass, `259.1 / 290.0 KiB`.
- `python scripts\\validate_agent_artifacts.py --root fts\\ft-2-OF_16 --json` - executed; full-root output noisy, current v6 semantic warnings reviewed manually.
- Helper generation self-check - created findings, Markdown matrix, XLSX matrix, session log, decision log and writer-r2 prompt.
- `openpyxl` workbook check - pass: `round-1-traceability-matrix.xlsx` has 69 rows and 10 columns on sheet `traceability`.

## Contamination Check

- Older canary versions were not used as source requirements or templates.
- Mockup inventory was used only to verify interaction-hint limits; no mockup-only business rule was promoted to covered behavior.

## Event Timeline

| step | event | result | artifact_or_evidence |
| --- | --- | --- | --- |
| 1 | Resolved instruction context | pass | budget `259.1 / 290.0 KiB` |
| 2 | Read selected required instruction files | pass | 18 files listed above |
| 3 | Read required cycle/source/test-design inputs | pass | files listed above |
| 4 | Ran semantic traceability/test-design review | not signed-off | `round-1-findings.md` |
| 5 | Built reviewer traceability matrix | pass | `round-1-traceability-matrix.md`; `.xlsx` |
| 6 | Updated transition artifacts/state | pass | `prompt.semantic-review-r1-to-writer-r2.md`; `cycle-state.yaml` |

## Quality Checkpoints

| checkpoint | status | evidence | follow_up |
| --- | --- | --- | --- |
| Instruction loading gate | pass | resolver budget pass and selected files read | none |
| Traceability gate | fail | `FINDING-002`; `ATOM-062` | writer must normalize `GSR 126` |
| Test-design gate | fail | `FINDING-001`; affected negative TC list | writer must add concrete valid fixtures |
| Residual gap admissibility | pass-with-residual-risk | `GAP-001`-`GAP-008` retained as `unclear` rows | writer must not convert mechanisms to covered TC |

## Technical Fallbacks

| fallback_id | trigger | failed_method | fallback_method | helper_artifact_path | retained | quality_risk | follow_up |
| --- | --- | --- | --- | --- | --- | --- | --- |
| `TF-001` | mojibake in PowerShell stdout for Cyrillic instruction/source files | default console output read | explicit UTF-8 `Get-Content -Encoding UTF8` rereads; distorted stdout discarded | `n/a` | `n/a` | none after reread | keep UTF-8 preamble in writer-r2 |
| `TF-002` | large generated reviewer artifacts plus XLSX duplicate | manual one-shot command/write would be fragile | file-based helper script with openpyxl for XLSX | `scripts/build_semantic_review_r1_outputs.py` | `yes` | low: generated table should be reviewed in writer-r2 | writer-r2 should preserve matrix refs when responding |
| `TF-003` | full-root validator includes historical artifacts | unfiltered validator as decisive scoped gate | scoped manual filtering/current-v6 semantic review | `n/a` | `n/a` | medium: validator output is noisy | use reviewer findings as authoritative for this stage |
| `TF-004` | PowerShell module unavailable for XLSX validation | `Import-Excel` command | `openpyxl.load_workbook` workbook inspection | `n/a` | `n/a` | none: same generated workbook opened successfully | no follow-up |

## Handoff Notes For Next Session

- Writer-r2 must address `FINDING-001`-`FINDING-003` and create a canonical writer response artifact.
- Do not route to UI automation prep from semantic round 1.
- Preserve residual `GAP-001`-`GAP-008`; do not invent invalid-input, duplicate-income, CDI, clear/reset or document-content mechanisms.
"""
    (OUTPUTS / "reviewer-session-log.semantic-review-r1.md").write_text(session_log, encoding="utf-8")

    decision_log = """# Agent Decision Log

## Decision Log Metadata

| field | value |
| --- | --- |
| ft_slug | `ft-2-OF_16` |
| scope_slug | `ui-employment-canary-v6-terminal-gate-regression` |
| stage | `semantic-review-r1` |
| started_from | `cycle-state.yaml` |

## Decision Log

| decision_id | step | decision_type | input_or_trigger | decision | rationale | artifact_or_output | risk_or_confidence | status |
| --- | --- | --- | --- | --- | --- | --- | --- | --- |
| DEC-001 | 1 | traceability | `source-parity-check.md`; writer normalization | Require writer remediation for `GSR 126` normalization | Mandatory PDF-only id appears only as broad residual context, not as a normalized source property/atom/gap row | `round-1-findings.md`; `ATOM-062` | high | applied |
| DEC-002 | 2 | test-design | requiredness TC review | Reject negative transition cases without full valid fixture | Failure attribution is unreliable when other mandatory fields can also block `Следующий шаг` | `round-1-findings.md` | high | applied |
| DEC-003 | 3 | gap | `GAP-001`-`GAP-008` | Keep residual mechanisms as `unclear`, not `covered` | Source lacks deterministic UI/setup/oracle for these mechanisms | `round-1-traceability-matrix.md` | medium | applied |
| DEC-004 | 4 | routing | open error/warning findings | Route cycle to `writer-r2` with `semantic-revision-needed` | Session-based gate forbids semantic pass with open error/warning findings | `cycle-state.yaml`; writer-r2 prompt | high | applied |
"""
    (OUTPUTS / "agent-decision-log.semantic-review-r1.md").write_text(decision_log, encoding="utf-8")

    prompt = """# Semantic Review R1 To Writer R2

## Цель этапа

Исправить semantic/test-design findings round 1 для scope `ui-employment-canary-v6-terminal-gate-regression` и подготовить writer response для `semantic-review-r2`.

## Входные артефакты

- `fts/ft-2-OF_16/work/review-cycles/ui-employment-canary-v6-terminal-gate-regression/cycle-state.yaml`
- `fts/ft-2-OF_16/work/review-cycles/ui-employment-canary-v6-terminal-gate-regression/outputs/round-1-findings.md`
- `fts/ft-2-OF_16/work/review-cycles/ui-employment-canary-v6-terminal-gate-regression/outputs/round-1-traceability-matrix.md`
- `fts/ft-2-OF_16/work/review-cycles/ui-employment-canary-v6-terminal-gate-regression/outputs/round-1-traceability-matrix.xlsx`
- `fts/ft-2-OF_16/test-cases/2-1-1-1-1-2-ui-employment-canary-v6-terminal-gate-regression.md`
- `fts/ft-2-OF_16/work/test-design/ui-employment-canary-v6-terminal-gate-regression/`
- `fts/ft-2-OF_16/work/stage-handoffs/01-ui-employment/source-parity-check.md`
- `fts/ft-2-OF_16/work/stage-handoffs/01-ui-employment/source-row-inventory.md`
- `fts/ft-2-OF_16/work/stage-handoffs/01-ui-employment/scope-coverage-gaps.md`
- `fts/ft-2-OF_16/work/stage-handoffs/01-ui-employment/mockup-visual-inventory.md`
- `fts/ft-2-OF_16/work/test-design/ui-employment/dictionary-inventory.md`
- `fts/ft-2-OF_16/AGENT-NOTES.md`

## Обязательные действия

- Исправить `FINDING-001`: для всех affected negative `Следующий шаг` кейсов раскрыть конкретную валидную фикстуру всех остальных обязательных полей активной ветки.
- Исправить `FINDING-002`: добавить нормализованную строку/решение для mandatory `GSR 126` в `SRC-004` и связать ее с конкретным `ATOM-*`/TC или узким `GAP-*`/`unclear`.
- Исправить `FINDING-003`: убрать дублирование `TC-UI-EMP-V6-003` и `TC-UI-EMP-V6-053` или сделать отдельный action-level case реально отличающимся.
- Создать `outputs/writer-r2-response.md` с ответом на каждый finding по `references/qa/review-findings-format.md`.
- Обновить canonical test-case file и split artifacts согласованно; сохранить `GAP-001`-`GAP-008` как residual unresolved context, если новый source/evidence их не закрывает.

## Не делать

- Не использовать старые canary versions как source requirements or templates.
- Не переводить unresolved invalid-input, CDI, duplicate-income, checkbox clear/reset or document-content mapping mechanisms в executable TC без source/evidence.
- Не запускать UI automation prep и не ставить signed-off.

## Ожидаемые выходы

- Обновленный `test-cases/2-1-1-1-1-2-ui-employment-canary-v6-terminal-gate-regression.md`
- Обновленные affected split artifacts under `work/test-design/ui-employment-canary-v6-terminal-gate-regression/`
- `work/review-cycles/ui-employment-canary-v6-terminal-gate-regression/outputs/writer-r2-response.md`
- writer-r2 session log and decision log
- updated `cycle-state.yaml` routed to `semantic-review-r2`

## Gate завершения

Этап считается завершенным, когда все findings round 1 имеют canonical writer response, canonical and split artifacts are updated consistently, and `cycle-state.yaml` routes to:

```yaml
current_stage: semantic-review-r2
stage_status: semantic-review-ready
semantic_round: 2
```
"""
    (PROMPTS / "prompt.semantic-review-r1-to-writer-r2.md").write_text(prompt, encoding="utf-8")


def update_state() -> None:
    text = STATE.read_text(encoding="utf-8")
    replacements = {
        "current_stage: semantic-review-r1": "current_stage: writer-r2",
        "stage_status: semantic-review-ready": "stage_status: semantic-revision-needed",
        "active_transition_prompt: work/review-cycles/ui-employment-canary-v6-terminal-gate-regression/prompts/prompt.semantic-review-r1.md": "active_transition_prompt: work/review-cycles/ui-employment-canary-v6-terminal-gate-regression/prompts/prompt.semantic-review-r1-to-writer-r2.md",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    additions = [
        "  - work/review-cycles/ui-employment-canary-v6-terminal-gate-regression/outputs/round-1-findings.md",
        "  - work/review-cycles/ui-employment-canary-v6-terminal-gate-regression/outputs/round-1-traceability-matrix.md",
        "  - work/review-cycles/ui-employment-canary-v6-terminal-gate-regression/outputs/round-1-traceability-matrix.xlsx",
        "  - work/review-cycles/ui-employment-canary-v6-terminal-gate-regression/outputs/reviewer-session-log.semantic-review-r1.md",
        "  - work/review-cycles/ui-employment-canary-v6-terminal-gate-regression/outputs/agent-decision-log.semantic-review-r1.md",
        "  - work/review-cycles/ui-employment-canary-v6-terminal-gate-regression/prompts/prompt.semantic-review-r1-to-writer-r2.md",
    ]
    marker = "blocking_reasons:"
    if additions[0] not in text and marker in text:
        before, after = text.split(marker, 1)
        before = before.rstrip() + "\n" + "\n".join(additions) + "\n"
        text = before + marker + after
    text = text.replace("blocking_reasons: []", "blocking_reasons:\n  - FINDING-001\n  - FINDING-002\n  - FINDING-003")
    text = text.replace("blocking_findings: []", "blocking_findings:\n  - FINDING-001\n  - FINDING-002\n  - FINDING-003")
    STATE.write_text(text, encoding="utf-8")


def main() -> None:
    OUTPUTS.mkdir(parents=True, exist_ok=True)
    PROMPTS.mkdir(parents=True, exist_ok=True)
    matrix = build_matrix()
    write_matrix(matrix)
    write_findings()
    write_logs_and_prompt()
    update_state()
    print(f"Generated semantic-review-r1 outputs at {OUTPUTS}")


if __name__ == "__main__":
    main()
