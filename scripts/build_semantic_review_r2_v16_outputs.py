from __future__ import annotations

import html
import re
import zipfile
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
FT_ROOT = ROOT / "fts" / "ft-2-OF_16"
SCOPE = "ui-employment-canary-v16-clean-runner-regression"
CYCLE = FT_ROOT / "work" / "review-cycles" / SCOPE
OUTPUTS = CYCLE / "outputs"
TEST_DESIGN = FT_ROOT / "work" / "test-design" / SCOPE
CANONICAL = FT_ROOT / "test-cases" / "2-1-1-1-1-2-ui-employment-canary-v16-clean-runner-regression.md"
WRITER_MATRIX = OUTPUTS / "writer-r2-traceability-matrix.md"
ROUND2_MATRIX = OUTPUTS / "round-2-traceability-matrix.md"
ROUND2_XLSX = OUTPUTS / "round-2-traceability-matrix.xlsx"


def parse_markdown_table(text: str) -> tuple[list[str], list[dict[str, str]]]:
    lines = [line.strip() for line in text.splitlines() if line.strip().startswith("|")]
    header = [cell.strip() for cell in lines[0].strip("|").split("|")]
    rows: list[dict[str, str]] = []
    for line in lines[2:]:
        cells = [cell.strip() for cell in line.strip("|").split("|")]
        rows.append(dict(zip(header, cells)))
    return header, rows


def render_table(header: list[str], rows: list[dict[str, str]]) -> str:
    out = ["| " + " | ".join(header) + " |", "| " + " | ".join(["---"] * len(header)) + " |"]
    for row in rows:
        out.append("| " + " | ".join(row.get(col, "") for col in header) + " |")
    return "\n".join(out)


def build_matrix() -> tuple[list[str], list[dict[str, str]]]:
    header, rows = parse_markdown_table(WRITER_MATRIX.read_text(encoding="utf-8"))
    blockers = {
        "ATOM-008": "Current `TC-EMP-V16-006` uses unsupported red-highlight / blocked-navigation oracle. Source and plan require non-acceptance of `1999`, not a specific UI feedback mechanism.",
        "ATOM-009": "Current numeric-class TCs are split, but each uses unsupported red-highlight / blocked-navigation oracle instead of proving the invalid value is not accepted/displayed.",
        "ATOM-015": "Current additional-income numeric-class TCs are split, but each uses unsupported red-highlight / blocked-navigation oracle instead of proving the invalid value is not accepted/displayed.",
    }
    for row in rows:
        atom = row.get("atom_id", "")
        if atom in blockers:
            row["covered_by_tc"] = "-"
            row["coverage_status"] = "gap"
            row["gap_note"] = blockers[atom]
    ROUND2_MATRIX.write_text("# Round 2 Traceability Matrix\n\n" + render_table(header, rows) + "\n", encoding="utf-8")
    return header, rows


def col_letter(index: int) -> str:
    letters = ""
    while index:
        index, rem = divmod(index - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def write_xlsx_stdlib(header: list[str], rows: list[dict[str, str]]) -> None:
    table = [header] + [[row.get(col, "") for col in header] for row in rows]
    sheet_rows = []
    for r_idx, row in enumerate(table, start=1):
        cells = []
        for c_idx, value in enumerate(row, start=1):
            ref = f"{col_letter(c_idx)}{r_idx}"
            text = html.escape(str(value))
            cells.append(f'<c r="{ref}" t="inlineStr"><is><t>{text}</t></is></c>')
        sheet_rows.append(f'<row r="{r_idx}">{"".join(cells)}</row>')
    dimension = f"A1:{col_letter(len(header))}{len(table)}"
    sheet_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        f'<dimension ref="{dimension}"/><sheetViews><sheetView workbookViewId="0"><pane ySplit="1" topLeftCell="A2" activePane="bottomLeft" state="frozen"/></sheetView></sheetViews>'
        f'<sheetData>{"".join(sheet_rows)}</sheetData><autoFilter ref="{dimension}"/></worksheet>'
    )
    workbook_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<sheets><sheet name="traceability" sheetId="1" r:id="rId1"/></sheets></workbook>'
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
        '</Relationships>'
    )
    wb_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
        '</Relationships>'
    )
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        '</Types>'
    )
    with zipfile.ZipFile(ROUND2_XLSX, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", content_types)
        zf.writestr("_rels/.rels", rels)
        zf.writestr("xl/workbook.xml", workbook_xml)
        zf.writestr("xl/_rels/workbook.xml.rels", wb_rels)
        zf.writestr("xl/worksheets/sheet1.xml", sheet_xml)


def write_xlsx(header: list[str], rows: list[dict[str, str]]) -> str:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except Exception:
        write_xlsx_stdlib(header, rows)
        return "stdlib-zipfile"

    wb = Workbook()
    ws = wb.active
    ws.title = "traceability"
    ws.append(header)
    for row in rows:
        ws.append([row.get(col, "") for col in header])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for column in ws.columns:
        width = min(70, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
        ws.column_dimensions[column[0].column_letter].width = width
        for cell in column:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    meta = wb.create_sheet("meta")
    meta.append(["field", "value"])
    meta.append(["ft_slug", "ft-2-OF_16"])
    meta.append(["scope_slug", SCOPE])
    meta.append(["round", "2"])
    meta.append(["markdown_artifact", "round-2-traceability-matrix.md"])
    wb.save(ROUND2_XLSX)
    return "openpyxl"


def write_findings() -> None:
    content = f"""# Semantic Review R2 Findings

## Human Summary

**review_mode:** semantic_traceability_test_design  
**verdict:** round-cap-reached  
**ft_slug:** `ft-2-OF_16`  
**scope_slug:** `{SCOPE}`  
**canonical_test_cases:** `test-cases/{CANONICAL.name}`  
**pdf_structural_cross_check:** yes, via `work/stage-handoffs/01-ui-employment/source-parity-check.md`  
**source_parity_checked:** yes  
**mockup_visual_inventory_checked:** yes, `opened = yes`, `not_used_as_requirement_source = yes`  
**writer_response_validated:** yes, all `FINDING-001` ... `FINDING-007` blocks use canonical writer response fields and allowed `resolution_status` values.

Reviewer checked the current canonical TC file, writer-r2 response, writer-r2 traceability matrix, scoped validator profile, source parity, mockup inventory and split test-design artifacts. Six round-1 findings are closed. `FINDING-003` is reopened in a narrower form as `FINDING-008`: unsupported clearing was removed, but numeric rejection cases now assert a different unsupported mechanism (`red highlight` plus blocked transition) and disagree with the source/plan oracle.

Coverage summary by round-2 reviewer matrix:

| status | count |
| --- | ---: |
| covered | 23 |
| gap | 6 |
| unclear | 0 |

Preserved residual source/setup gaps: `GAP-002`, `GAP-003`, `GAP-004`, `GAP-005`. No new source or evidence resolves them.

## Round-1 Finding Closure

| finding_id | writer_response_status | reviewer_r2_decision | evidence |
| --- | --- | --- | --- |
| `FINDING-001` | `fixed` | closed | `ATOM-018` now maps to dedicated `TC-EMP-V16-018` in canonical TC, ledger, TDDT, plan and writer-r2 matrix. |
| `FINDING-002` | `fixed` | closed | `ATOM-019` now maps to `TC-EMP-V16-027`; the case checks no-blocking branch without inventing an empty toggle-state. |
| `FINDING-003` | `fixed` | reopened as `FINDING-008` | Clearing oracle was removed, but canonical numeric TCs still assert unsupported red-highlight/blocked-navigation and disagree with plan/source non-acceptance oracle. |
| `FINDING-004` | `fixed` | closed | Numeric invalid classes are split into separate TC IDs for main and additional income. |
| `FINDING-005` | `fixed` | closed | `GSR 143 / ATOM-027` is consistently `covered` by `TC-EMP-V16-017` across source-row completeness, ledger, TDDT, plan, canonical TC and matrix. |
| `FINDING-006` | `fixed` | closed | `Package Test Design Plan.check_type` uses only canonical enum values. |
| `FINDING-007` | `fixed` | closed | The explicit `Test-design Applicability Model` substitute is accepted for this scoped run; split applicability artifacts are listed and validator has only nonblocking `info`. |

## Review Findings

### FINDING-008
**Review Mode:** test-design  
**Severity:** error  
**Category:** expected-result  
**Coverage Dimension:** numeric  
**Test Case ID:** TC-EMP-V16-006; TC-EMP-V16-007; TC-EMP-V16-011; TC-EMP-V16-019; TC-EMP-V16-020; TC-EMP-V16-021; TC-EMP-V16-022; TC-EMP-V16-023; TC-EMP-V16-024; TC-EMP-V16-025; TC-EMP-V16-026  
**Traceability Ref:** ATOM-008; ATOM-009; ATOM-015  
**Title:** Numeric rejection cases still use unsupported UI feedback and drift from the plan/source oracle

**Problem:** Writer-r2 removed the unsupported clearing oracle, but replaced it with another unsupported mechanism: each numeric rejection TC expects red highlighting and `Анкета клиента` not opening. Source normalization for `SRC-003.P006`, `SRC-003.P007` and `SRC-012.P003` only says the invalid value is not accepted; red highlighting is source-backed for empty required fields through `SRC-018.P001 / GSR 142`, not for invalid numeric values. The split `Package Test Design Plan` also expects `field displays a value different from <invalid>`, while the canonical cases assert red/blocked navigation instead.

**Evidence:**
- `source-table-normalization.md`: `SRC-003.P006` says `Значение ниже 2000 не принимается`; `SRC-003.P007` says `Поле не принимает нечисловые символы`; `SRC-012.P003` says `Поле не принимает нечисловые символы`.
- `source-table-normalization.md`: red highlighting is tied to `SRC-018.P001 / GSR 142` for `Есть незаполненное обязательное поле`, not to invalid numeric classes.
- `package-test-design-plan.md`: `PD-006`, `PD-007`, `PD-011`, `PD-021`-`PD-028` use the oracle `Value <invalid> is not accepted; field displays a value different from <invalid>`.
- Canonical `TC-EMP-V16-006`, `TC-EMP-V16-007`, `TC-EMP-V16-011`, `TC-EMP-V16-019`-`TC-EMP-V16-026` all expect the field to be red-highlighted and `Анкета клиента` not opened.
- `coverage-obligation-table.md` still contains duplicate validator-completion rows (`OBL-014`-`OBL-018`, `OBL-020`-`OBL-024`) that map several invalid classes back to `TC-EMP-V16-007` / `TC-EMP-V16-011`, even though those cases now cover only `letters`.

**Required Change:** Replace numeric invalid expected results with one deterministic source-backed oracle that directly proves non-acceptance, for example the invalid literal is not displayed/retained after focus loss, or provide approved source/UI evidence for red highlighting/blocked navigation on invalid numeric values. Align canonical TC, `Package Test Design Plan`, `Coverage Obligation Table`, TDDT, ledger and traceability matrix so each invalid class points to its actual atomic TC and no duplicate obligation row overstates coverage.

**Source Reference:** `source-table-normalization.md` rows `SRC-003.P006`, `SRC-003.P007`, `SRC-012.P003`, `SRC-018.P001`; `package-test-design-plan.md` rows `PD-006`, `PD-007`, `PD-011`, `PD-021`-`PD-028`; `references/agent/test-design-defect-taxonomy.md` defect class `unsupported-ui-mechanism`

**Status:** open

## Final Residual Risk

**remaining_blocking_findings:** FINDING-008  
**remaining_traceability_gaps:** ATOM-008; ATOM-009; ATOM-015; ATOM-012; ATOM-028; ATOM-029  
**remaining_coverage_gaps:** GAP-002; GAP-003; GAP-004; GAP-005  
**remaining_unclear_items:** none  
**decision_rationale:** Semantic round 2 is the final allowed semantic round. Because a blocking expected-result/test-design finding remains open, the cycle must stop at `round-cap-reached` rather than move to format review or UI automation prep.  
**next_action:** A future run must revise the numeric rejection oracle and cross-artifact mappings, then start a new review cycle or explicitly reset the semantic loop under a new cycle id.

## Gate Decision Result

Review-cycle stopped at `round-cap-reached`: unresolved semantic work remains and the workflow must not route to UI automation prep.
"""
    (OUTPUTS / "round-2-findings.md").write_text(content, encoding="utf-8")


def write_session_log(xlsx_method: str) -> None:
    files = [
        "AGENTS.md",
        "skills/README.md",
        "references/agent/session-based-review-cycle-format.md",
        "references/agent/codex-sdk-orchestration-format.md",
        "skills/ft-test-case-reviewer/SKILL.md",
        "references/agent/reviewer-output-format.md",
        "references/agent/package-test-design-plan-format.md",
        "references/agent/dictionary-inventory-format.md",
        "references/agent/test-design-defect-taxonomy.md",
        "references/qa/review-findings-format.md",
        "references/qa/traceability-matrix-format.md",
        "references/qa/test-design-review-rubric.md",
        "references/qa/coverage-runtime-checklist.md",
        "references/qa/traceability-rules.md",
        "references/agent/workflow-state-format.md",
        "references/agent/session-log-format.md",
        "references/agent/agent-decision-log-format.md",
        "references/agent/next-step-prompt-format.md",
    ]
    input_lines = "\n".join(f"- `{path}` - selected required instruction file read before reviewer decisions." for path in files)
    content = f"""# Reviewer Semantic Review R2 Session Log

## Session Metadata

| field | value |
| --- | --- |
| skill | `ft-test-case-reviewer` |
| mode | `semantic_traceability_test_design` |
| ft_slug | `ft-2-OF_16` |
| scope_slug | `{SCOPE}` |
| started_from | `cycle-state.yaml` |
| status_after | `round-cap-reached` |

## Inputs Read

- `python scripts/resolve_instruction_context.py --scenario reviewer.semantic_traceability_test_design --budget-report --fail-on-budget` - resolver command executed; budget status `pass (261.7 / 290.0 KiB)`; selected files below.
{input_lines}
- `fts/ft-2-OF_16/work/review-cycles/{SCOPE}/cycle-state.yaml` - session source of truth.
- `fts/ft-2-OF_16/test-cases/{CANONICAL.name}` - canonical TC under review.
- `fts/ft-2-OF_16/work/test-design/{SCOPE}/` - split test-design artifacts checked for affected atoms and numeric classes.
- `fts/ft-2-OF_16/work/review-cycles/{SCOPE}/outputs/round-1-findings.md` - round-1 findings to close/reopen.
- `fts/ft-2-OF_16/work/review-cycles/{SCOPE}/outputs/round-1-traceability-matrix.md` - previous reviewer matrix.
- `fts/ft-2-OF_16/work/review-cycles/{SCOPE}/outputs/writer-r2-response.md` - writer response validation.
- `fts/ft-2-OF_16/work/review-cycles/{SCOPE}/outputs/writer-r2-traceability-matrix.md` - writer-r2 traceability baseline.
- `fts/ft-2-OF_16/work/review-cycles/{SCOPE}/outputs/scoped-validator-profile.writer-r2.json` - scoped validator blocker count.
- `fts/ft-2-OF_16/work/review-cycles/{SCOPE}/outputs/validator-report.writer-r2.full.json` - root validator report; used only as background because it contains historical noise outside active scope.
- `fts/ft-2-OF_16/work/review-cycles/{SCOPE}/outputs/writer-session-log.writer-r2.md` - writer technical fallbacks and reviewer focus.
- `fts/ft-2-OF_16/work/review-cycles/{SCOPE}/outputs/agent-decision-log.writer-r2.md` - writer decisions.
- `fts/ft-2-OF_16/work/stage-handoffs/01-ui-employment/source-parity-check.md` - PDF/DOCX parity and mandatory GSR ids.
- `fts/ft-2-OF_16/work/stage-handoffs/01-ui-employment/mockup-visual-inventory.md` - UI hints only, not business rules.
- `fts/ft-2-OF_16/AGENT-NOTES.md` - package notes.

## Inputs Not Used

- `codex-session-map.yaml` - SDK runner owned; not edited.
- UI automation artifacts - explicitly out of scope for semantic review r2.
- Historical canary versions outside `{SCOPE}` - excluded from source of truth; root validator historical findings were not used as current-scope evidence.

## Key Decisions

- Writer response format is canonical for all seven round-1 findings.
- Closed `FINDING-001`, `FINDING-002`, `FINDING-004`, `FINDING-005`, `FINDING-006`, `FINDING-007`.
- Reopened `FINDING-003` as focused `FINDING-008` because numeric invalid TC still assert unsupported red-highlight/blocked-navigation and drift from source/plan non-acceptance oracle.
- Accepted the explicit `Test-design Applicability Model` substitute for this scoped runner regression; no duplicate inline matrix required in R2.
- Preserved `GAP-002`, `GAP-003`, `GAP-004`, `GAP-005`.
- Routed terminally to `round-cap-reached` because semantic round 2 is the max semantic round and a blocking finding remains.

## Risks And Fallbacks

- Root validator report contains historical unrelated findings; scoped profile is the active current-scope validator evidence.
- Numeric expected-result issue is semantic and not fully caught by the scoped validator profile.
- Initial PowerShell reads produced mojibake; affected instruction/source artifacts were reread using explicit UTF-8 and distorted stdout was not used as evidence.

## Validation

- Instruction resolver - pass, `261.7 / 290.0 KiB`.
- Writer response canonical fields/enums - pass.
- `round-2-traceability-matrix.xlsx` generation - pass via `{xlsx_method}`.
- Cycle state updated to terminal `round-cap-reached`; active prompt cleared to `none`.

## Contamination Check

- Review stayed within `ft-2-OF_16` and scope `{SCOPE}`.
- Mockup inventory used only for UI aliases/hints; no mockup-only validation or expected result promoted to requirement.
- Historical canary runs and root validator historical findings were not used to open current semantic findings.

## Event Timeline

| step | event | result | artifact_or_evidence |
| --- | --- | --- | --- |
| 1 | Routed to reviewer semantic r2 | `ft-test-case-reviewer` selected | active prompt |
| 2 | Resolved instruction context | pass | resolver output |
| 3 | Read required instructions and inputs | pass | selected files and cycle artifacts |
| 4 | Validated writer response | pass | `writer-r2-response.md` |
| 5 | Checked affected atoms and numeric TC | one blocker found | `FINDING-008` |
| 6 | Built round-2 findings and matrix | terminal round-cap evidence | `round-2-findings.md`; `round-2-traceability-matrix.md`; `.xlsx` |
| 7 | Updated cycle state | terminal | `cycle-state.yaml` |

## Quality Checkpoints

| checkpoint | status | evidence | follow_up |
| --- | --- | --- | --- |
| Round-1 closure | fail | `FINDING-003` reopened as `FINDING-008` | future revision must fix numeric oracle |
| Traceability refs | fail | `ATOM-008`, `ATOM-009`, `ATOM-015` marked `gap` in round-2 matrix | align TC/plan/source oracle |
| Applicability model | pass | explicit split substitute and scoped validator `info` only | none |
| GSR 143 consistency | pass | `SRC-018.P002`, `ATOM-027`, `TC-EMP-V16-017` consistent | none |

## Technical Fallbacks

| fallback_id | trigger | failed_method | fallback_method | helper_artifact_path | retained | quality_risk | follow_up |
| --- | --- | --- | --- | --- | --- | --- | --- |
| `TF-001` | PowerShell stdout mojibake on Russian instruction/source reads | plain `Get-Content -Raw` console output | explicit UTF-8 reread with `Get-Content -Encoding UTF8`; distorted stdout discarded | n/a | n/a | none after reread | none |
| `TF-002` | generated reviewer matrix plus required XLSX duplicate | manual spreadsheet edit would be fragile | file-based helper script with `{xlsx_method}` | `scripts/build_semantic_review_r2_v16_outputs.py` | yes | low | none |

## Handoff Notes For Next Session

- Cycle is terminal `round-cap-reached`; do not route to UI automation prep.
- A future writer must fix numeric invalid expected results and remove duplicated/incorrect numeric obligation mappings before another semantic sign-off attempt.
"""
    (OUTPUTS / "reviewer-session-log.semantic-review-r2.md").write_text(content, encoding="utf-8")


def write_decision_log() -> None:
    content = f"""# Agent Decision Log

## Decision Log Metadata

| field | value |
| --- | --- |
| ft_slug | `ft-2-OF_16` |
| scope_slug | `{SCOPE}` |
| stage | `semantic-review-r2` |
| started_from | `cycle-state.yaml` |

## Decision Log

| decision_id | step | decision_type | input_or_trigger | decision | rationale | artifact_or_output | risk_or_confidence | status |
| --- | --- | --- | --- | --- | --- | --- | --- | --- |
| `DEC-001` | 1 | `routing` | active prompt | Use `ft-test-case-reviewer` semantic r2 mode. | Stage is reviewer semantic round 2. | session log | high | applied |
| `DEC-002` | 2 | `validation` | resolver output | Allow domain review to proceed. | Resolver succeeded and budget status was pass. | session log | high | applied |
| `DEC-003` | 3 | `structure` | writer response | Accept writer response structure. | All seven response blocks use canonical fields and enum values. | `round-2-findings.md` | high | applied |
| `DEC-004` | 4 | `traceability` | `FINDING-001`; `FINDING-002`; `FINDING-005` | Close affected traceability findings. | `ATOM-018`, `ATOM-019`, `ATOM-027` have consistent current mappings. | `round-2-findings.md`; matrix | high | applied |
| `DEC-005` | 5 | `test-design` | `FINDING-003`; numeric TC | Reopen numeric expected-result defect as `FINDING-008`. | Source/plan prove non-acceptance, not red-highlight/blocked-navigation for invalid numeric values. | `round-2-findings.md`; matrix | high | applied |
| `DEC-006` | 6 | `test-design` | `FINDING-004` | Close class-splitting finding but include obligation drift as evidence under `FINDING-008`. | Canonical TC are split, but coverage obligations still have duplicate overbroad mappings. | `round-2-findings.md` | medium | applied |
| `DEC-007` | 7 | `test-design` | `FINDING-007` | Accept split `Test-design Applicability Model` substitute. | Substitute is explicit and validator reports only current-scope `info`, not warning/error. | `round-2-findings.md` | medium | applied |
| `DEC-008` | 8 | `gap` | prompt required preserved gaps | Preserve `GAP-002`, `GAP-003`, `GAP-004`, `GAP-005`. | No new source/evidence resolves duplicate mechanism, backend effects, CDI setup or inverse branch. | matrix; final residual risk | high | applied |
| `DEC-009` | 9 | `routing` | max semantic rounds = 2 and open error | Route to `round-cap-reached`. | Round 2 is final semantic round and `FINDING-008` blocks semantic pass. | `cycle-state.yaml` | high | applied |
"""
    (OUTPUTS / "agent-decision-log.semantic-review-r2.md").write_text(content, encoding="utf-8")


def update_cycle_state() -> None:
    latest = [
        "test-cases/2-1-1-1-1-2-ui-employment-canary-v16-clean-runner-regression.md",
        f"work/test-design/{SCOPE}",
        f"work/review-cycles/{SCOPE}/outputs/round-2-findings.md",
        f"work/review-cycles/{SCOPE}/outputs/round-2-traceability-matrix.md",
        f"work/review-cycles/{SCOPE}/outputs/round-2-traceability-matrix.xlsx",
        f"work/review-cycles/{SCOPE}/outputs/writer-r2-response.md",
        f"work/review-cycles/{SCOPE}/outputs/reviewer-session-log.semantic-review-r2.md",
        f"work/review-cycles/{SCOPE}/outputs/agent-decision-log.semantic-review-r2.md",
        f"work/review-cycles/{SCOPE}/outputs/scoped-validator-profile.writer-r2.json",
        f"work/review-cycles/{SCOPE}/outputs/validator-report.writer-r2.full.json",
    ]
    latest_yaml = "\n".join(f"  - {item}" for item in latest)
    state = f"""cycle_id: ft-2-OF_16-ui-employment-canary-v16-clean-runner-regression
ft_slug: ft-2-OF_16
scope_slug: {SCOPE}
section_id: 2.1.1.1.1.2
current_stage: semantic-review-r2
stage_status: round-cap-reached
semantic_round: 2
max_semantic_rounds: 2
canonical_test_cases: test-cases/2-1-1-1-1-2-ui-employment-canary-v16-clean-runner-regression.md
test_design_dir: work/test-design/{SCOPE}
active_snapshot: none
active_transition_prompt: none
sessions: []
latest_artifacts:
{latest_yaml}
accepted_risks:
  - V16 is a targeted clean-run runner regression over selected employment UI rows, not a full employment-section coverage run.
  - GAP-002, GAP-003, GAP-004 and GAP-005 remain explicit residual source/setup gaps rather than invented executable behavior.
blocking_reasons:
  - FINDING-008 blocks semantic sign-off: numeric invalid TC use unsupported red-highlight/blocked-navigation oracle and drift from source/plan non-acceptance oracle.
blocking_findings:
  - FINDING-008
open_questions: []
"""
    (CYCLE / "cycle-state.yaml").write_text(state, encoding="utf-8")


def main() -> None:
    OUTPUTS.mkdir(parents=True, exist_ok=True)
    header, rows = build_matrix()
    xlsx_method = write_xlsx(header, rows)
    write_findings()
    write_session_log(xlsx_method)
    write_decision_log()
    update_cycle_state()
    print(f"wrote semantic r2 outputs at {datetime.now().isoformat(timespec='seconds')}; xlsx={xlsx_method}")


if __name__ == "__main__":
    main()
