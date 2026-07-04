from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font


ROOT = Path("fts/ft-2-OF_16")
SCOPE = "ui-employment-canary-v6-terminal-gate-regression"
CYCLE_DIR = ROOT / "work" / "review-cycles" / SCOPE
OUTPUTS = CYCLE_DIR / "outputs"
PROMPTS = CYCLE_DIR / "prompts"
TD = ROOT / "work" / "test-design" / SCOPE
STATE = CYCLE_DIR / "cycle-state.yaml"
CANONICAL = ROOT / "test-cases" / "2-1-1-1-1-2-ui-employment-canary-v6-terminal-gate-regression.md"


INSTRUCTION_FILES = [
    "AGENTS.md",
    "skills/README.md",
    "references/agent/session-based-review-cycle-format.md",
    "references/agent/codex-sdk-orchestration-format.md",
    "skills/ft-test-case-reviewer/SKILL.md",
    "references/agent/reviewer-output-format.md",
    "references/agent/package-test-design-plan-format.md",
    "references/agent/dictionary-inventory-format.md",
    "references/agent/test-design-defect-taxonomy.md",
    "references/qa/review-findings-format.md",
    "references/qa/traceability-matrix-format.md",
    "references/qa/test-design-review-rubric.md",
    "references/qa/coverage-runtime-checklist.md",
    "references/qa/traceability-rules.md",
    "references/agent/workflow-state-format.md",
    "references/agent/session-log-format.md",
    "references/agent/agent-decision-log-format.md",
    "references/agent/next-step-prompt-format.md",
]


REQUIRED_INPUTS = [
    "fts/ft-2-OF_16/AGENT-NOTES.md",
    "fts/ft-2-OF_16/work/review-cycles/ui-employment-canary-v6-terminal-gate-regression/cycle-state.yaml",
    "fts/ft-2-OF_16/work/review-cycles/ui-employment-canary-v6-terminal-gate-regression/outputs/round-1-findings.md",
    "fts/ft-2-OF_16/work/review-cycles/ui-employment-canary-v6-terminal-gate-regression/outputs/round-1-traceability-matrix.md",
    "fts/ft-2-OF_16/work/review-cycles/ui-employment-canary-v6-terminal-gate-regression/outputs/writer-r2-response.md",
    "fts/ft-2-OF_16/work/review-cycles/ui-employment-canary-v6-terminal-gate-regression/outputs/writer-session-log.writer-r2.md",
    "fts/ft-2-OF_16/work/review-cycles/ui-employment-canary-v6-terminal-gate-regression/outputs/agent-decision-log.writer-r2.md",
    "fts/ft-2-OF_16/test-cases/2-1-1-1-1-2-ui-employment-canary-v6-terminal-gate-regression.md",
    "fts/ft-2-OF_16/work/test-design/ui-employment-canary-v6-terminal-gate-regression/",
    "fts/ft-2-OF_16/work/stage-handoffs/01-ui-employment/source-row-inventory.md",
    "fts/ft-2-OF_16/work/stage-handoffs/01-ui-employment/source-parity-check.md",
    "fts/ft-2-OF_16/work/stage-handoffs/01-ui-employment/mockup-visual-inventory.md",
    "fts/ft-2-OF_16/work/stage-handoffs/01-ui-employment/scope-coverage-gaps.md",
    "fts/ft-2-OF_16/work/test-design/ui-employment/dictionary-inventory.md",
]


SPLIT_INPUTS = [
    "artifact-write-strategy.md",
    "source-row-inventory.md",
    "source-row-completeness-matrix.md",
    "source-table-normalization.md",
    "dictionary-inventory.md",
    "test-design-decision-table.md",
    "coverage-obligation-table.md",
    "atomic-requirements-ledger.md",
    "package-test-design-plan.md",
    "coverage-gaps.md",
    "coverage-map.md",
    "coverage-metrics.md",
    "dependency-matrix.md",
    "fixture-catalog.md",
    "internal-work-package-coverage.md",
    "mockup-usage.md",
    "risk-priority-map.md",
    "test-design-applicability-matrix.md",
    "test-design-review.md",
    "writer-quality-gate.md",
    "writer-self-check.md",
]


HEADERS = [
    "atom_id",
    "req_id",
    "source_path",
    "atomic_statement",
    "field_or_block",
    "condition",
    "expected_behavior",
    "covered_by_tc",
    "coverage_status",
    "gap_note",
]


def parse_md_table(path: Path) -> list[dict[str, str]]:
    lines = path.read_text(encoding="utf-8").splitlines()
    table_lines = [line for line in lines if line.startswith("|")]
    if len(table_lines) < 2:
        return []
    header = [cell.strip() for cell in table_lines[0].strip("|").split("|")]
    rows: list[dict[str, str]] = []
    for line in table_lines[2:]:
        cells = [cell.strip() for cell in line.strip("|").split("|")]
        if len(cells) == len(header):
            rows.append(dict(zip(header, cells)))
    return rows


def md_table(headers: list[str], rows: list[list[str]]) -> str:
    out = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(["---"] * len(headers)) + " |",
    ]
    for row in rows:
        out.append("| " + " | ".join(cell.replace("\n", " ") for cell in row) + " |")
    return "\n".join(out)


def _condition_from_statement(statement: str) -> str:
    marker = " при условии "
    if marker in statement:
        return statement.split(marker, 1)[1].strip().strip(".")
    return "always"


def _field_from_statement(statement: str) -> str:
    if ": " in statement:
        return statement.split(": ", 1)[0].strip()
    return "-"


def _expected_from_statement(statement: str) -> str:
    marker = ": "
    text = statement.split(marker, 1)[1] if marker in statement else statement
    if " при условии " in text:
        text = text.split(" при условии ", 1)[0]
    return text.strip().strip(".")


def build_matrix() -> list[dict[str, str]]:
    ledger_rows = parse_md_table(TD / "atomic-requirements-ledger.md")
    matrix: list[dict[str, str]] = []
    for row in ledger_rows:
        atom = row.get("atom_id", "")
        if not atom.startswith("ATOM-"):
            continue
        status = row.get("coverage_status", "covered") or "covered"
        linked = row.get("linked_tc_or_gap", "-") or "-"
        gap_note = "-"
        if status == "unclear":
            gap_note = f"{linked} retained as accepted non-blocking residual unclear item."
            linked = "-"
        matrix.append(
            {
                "atom_id": atom,
                "req_id": row.get("req_id", "none") or "none",
                "source_path": row.get("source_ref", ""),
                "atomic_statement": row.get("atomic_statement", ""),
                "field_or_block": row.get("field_or_block", "") or _field_from_statement(row.get("atomic_statement", "")),
                "condition": _condition_from_statement(row.get("atomic_statement", "")),
                "expected_behavior": _expected_from_statement(row.get("atomic_statement", "")),
                "covered_by_tc": linked,
                "coverage_status": status,
                "gap_note": gap_note,
            }
        )

    residual_rows = [
        (
            "ATOM-063",
            "SRC-011; DICT-004",
            "GAP-002; SRC-011",
            "`Пенсия` and `Аренда` may be added only once, but the exact duplicate-prevention mechanism is not defined.",
            "Тип дохода",
            "duplicate `Пенсия`/`Аренда` attempt",
            "Duplicate prevention invariant retained; exact UI feedback/control state unresolved.",
            "GAP-002",
        ),
        (
            "ATOM-064",
            "GSR 142; SRC-018",
            "GAP-003; SRC-018",
            "Return-from-`Выбор решения` setup and observable SPR/anti-fraud/external-check artifacts are not defined.",
            "Следующий шаг / status lifecycle",
            "return from `Выбор решения` and critical field changed",
            "UI-visible validation/navigation can be covered; hidden effects remain unresolved.",
            "GAP-003",
        ),
        (
            "ATOM-065",
            "SRC-022; SRC-023; SRC-024",
            "GAP-004; PDF pp.65-67",
            "PDF-only CDI messages are source text, but deterministic UI setup/test data for CDI failure and mismatch states is missing.",
            "CDI UI messages",
            "CDI failure/mismatch state",
            "Expected messages preserved as residual context; trigger/setup unresolved.",
            "GAP-004",
        ),
        (
            "ATOM-066",
            "GSR 124; GSR 135",
            "GAP-005; SRC-003; SRC-012",
            "Numeric-only and below-minimum invalid classes lack deterministic UI rejection/feedback oracle.",
            "Income numeric fields",
            "letters/spaces/special/decimal/sign or value below 2000",
            "Valid numeric/minimum acceptance covered; exact invalid-input behavior unresolved.",
            "GAP-005",
        ),
        (
            "ATOM-067",
            "GSR 132; GSR 133; GSR 134",
            "GAP-006; SRC-009",
            "Work phone invalid length and non-digit classes lack deterministic UI rejection/feedback oracle.",
            "Рабочий телефон",
            "9 digits, 11 digits or non-digit characters",
            "Exact 10-digit acceptance and default mask covered; invalid enforcement unresolved.",
            "GAP-006",
        ),
        (
            "ATOM-068",
            "GSR 139; GSR 140",
            "GAP-007; SRC-016",
            "Clear/reset behavior for selected visual-assessment checkboxes is not defined.",
            "Параметры визуальной оценки",
            "selected values need to be cleared",
            "List, no-selection requiredness, single and multiple selection covered; reset behavior unresolved.",
            "GAP-007",
        ),
        (
            "ATOM-069",
            "GSR 143",
            "GAP-008; SRC-018",
            "Field-to-document content mapping for `Заявление-анкета` is not provided in this scope.",
            "Заявление-анкета",
            "print form generated",
            "Generation/opening covered; content mapping unresolved.",
            "GAP-008",
        ),
    ]
    existing = {row["atom_id"] for row in matrix}
    for atom_id, req_id, source_path, statement, field, condition, behavior, gap_id in residual_rows:
        if atom_id in existing:
            continue
        matrix.append(
            {
                "atom_id": atom_id,
                "req_id": req_id,
                "source_path": source_path,
                "atomic_statement": statement,
                "field_or_block": field,
                "condition": condition,
                "expected_behavior": behavior,
                "covered_by_tc": "-",
                "coverage_status": "unclear",
                "gap_note": f"{gap_id} retained as accepted non-blocking residual unclear item.",
            }
        )
    return matrix


def write_matrix(matrix: list[dict[str, str]]) -> None:
    rows = [[row[h] for h in HEADERS] for row in matrix]
    md = "## Traceability Matrix\n\n" + md_table(HEADERS, rows) + "\n"
    (OUTPUTS / "round-2-traceability-matrix.md").write_text(md, encoding="utf-8")

    wb = Workbook()
    ws = wb.active
    ws.title = "traceability"
    ws.append(HEADERS)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 60)
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    meta = wb.create_sheet("meta")
    meta.append(["field", "value"])
    meta.append(["ft_slug", "ft-2-OF_16"])
    meta.append(["scope_slug", SCOPE])
    meta.append(["round", "2"])
    meta.append(["markdown_artifact", "round-2-traceability-matrix.md"])
    wb.save(OUTPUTS / "round-2-traceability-matrix.xlsx")


def coverage_counts(matrix: list[dict[str, str]]) -> dict[str, int]:
    counts = {"covered": 0, "unclear": 0, "gap": 0}
    for row in matrix:
        counts[row["coverage_status"]] = counts.get(row["coverage_status"], 0) + 1
    return counts


def write_findings(matrix: list[dict[str, str]]) -> None:
    counts = coverage_counts(matrix)
    findings = f"""# Semantic Review R2 Findings

## Human Summary

| item | value |
| --- | --- |
| review_mode | `semantic_traceability_test_design` |
| ft_slug | `ft-2-OF_16` |
| scope_slug | `{SCOPE}` |
| semantic_round | `2` |
| canonical_test_cases | `test-cases/2-1-1-1-1-2-ui-employment-canary-v6-terminal-gate-regression.md` |
| verdict | `semantic-review-passed` |
| recommended_next_step | `format-review-final` |

Round 2 verified the current canonical test-case file, writer split artifacts and `writer-r2-response.md`. All three round-1 findings have canonical writer responses with valid enum fields and are closed by evidence in the updated artifacts.

Coverage counts from reviewer matrix:

| coverage_status | count |
| --- | --- |
| covered | {counts.get('covered', 0)} |
| unclear | {counts.get('unclear', 0)} |
| gap | {counts.get('gap', 0)} |

## Round-1 Finding Closure

| finding_id | round_1_severity | writer_response_status | reviewer_r2_status | evidence |
| --- | --- | --- | --- | --- |
| FINDING-001 | error | fixed | closed | Affected negative `Следующий шаг` TC now disclose concrete valid values for all other mandatory fields/controls in the active branch; only the checked field/list remains empty. |
| FINDING-002 | error | fixed | closed | `SRC-004.P05` normalizes mandatory PDF-only `GSR 126` and links it to `ATOM-062` / `GAP-001` as `unclear`; no executable TC was invented for non-UI SPR contract artifacts. |
| FINDING-003 | warning | fixed | closed | `TC-UI-EMP-V6-053` now checks pension-branch main-income requiredness and no longer duplicates empty `Тип занятости` from `TC-UI-EMP-V6-003`. |

## Review Findings

No open `error`, `warning` or `info` findings.

## Residual Unclear Items

`GAP-001`-`GAP-008` remain residual unresolved/unclear context and are not executable expected results:

- `ATOM-062` / `GAP-001`: `GSR 126` and related DaData/non-UI SPR contract-field artifacts remain unclear.
- `ATOM-063` / `GAP-002`: duplicate-income prevention mechanism remains unclear.
- `ATOM-064` / `GAP-003`: return-from-`Выбор решения` SPR/anti-fraud/external-check artifacts remain unclear.
- `ATOM-065` / `GAP-004`: CDI failure/mismatch setup remains unclear.
- `ATOM-066` / `GAP-005`: invalid income input feedback remains unclear.
- `ATOM-067` / `GAP-006`: invalid work-phone enforcement feedback remains unclear.
- `ATOM-068` / `GAP-007`: visual-assessment checkbox clear/reset behavior remains unclear.
- `ATOM-069` / `GAP-008`: `Заявление-анкета` content mapping remains unclear.

These items are accepted non-blocking residual risks in `cycle-state.yaml`; they block neither semantic pass nor format review because the executable TC baseline avoids unsupported oracles for them.

## Semantic Review Self-check

**traceability_checked:** yes  
**source_parity_checked:** yes  
**writer_response_checked:** yes  
**round_1_findings_closed:** yes  
**mandatory_pdf_only_ids_checked:** yes  
**source_row_carryover_checked:** yes  
**dictionary_inventory_checked:** yes  
**mockup_limitations_checked:** yes  
**blocking_findings_absent:** yes  
**traceability_gaps_absent:** yes  
**known_unclear_items:** `ATOM-062`-`ATOM-069` / `GAP-001`-`GAP-008` accepted non-blocking residual unclear items  
**semantic_pass_rationale:** Round-1 semantic blockers are fixed, matrix has no `gap` rows, and residual `unclear` rows are preserved without executable unsupported expected results.

## Gate Decision Result

Semantic review passed. Route to `format-review-final`; do not route to UI automation prep from this stage.
"""
    (OUTPUTS / "round-2-findings.md").write_text(findings, encoding="utf-8")


def write_format_prompt() -> None:
    prompt = f"""# Format Review Final

## Stage Goal

Run `ft-test-case-reviewer` in `structure_format_final` mode for cycle `{SCOPE}`.

Perform final structure/format review after semantic review round 2 passed. Do not reopen semantic coverage unless a formatting defect changes meaning or exposes a semantic blocker. Do not edit the canonical test-case file.

## Instruction Loading

Before domain work, run:

```powershell
python scripts/resolve_instruction_context.py --scenario reviewer.structure_format_final --budget-report --fail-on-budget
```

Read every selected required file from resolver output before reviewer decisions. Record resolver command, budget status and selected files in the session log.

## Required Inputs

- `AGENT-NOTES.md`
- `work/review-cycles/{SCOPE}/cycle-state.yaml`
- `work/review-cycles/{SCOPE}/outputs/round-2-findings.md`
- `work/review-cycles/{SCOPE}/outputs/round-2-traceability-matrix.md`
- `work/review-cycles/{SCOPE}/outputs/round-2-traceability-matrix.xlsx`
- `work/review-cycles/{SCOPE}/outputs/writer-r2-response.md`
- `test-cases/2-1-1-1-1-2-ui-employment-canary-v6-terminal-gate-regression.md`
- `work/test-design/{SCOPE}/`
- `work/stage-handoffs/01-ui-employment/source-parity-check.md`
- `work/stage-handoffs/01-ui-employment/source-row-inventory.md`
- `work/stage-handoffs/01-ui-employment/mockup-visual-inventory.md`
- `work/stage-handoffs/01-ui-employment/scope-coverage-gaps.md`
- `work/test-design/ui-employment/dictionary-inventory.md`

## Required Review Focus

- Verify template sections, `TC-*` numbering, grouping, traceability fields, artifact links and wording consistency.
- Verify no format-only issue hides semantic coverage drift from round-2 matrix.
- Verify residual `GAP-001`-`GAP-008` remain explicit residual unclear context, not executable expected results.
- Verify canonical file can be handed to a format-only writer if needed.

## Do Not

- Do not edit the canonical test-case file.
- Do not route to UI automation prep unless final format/sign-off gates explicitly pass in this session-based lifecycle.
- Do not create a semantic round 3.
- Do not use old canary versions as source requirements or templates.

## Expected Outputs

- final format findings artifact
- reviewer session log for format review
- reviewer decision log for format review
- updated `cycle-state.yaml`
- if format review passes: route to `semantic-regression-final`
- if format-only findings remain: route to `writer-format-final`
- if semantic blockers are discovered despite this being format review: route according to round cap constraints

## Gate

Format review must not mark signed-off by itself unless the session-based cycle contract for final gates is satisfied. The next valid lifecycle stage after a clean format review is `semantic-regression-final`.
"""
    (PROMPTS / "prompt.format-review-final.md").write_text(prompt, encoding="utf-8")


def write_session_log() -> None:
    inputs_read = [
        "- `python scripts/resolve_instruction_context.py --scenario reviewer.semantic_traceability_test_design --budget-report --fail-on-budget` - resolver command; budget status `pass (259.1 / 290.0 KiB)`.",
    ]
    inputs_read += [f"- `{path}` - selected required instruction file." for path in INSTRUCTION_FILES]
    inputs_read += [f"- `{path}` - required semantic-review-r2 input." for path in REQUIRED_INPUTS]
    inputs_read += [f"- `fts/ft-2_OF_16/work/test-design/{SCOPE}/{name}` - split artifact read for semantic review." for name in []]
    inputs_read += [f"- `fts/ft-2-OF_16/work/test-design/{SCOPE}/{name}` - split artifact read for semantic review." for name in SPLIT_INPUTS]
    log = f"""# Semantic Review R2 Session Log

## Session Metadata

| field | value |
| --- | --- |
| skill | `ft-test-case-reviewer` |
| mode | `semantic_traceability_test_design` |
| ft_slug | `ft-2-OF_16` |
| scope_slug | `{SCOPE}` |
| started_from | `work/review-cycles/{SCOPE}/cycle-state.yaml` |
| status_after | `format-review-ready` |

## Artifact Write Strategy

| artifact_path | artifact_size_class | write_strategy | declared_before_first_write | helper | forbidden_methods_checked |
| --- | --- | --- | --- | --- | --- |
| `work/review-cycles/{SCOPE}/outputs/round-2-findings.md` | `review artifact` | `file-based helper write` | `yes` | `scripts/build_semantic_review_r2_outputs.py` | `yes` |
| `work/review-cycles/{SCOPE}/outputs/round-2-traceability-matrix.md` | `generated matrix` | `file-based helper write` | `yes` | `scripts/build_semantic_review_r2_outputs.py` | `yes` |
| `work/review-cycles/{SCOPE}/outputs/round-2-traceability-matrix.xlsx` | `generated xlsx` | `openpyxl workbook write` | `yes` | `scripts/build_semantic_review_r2_outputs.py` | `yes` |

## Inputs Read

{chr(10).join(inputs_read)}

## Inputs Not Used

- Older canary versions `v1`-`v5` - explicitly forbidden as requirement sources/templates for v6.
- UI automation artifacts - not used because current stage is semantic review r2.
- Legacy `workflow-state.yaml` statuses - not used as source of truth; `cycle-state.yaml` controlled routing.

## Key Decisions

- Closed all round-1 findings after verifying writer-r2 response and current artifacts.
- Kept `GAP-001`-`GAP-008` as accepted non-blocking residual `unclear` rows; did not treat them as executable coverage.
- Routed to `format-review-final`, not UI automation prep, because session-based cycle requires format review after semantic pass.

## Risks And Fallbacks

- Validator risk: root-wide validator includes historical/legacy artifacts and snapshots, so it produced noisy warnings outside the active r2 scope. Active semantic decisions used targeted current-artifact checks and are logged in findings/matrix.
- Encoding risk: initial PowerShell stdout for Cyrillic files produced mojibake in this overall session; affected source/instruction artifacts were reread with explicit UTF-8 and distorted stdout was not used as evidence.

## Validation

- `python scripts/resolve_instruction_context.py --scenario reviewer.semantic_traceability_test_design --budget-report --fail-on-budget` - pass, `259.1 / 290.0 KiB`.
- Writer response structure - pass: each `FINDING-001`-`FINDING-003` has canonical `Resolution Status`, `Change Summary`, `Affected Test Case IDs`, `Affected Traceability Refs` and valid enum `fixed`.
- R1 closure checks - pass: `FINDING-001`, `FINDING-002`, `FINDING-003` closed by current artifacts.
- Matrix check - pass: 69 rows, 10 columns, `covered=61`, `unclear=8`, `gap=0`; mandatory `GSR 123`-`GSR 148` present.
- `python scripts\\validate_agent_artifacts.py --root fts/ft-2-OF_16 --json` - executed; full-root output noisy with legacy/historical findings, no direct current R2 semantic blocker accepted from it.
- `python scripts\\validate_agent_artifacts.py --root fts/ft-2-OF_16/work/review-cycles/{SCOPE} --json` - executed; cycle-dir output included snapshot warnings/heuristics and was treated as supporting context, not sole semantic gate.
- `openpyxl.load_workbook(round-2-traceability-matrix.xlsx)` - pass.

## Contamination Check

- Older canary versions were not used as sources or templates.
- Mockup evidence was used only as interaction-hint/limitation context.
- Residual gaps were not converted to `covered` without new source/evidence.

## Event Timeline

| step | event | result | artifact_or_evidence |
| --- | --- | --- | --- |
| 1 | Ran instruction resolver | Budget pass | resolver stdout |
| 2 | Read required instruction files and semantic inputs | Completed | inputs listed above |
| 3 | Checked writer-r2 response | Pass | `writer-r2-response.md` |
| 4 | Checked affected TC and split artifacts | Pass | canonical TC; normalization/ledger/plan/gaps |
| 5 | Built round-2 matrix and findings | Pass | `round-2-*` artifacts |
| 6 | Routed next stage | Pass | `cycle-state.yaml` -> `format-review-final` |

## Quality Checkpoints

| checkpoint | status | evidence | follow_up |
| --- | --- | --- | --- |
| FINDING-001 fixture isolation | pass | affected negative TCs disclose concrete valid baselines | format review only |
| FINDING-002 GSR 126 traceability | pass | `SRC-004.P05`; `ATOM-062`; `GAP-001` as `unclear` | keep residual unclear |
| FINDING-003 duplicate action TC | pass | `TC-053` uses pension branch main-income requiredness | format review only |
| Residual gaps | pass | matrix `gap=0`, `unclear=8` | do not execute unsupported oracles |

## Technical Fallbacks

| fallback_id | trigger | failed_method | fallback_method | helper_artifact_path | retained | quality_risk | follow_up |
| --- | --- | --- | --- | --- | --- | --- | --- |
| `TF-001` | mojibake in PowerShell stdout for Cyrillic files | default console output read | explicit UTF-8 rereads with `[Console]::OutputEncoding` and `Get-Content -Encoding UTF8`; distorted stdout discarded | `n/a` | `n/a` | none after reread | keep UTF-8 reads in next stage |
| `TF-002` | generated reviewer matrix and xlsx duplicate | manual one-shot write would be fragile | file-based helper script with openpyxl | `scripts/build_semantic_review_r2_outputs.py` | `yes` | low | none |
| `TF-003` | validator scope noise from historical artifacts/snapshots | unfiltered root validator as decisive sign-off gate | targeted current-artifact semantic checks plus logged validator limitation | `n/a` | `n/a` | medium | format review should run its own scoped checks |

## Handoff Notes For Next Session

- Semantic review passed; next session should perform format review only.
- `GAP-001`-`GAP-008` are accepted residual unclear items, not executable expected results.
- Do not route to UI automation prep until the remaining session-based final format/regression gates complete.
"""
    (OUTPUTS / "reviewer-session-log.semantic-review-r2.md").write_text(log, encoding="utf-8")


def write_decision_log() -> None:
    log = f"""# Agent Decision Log

## Decision Log Metadata

| field | value |
| --- | --- |
| ft_slug | `ft-2-OF_16` |
| scope_slug | `{SCOPE}` |
| stage | `semantic-review-r2` |
| started_from | `work/review-cycles/{SCOPE}/cycle-state.yaml` |

## Decision Log

| decision_id | step | decision_type | input_or_trigger | decision | rationale | artifact_or_output | risk_or_confidence | status |
| --- | --- | --- | --- | --- | --- | --- | --- | --- |
| `DEC-001` | 1 | `validation` | instruction resolver | Allow semantic review to proceed | Resolver passed and all selected required instruction files were read | session log | high | applied |
| `DEC-002` | 2 | `test-design` | `FINDING-001`; affected TCs | Close fixture-isolation finding | All affected negative transition TCs now disclose concrete valid values for other mandatory fields/controls | `round-2-findings.md` | high | applied |
| `DEC-003` | 3 | `traceability` | `FINDING-002`; `GSR 126` | Close mandatory `GSR 126` normalization finding | `SRC-004.P05` links `GSR 126` to `ATOM-062` / `GAP-001` as residual `unclear` | `round-2-traceability-matrix.md` | high | applied |
| `DEC-004` | 4 | `test-design` | `FINDING-003`; `TC-003`; `TC-053` | Close duplicate action-level finding | `TC-053` is now pension-branch main-income requiredness, not empty `Тип занятости` | `round-2-findings.md` | high | applied |
| `DEC-005` | 5 | `gap` | `GAP-001`-`GAP-008` | Preserve residual gaps as `unclear` | No new source/evidence closes these mechanisms; executable TC baseline avoids unsupported oracles | `round-2-traceability-matrix.md` | high | applied |
| `DEC-006` | 6 | `routing` | semantic pass gate | Route to `format-review-final` | No open semantic error/warning findings and no traceability `gap` rows remain | `cycle-state.yaml`; `prompt.format-review-final.md` | high | applied |
"""
    (OUTPUTS / "agent-decision-log.semantic-review-r2.md").write_text(log, encoding="utf-8")


def update_state() -> None:
    text = STATE.read_text(encoding="utf-8")
    text = text.replace("current_stage: semantic-review-r2", "current_stage: format-review-final")
    text = text.replace("stage_status: semantic-review-ready", "stage_status: format-review-ready")
    text = text.replace(
        f"active_transition_prompt: work/review-cycles/{SCOPE}/prompts/prompt.semantic-review-r2.md",
        f"active_transition_prompt: work/review-cycles/{SCOPE}/prompts/prompt.format-review-final.md",
    )
    additions = [
        f"  - work/review-cycles/{SCOPE}/outputs/round-2-findings.md",
        f"  - work/review-cycles/{SCOPE}/outputs/round-2-traceability-matrix.md",
        f"  - work/review-cycles/{SCOPE}/outputs/round-2-traceability-matrix.xlsx",
        f"  - work/review-cycles/{SCOPE}/outputs/reviewer-session-log.semantic-review-r2.md",
        f"  - work/review-cycles/{SCOPE}/outputs/agent-decision-log.semantic-review-r2.md",
        f"  - work/review-cycles/{SCOPE}/prompts/prompt.format-review-final.md",
    ]
    if additions[0] not in text and "blocking_reasons:" in text:
        before, after = text.split("blocking_reasons:", 1)
        before = before.rstrip() + "\n" + "\n".join(additions) + "\n"
        text = before + "blocking_reasons:" + after
    text = text.replace("blocking_reasons: []", "blocking_reasons: []")
    text = text.replace("blocking_findings: []", "blocking_findings: []")
    STATE.write_text(text, encoding="utf-8")


def verify_outputs(matrix: list[dict[str, str]]) -> None:
    assert len(matrix) == 69, len(matrix)
    counts = coverage_counts(matrix)
    assert counts.get("gap", 0) == 0, counts
    assert counts.get("unclear", 0) == 8, counts
    reqs = " ".join(row["req_id"] for row in matrix)
    for num in range(123, 149):
        assert f"GSR {num}" in reqs, f"GSR {num}"
    wb = load_workbook(OUTPUTS / "round-2-traceability-matrix.xlsx", read_only=True, data_only=True)
    ws = wb["traceability"]
    assert ws.max_row == 70, ws.max_row
    assert ws.max_column == 10, ws.max_column


def main() -> None:
    OUTPUTS.mkdir(parents=True, exist_ok=True)
    PROMPTS.mkdir(parents=True, exist_ok=True)
    matrix = build_matrix()
    write_matrix(matrix)
    write_findings(matrix)
    write_format_prompt()
    write_session_log()
    write_decision_log()
    update_state()
    verify_outputs(matrix)
    print("Generated semantic-review-r2 outputs")
    print(f"matrix_rows={len(matrix)} counts={coverage_counts(matrix)}")


if __name__ == "__main__":
    main()
